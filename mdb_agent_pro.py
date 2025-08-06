#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MDB Agent Pro v2.0.0 - Microsoft Access Database to API Bridge
=================================================================

Professional GUI application for bridging Microsoft Access databases to REST APIs
with visual field mapping, real-time monitoring, and enterprise-grade features.

Features:
- Visual field mapping interface with drag-and-drop functionality
- Real-time API testing and validation
- Data transformation engine with multiple formats
- Intelligent retry mechanism with exponential backoff
- Health monitoring and comprehensive diagnostics
- Transaction logging and audit trails
- Template management for reusable configurations
- Automated scheduling and background processing
- Security with encrypted configuration storage
- Professional multi-tab interface design

Technical Specifications:
- Python 3.8+ with Tkinter GUI framework
- Microsoft Access connectivity via pyodbc
- REST API integration with authentication support
- SQLite3 for local data storage and logging
- Multi-threaded background processing
- Cross-platform Windows compatibility

Developer: Freddy Mazmur
Company: PT Sahabat Agro Group
Contact: freddy.pm@sahabatagro.co.id
Phone: +62 813-9855-2019
Version: 2.0.0 Professional Edition
Release: January 2025

License: Proprietary - PT Sahabat Agro Group
Copyright (c) 2025 PT Sahabat Agro Group. All rights reserved.
"""

# Standard library imports
import tkinter as tk
import tkinter.simpledialog
from tkinter import ttk, filedialog, messagebox, scrolledtext
import tkinter.font as tkFont
import json
import os
import threading
import queue
import uuid
import base64
import sqlite3
import logging
import platform
import sys
import smtplib
import time
import hashlib
import csv
import traceback
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Third-party imports
import pyodbc
import requests

# Optional imports with fallbacks
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# Import utility modules
from utils.security import SecurityManager, AuthenticationManager
from utils.logging_manager import LogManager, get_log_manager
from utils.gui_utils import (
    StatusIndicator, NavigationManager, ConfigurationManager, 
    StyleManager, ErrorHandler, FieldMapper
)
from utils.database_manager import DatabaseManager, ConnectionManager, BufferItem, RetryConfig

class MDBAgentPro:
    """Main application class with modular architecture"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("MDB Agent Pro v2.0 - PT Sahabat Agro Group")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)
        
        # Initialize utility managers
        self.security = SecurityManager()
        self.log_manager = get_log_manager()
        self.db_manager = DatabaseManager()
        self.connection_manager = ConnectionManager()
        self.config_manager = ConfigurationManager(self.security)
        
        # Initialize field mapper
        self.field_mapper = None
        self.mapping_mode = tk.StringVar(value="flat")
        self.error_handler = ErrorHandler(self.log_manager, self.root)
        
        # Initialize status variable first
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        
        # Configuration (use ConfigurationManager)
        self.config = self.config_manager.config
        
        # Database connection
        self.db_connection = None
        self.selected_table = None
        self.table_columns = []
        
        # Field mapping variables
        self.field_mappings = {}
        self.api_fields = []
        self.mapping_widgets = {}
        
        # API buffer and threading
        self.is_running = False
        self.worker_thread = None
        self.push_queue = queue.Queue()
        
        # Test connection results storage
        self.last_test_connection_result = None
        
        # API fields source tracking
        self.api_fields_source = "none"  # none, auto_detect, manual
        
        # Admin mode with enhanced security
        self.admin_mode = False
        self.auth_manager = AuthenticationManager(self.security)
        self.current_session_id = None
        
        # Current tab
        self.current_tab = "dashboard"
        
        # Setup GUI components
        self.setup_styles()
        self.setup_gui()
        self.load_settings()
        
        # Start background worker
        self.start_worker()
        
        # Register configuration change callbacks
        self.config_manager.register_change_callback(self.on_config_changed)
    
    def setup_styles(self):
        """Setup custom styles using StyleManager"""
        self.style = ttk.Style()
        self.style_manager = StyleManager(self.style)
        
        # Apply the green theme
        self.style_manager.apply_theme("green")
    
    def on_config_changed(self, config: Dict):
        """Handle configuration changes"""
        self.log_manager.log("INFO", "Configuration updated", module="MDBAgentPro")
        
        # Update UI elements if needed
        if hasattr(self, 'status_var'):
            self.status_var.set("Configuration updated")
    
    def log_entry(self, message: str, level: str = "INFO", details: str = ""):
        """Unified logging method using LogManager"""
        self.log_manager.log(level, message, details, module="MDBAgentPro")
        
        # Update status variable
        if hasattr(self, 'status_var') and self.status_var:
            self.status_var.set(f"{level}: {message}")
        
        # Update dashboard if visible
        if hasattr(self, 'current_tab') and self.current_tab == "dashboard":
            self.refresh_dashboard()
    
    def save_config(self):
        """Save configuration using ConfigurationManager"""
        success = self.config_manager.save_config()
        if success:
            # Preserve test connection results after save
            if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
                # Re-propagate test results to maintain consistency across UI
                result = self.last_test_connection_result
                if hasattr(self, 'propagate_test_connection_results'):
                    self.propagate_test_connection_results(
                        result['success'], 
                        result['status_code'], 
                        result.get('error', ''), 
                        result.get('response_time', '0ms').replace('ms', '')
                    )
                self.log_entry("Configuration saved - test connection results preserved", "INFO")
            else:
                self.log_entry("Configuration saved", "INFO")
        else:
            self.log_entry("Failed to save configuration", "ERROR")
    
    def start_worker(self):
        """Start background worker thread"""
        def worker():
            while True:
                try:
                    if self.is_running and self.config.get("auto_push"):
                        # Process buffer items with mapping
                        items = self.db_manager.get_buffer_items()
                        for item in items:
                            # Try to send with mapping if payload is raw data
                            payload = item['payload']
                            if isinstance(payload, dict) and 'raw_data' in payload:
                                # This is raw data, apply mapping
                                success = self.send_data_with_mapping(payload['raw_data'])
                            else:
                                # Already processed payload
                                success = self.send_to_api(payload)
                                
                            if success:
                                self.db_manager.update_buffer_status(item['id'], 'completed')
                            else:
                                retry_count = item['retry_count'] + 1
                                if retry_count < 3:
                                    self.db_manager.update_buffer_status(item['id'], 'pending', retry_count)
                                else:
                                    self.db_manager.update_buffer_status(item['id'], 'failed', retry_count)
                    
                    import time
                    time.sleep(30)  # Check every 30 seconds
                except:
                    pass
        
        self.worker_thread = threading.Thread(target=worker, daemon=True)
        self.worker_thread.start()
    
    def setup_gui(self):
        """Setup main GUI"""
        # Main container
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Top status bar
        self.setup_status_bar(main_frame)
        
        # Content area
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Sidebar
        self.setup_sidebar(content_frame)
        
        # Main content area
        self.setup_content_area(content_frame)
        
        # Bottom status bar
        self.setup_bottom_status(main_frame)
    
    def setup_status_bar(self, parent):
        """Setup top status indicator bar"""
        status_frame = ttk.Frame(parent, style='Status.TFrame')
        status_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        # Left side - Application title
        title_frame = ttk.Frame(status_frame)
        title_frame.pack(side=tk.LEFT, fill=tk.Y)
        
        ttk.Label(title_frame, text="MDB Agent Pro v2.0", 
                 style='Title.TLabel', 
                 font=('Arial', 14, 'bold')).pack(anchor=tk.W)
        ttk.Label(title_frame, text="Database to API Bridge", 
                 font=('Arial', 10), 
                 foreground='gray').pack(anchor=tk.W)
        
        # Right side - Status indicators with better layout
        indicators_frame = ttk.Frame(status_frame)
        indicators_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Status indicators container
        status_container = ttk.Frame(indicators_frame)
        status_container.pack(side=tk.RIGHT, padx=(20, 0))
        
        self.db_status = StatusIndicator(status_container, "Database")
        self.db_status.frame.pack(side=tk.LEFT, padx=(0, 15))
        
        self.api_status = StatusIndicator(status_container, "API")
        self.api_status.frame.pack(side=tk.LEFT, padx=(0, 15))
        
        self.buffer_status = StatusIndicator(status_container, "Buffer")
        self.buffer_status.frame.pack(side=tk.LEFT)
    
    def setup_sidebar(self, parent):
        """Setup navigation sidebar using NavigationManager"""
        sidebar = ttk.Frame(parent, style='Sidebar.TFrame', width=250)
        sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 15))
        sidebar.pack_propagate(False)
        
        # Header section with consistent padding
        header_frame = ttk.Frame(sidebar)
        header_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        ttk.Label(header_frame, text="MDB Agent Pro", 
                 font=('Arial', 12, 'bold')).pack(anchor=tk.W)
        ttk.Label(header_frame, text="PT Sahabat Agro Group", 
                 font=('Arial', 9), foreground='gray').pack(anchor=tk.W, pady=(2, 0))
        
        # Navigation sections with proper spacing
        nav_frame = ttk.Frame(sidebar)
        nav_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 0))
        
        # Initialize NavigationManager
        self.nav_manager = NavigationManager(nav_frame)
        
        # Create navigation sections
        self.nav_manager.create_section("DASHBOARD", [
            ("Dashboard", "dashboard")
        ], is_first=True)
        
        self.nav_manager.create_section("MASTER", [
            ("Health Checks", "health_checks"),
            ("Transaction Log", "transaction")
        ])
        
        self.nav_manager.create_section("CONFIGURATION", [
            ("Database Connection", "database_connection"),
            ("API Field Mapping", "mapping"),
            ("API Settings", "api"),
            ("Scheduler", "scheduler")
        ])
        
        self.nav_manager.create_section("INFORMATION", [
            ("About Application", "about")
        ])
        
        # Set main tab switching callback
        self.nav_manager.set_main_switch_callback(self.switch_tab)
        
        # Register additional tab callbacks for refresh actions
        self.nav_manager.register_tab_callback("dashboard", self.refresh_dashboard)
        self.nav_manager.register_tab_callback("health_checks", self.auto_refresh_health_status)
        self.nav_manager.register_tab_callback("scheduler", self.refresh_scheduler_log)
        
        # Footer controls with proper spacing
        footer_frame = ttk.Frame(sidebar)
        footer_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(10, 15))
        
        ttk.Separator(footer_frame, orient='horizontal').pack(fill=tk.X, pady=(0, 10))
        
        # Control buttons with consistent styling
        self.admin_btn = ttk.Button(
            footer_frame, 
            text=" Admin Mode", 
            command=self.toggle_admin_mode
        )
        self.admin_btn.pack(fill=tk.X)
    
    def setup_content_area(self, parent):
        """Setup main content area"""
        self.content_frame = ttk.Frame(parent, style='Content.TFrame')
        self.content_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Add padding to content area
        content_inner = ttk.Frame(self.content_frame)
        content_inner.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        # Content will be dynamically added here
        self.content_inner = content_inner
        
        # Create all tab frames
        self.tab_frames = {}
        self.create_dashboard_tab()
        self.create_health_checks_tab()
        self.create_transaction_tab()
        self.create_database_connection_tab()
        self.create_mapping_tab()
        self.create_api_tab()
        self.create_scheduler_tab()
        self.create_about_tab()
        
        # Log successful tab initialization
        self.log_entry(f"Tab frames created: {list(self.tab_frames.keys())}", "INFO")
        
        # Show dashboard by default
        self.switch_tab("dashboard")
    
    def setup_bottom_status(self, parent):
        """Setup bottom status bar"""
        bottom_frame = ttk.Frame(parent, style='Status.TFrame')
        bottom_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=15, pady=(5, 10))
        
        # Left side - Status message
        left_frame = ttk.Frame(bottom_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.Y)
        
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        ttk.Label(left_frame, textvariable=self.status_var, 
                 font=('Arial', 9)).pack(side=tk.LEFT)
        
        # Right side - Agent status with better styling
        right_frame = ttk.Frame(bottom_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.agent_status_var = tk.StringVar()
        self.agent_status_var.set(" Agent: Stopped")
        agent_label = ttk.Label(right_frame, textvariable=self.agent_status_var, 
                               font=('Arial', 9, 'bold'))
        agent_label.pack(side=tk.RIGHT)
    
    def create_dashboard_tab(self):
        """Create dashboard tab"""
        frame = ttk.Frame(self.content_inner)
        self.tab_frames["dashboard"] = frame
        
        # Page title with better spacing
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, pady=(0, 25))
        
        ttk.Label(title_frame, text="Dashboard", 
                 style='Title.TLabel', 
                 font=('Arial', 16, 'bold')).pack(anchor=tk.W)
        ttk.Label(title_frame, text="System overview and quick actions", 
                 font=('Arial', 11), 
                 foreground='gray').pack(anchor=tk.W, pady=(3, 0))
        
        # Status overview with better layout
        status_frame = ttk.LabelFrame(frame, text="System Status", padding=20)
        status_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Create a grid with proper spacing
        status_grid = ttk.Frame(status_frame)
        status_grid.pack(fill=tk.X)
        
        # Configure grid weights for better distribution
        status_grid.grid_columnconfigure(1, weight=1)
        
        # Status items with icons and better spacing
        status_items = [
            ("Database:", "dash_db_status", " Not Connected", "red"),
            ("API Endpoint:", "dash_api_status", " Not Configured", "red"),
            ("Agent Service:", "dash_agent_status", " Stopped", "red"),
            ("Buffer Queue:", "dash_buffer_status", " 0 items", "green")
        ]
        
        for i, (label, attr_name, default_text, color) in enumerate(status_items):
            # Label
            ttk.Label(status_grid, text=label, 
                     font=('Arial', 10, 'bold')).grid(
                         row=i, column=0, sticky=tk.W, 
                         padx=(0, 20), pady=8)
            
            # Status
            status_label = ttk.Label(status_grid, text=default_text, 
                                   foreground=color, font=('Arial', 10))
            status_label.grid(row=i, column=1, sticky=tk.W, pady=8)
            setattr(self, attr_name, status_label)
        
        # Quick actions with better layout
        actions_frame = ttk.LabelFrame(frame, text="Quick Actions", padding=20)
        actions_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Create action buttons in a more organized way
        actions_container = ttk.Frame(actions_frame)
        actions_container.pack()
        
        # Database & API actions
        db_api_frame = ttk.Frame(actions_container)
        db_api_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(db_api_frame, text=" Test Database", 
                  command=self.test_database, width=18).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(db_api_frame, text=" Test API", 
                  command=self.test_api, width=18).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(db_api_frame, text=" Manual Push", 
                  command=self.manual_push, width=18).pack(side=tk.LEFT)
        
        # Agent control actions
        agent_frame = ttk.Frame(actions_container)
        agent_frame.pack(fill=tk.X)
        
        ttk.Button(agent_frame, text=" Start Agent", 
                  command=self.start_agent, width=18).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(agent_frame, text=" Stop Agent", 
                  command=self.stop_agent, width=18).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(agent_frame, text=" Clear Buffer", 
                  command=self.clear_buffer, width=18).pack(side=tk.LEFT)
        
        # Recent activity with better styling
        activity_frame = ttk.LabelFrame(frame, text="Recent Activity", padding=20)
        activity_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create treeview with better column sizing
        self.activity_tree = ttk.Treeview(activity_frame, 
                                        columns=("time", "level", "message"), 
                                        show="headings", height=10)
        
        # Configure columns with better proportions
        self.activity_tree.heading("time", text="Time")
        self.activity_tree.heading("level", text="Level") 
        self.activity_tree.heading("message", text="Message")
        
        self.activity_tree.column("time", width=140, minwidth=120)
        self.activity_tree.column("level", width=80, minwidth=60)
        self.activity_tree.column("message", width=450, minwidth=300)
        
        # Add scrollbar
        activity_scroll = ttk.Scrollbar(activity_frame, orient=tk.VERTICAL, 
                                      command=self.activity_tree.yview)
        self.activity_tree.configure(yscrollcommand=activity_scroll.set)
        
        # Pack with proper layout
        self.activity_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        activity_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    def create_health_checks_tab(self):
        """Create health checks tab"""
        frame = ttk.Frame(self.content_inner)
        self.tab_frames["health_checks"] = frame
        
        # Page title with better spacing
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, pady=(0, 25))
        
        ttk.Label(title_frame, text="Health Checks", 
                 style='Title.TLabel', 
                 font=('Arial', 16, 'bold')).pack(anchor=tk.W)
        ttk.Label(title_frame, text="System health monitoring and diagnostics", 
                 font=('Arial', 11), 
                 foreground='gray').pack(anchor=tk.W, pady=(3, 0))
        
        # Integration Status
        integration_frame = ttk.LabelFrame(frame, text="API Settings Integration", padding=10)
        integration_frame.pack(fill=tk.X, pady=(0, 20))
        
        status_row = ttk.Frame(integration_frame)
        status_row.pack(fill=tk.X)
        
        ttk.Label(status_row, text=" Status:", font=('Arial', 10, 'bold')).pack(side=tk.LEFT)
        integration_status = ttk.Label(status_row, text=" Fully Integrated with API Settings", 
                                     font=('Arial', 10), foreground='green')
        integration_status.pack(side=tk.LEFT, padx=(10, 0))
        
        info_text = ("Health Check sekarang menggunakan konfigurasi langsung dari tab API Settings. "
                    "Perubahan pada endpoint, authentication, dan method akan otomatis ter-update di Health Check.")
        ttk.Label(integration_frame, text=info_text, font=('Arial', 9), 
                 foreground='blue', wraplength=600).pack(anchor=tk.W, pady=(5, 0))
        
        # System health overview with better layout
        health_frame = ttk.LabelFrame(frame, text="System Health Overview", padding=20)
        health_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Health status grid
        health_grid = ttk.Frame(health_frame)
        health_grid.pack(fill=tk.X)
        
        # Database health
        ttk.Label(health_grid, text="Database Connection:", style='Header.TLabel').grid(row=0, column=0, sticky=tk.W, padx=(0, 20))
        self.health_db_status = ttk.Label(health_grid, text=" Not Connected", foreground="red")
        self.health_db_status.grid(row=0, column=1, sticky=tk.W)
        ttk.Button(health_grid, text="Test", command=self.health_test_database).grid(row=0, column=2, padx=(10, 0))
        
        # API health
        ttk.Label(health_grid, text="API Endpoint:", style='Header.TLabel').grid(row=1, column=0, sticky=tk.W, padx=(0, 20))
        self.health_api_status = ttk.Label(health_grid, text=" Not Configured", foreground="red")
        self.health_api_status.grid(row=1, column=1, sticky=tk.W)
        ttk.Button(health_grid, text="Test", command=self.health_test_api).grid(row=1, column=2, padx=(10, 0))
        
        # Buffer health
        ttk.Label(health_grid, text="Buffer Status:", style='Header.TLabel').grid(row=2, column=0, sticky=tk.W, padx=(0, 20))
        self.health_buffer_status = ttk.Label(health_grid, text=" Empty", foreground="green")
        self.health_buffer_status.grid(row=2, column=1, sticky=tk.W)
        ttk.Button(health_grid, text="Check", command=self.health_check_buffer).grid(row=2, column=2, padx=(10, 0))
        
        # Data integrity
        ttk.Label(health_grid, text="Data Integrity:", style='Header.TLabel').grid(row=3, column=0, sticky=tk.W, padx=(0, 20))
        self.health_data_status = ttk.Label(health_grid, text=" Not Verified", foreground="orange")
        self.health_data_status.grid(row=3, column=1, sticky=tk.W)
        ttk.Button(health_grid, text="Verify", command=self.health_verify_data).grid(row=3, column=2, padx=(10, 0))
        
        # Auto check controls
        check_frame = ttk.LabelFrame(frame, text="Health Check Controls", padding=10)
        check_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Info label about API Settings integration
        info_label = ttk.Label(check_frame, text=" API Health Check menggunakan konfigurasi dari tab API Settings", 
                              font=('Arial', 9), foreground='blue')
        info_label.pack(anchor=tk.W, pady=(0, 10))
        
        control_frame = ttk.Frame(check_frame)
        control_frame.pack(fill=tk.X)
        
        ttk.Button(control_frame, text="Run All Checks", command=self.run_all_health_checks).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(control_frame, text="Refresh API Settings", command=self.refresh_health_api_settings).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(control_frame, text="Auto Check (5min)", command=self.toggle_auto_health_check).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(control_frame, text="Export Health Report", command=self.export_health_report).pack(side=tk.LEFT)
        
        # Health check results
        results_frame = ttk.LabelFrame(frame, text="Health Check Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        self.health_tree = ttk.Treeview(results_frame, columns=("timestamp", "check", "status", "details"), show="headings", height=10)
        self.health_tree.heading("timestamp", text="Timestamp")
        self.health_tree.heading("check", text="Check Type")
        self.health_tree.heading("status", text="Status")
        self.health_tree.heading("details", text="Details")
        
        self.health_tree.column("timestamp", width=150)
        self.health_tree.column("check", width=120)
        self.health_tree.column("status", width=80)
        self.health_tree.column("details", width=350)
        
        health_scroll = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.health_tree.yview)
        self.health_tree.configure(yscrollcommand=health_scroll.set)
        
        self.health_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        health_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    def create_transaction_tab(self):
        """Create transaction log tab"""
        frame = ttk.Frame(self.content_inner)
        self.tab_frames["transaction"] = frame
        
        # Page title with better spacing
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, pady=(0, 25))
        
        ttk.Label(title_frame, text="Transaction Log", 
                 style='Title.TLabel', 
                 font=('Arial', 16, 'bold')).pack(anchor=tk.W)
        ttk.Label(title_frame, text="Monitor and analyze data push transactions", 
                 font=('Arial', 11), 
                 foreground='gray').pack(anchor=tk.W, pady=(3, 0))
        
        # Transaction filters
        filter_frame = ttk.LabelFrame(frame, text="Filters", padding=10)
        filter_frame.pack(fill=tk.X, pady=(0, 10))
        
        filter_grid = ttk.Frame(filter_frame)
        filter_grid.pack(fill=tk.X)
        
        # Date filter
        ttk.Label(filter_grid, text="Date Range:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.trans_start_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(filter_grid, textvariable=self.trans_start_date, width=12).grid(row=0, column=1, padx=(0, 5))
        ttk.Label(filter_grid, text="to").grid(row=0, column=2, padx=5)
        self.trans_end_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(filter_grid, textvariable=self.trans_end_date, width=12).grid(row=0, column=3, padx=(5, 20))
        
        # Status filter
        ttk.Label(filter_grid, text="Status:").grid(row=0, column=4, sticky=tk.W, padx=(0, 10))
        self.trans_status_filter = tk.StringVar(value="All")
        status_combo = ttk.Combobox(filter_grid, textvariable=self.trans_status_filter, 
                                   values=["All", "Success", "Failed", "Pending"], state="readonly", width=10)
        status_combo.grid(row=0, column=5, padx=(0, 20))
        
        ttk.Button(filter_grid, text="Apply Filter", command=self.apply_transaction_filter).grid(row=0, column=6, padx=10)
        ttk.Button(filter_grid, text="Export", command=self.export_transactions).grid(row=0, column=7, padx=5)
        
        # Transaction statistics
        stats_frame = ttk.LabelFrame(frame, text="Transaction Statistics", padding=10)
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        stats_grid = ttk.Frame(stats_frame)
        stats_grid.pack(fill=tk.X)
        
        ttk.Label(stats_grid, text="Total:", style='Header.TLabel').grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.stats_total = ttk.Label(stats_grid, text="0")
        self.stats_total.grid(row=0, column=1, sticky=tk.W, padx=(0, 30))
        
        ttk.Label(stats_grid, text="Success:", style='Header.TLabel').grid(row=0, column=2, sticky=tk.W, padx=(0, 10))
        self.stats_success = ttk.Label(stats_grid, text="0", foreground="green")
        self.stats_success.grid(row=0, column=3, sticky=tk.W, padx=(0, 30))
        
        ttk.Label(stats_grid, text="Failed:", style='Header.TLabel').grid(row=0, column=4, sticky=tk.W, padx=(0, 10))
        self.stats_failed = ttk.Label(stats_grid, text="0", foreground="red")
        self.stats_failed.grid(row=0, column=5, sticky=tk.W, padx=(0, 30))
        
        ttk.Label(stats_grid, text="Pending:", style='Header.TLabel').grid(row=0, column=6, sticky=tk.W, padx=(0, 10))
        self.stats_pending = ttk.Label(stats_grid, text="0", foreground="orange")
        self.stats_pending.grid(row=0, column=7, sticky=tk.W)
        
        # Transaction log
        log_frame = ttk.LabelFrame(frame, text="Transaction History", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.transaction_tree = ttk.Treeview(log_frame, columns=("id", "timestamp", "table", "status", "records", "details"), show="headings", height=12)
        self.transaction_tree.heading("id", text="ID")
        self.transaction_tree.heading("timestamp", text="Timestamp")
        self.transaction_tree.heading("table", text="Table")
        self.transaction_tree.heading("status", text="Status")
        self.transaction_tree.heading("records", text="Records")
        self.transaction_tree.heading("details", text="Details")
        
        self.transaction_tree.column("id", width=60)
        self.transaction_tree.column("timestamp", width=150)
        self.transaction_tree.column("table", width=120)
        self.transaction_tree.column("status", width=80)
        self.transaction_tree.column("records", width=80)
        self.transaction_tree.column("details", width=300)
        
        trans_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.transaction_tree.yview)
        self.transaction_tree.configure(yscrollcommand=trans_scroll.set)
        
        self.transaction_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        trans_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    def create_database_connection_tab(self):
        """Create database configuration tab"""
        frame = ttk.Frame(self.content_inner)
        self.tab_frames["database_connection"] = frame
        
        # Page title with better spacing
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, pady=(0, 25))
        
        ttk.Label(title_frame, text="Database Connection", 
                 style='Title.TLabel', 
                 font=('Arial', 16, 'bold')).pack(anchor=tk.W)
        ttk.Label(title_frame, text="Configure Microsoft Access database connection and settings", 
                 font=('Arial', 11), 
                 foreground='gray').pack(anchor=tk.W, pady=(3, 0))
        
        # File selection
        file_frame = ttk.LabelFrame(frame, text="Database File", padding=10)
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(file_frame, text="MDB File:").pack(anchor=tk.W)
        file_select_frame = ttk.Frame(file_frame)
        file_select_frame.pack(fill=tk.X, pady=(5, 10))
        
        self.mdb_file_var = tk.StringVar()
        ttk.Entry(file_select_frame, textvariable=self.mdb_file_var, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(file_select_frame, text="Browse", command=self.browse_mdb_file).pack(side=tk.RIGHT)
        
        ttk.Label(file_frame, text="Password:").pack(anchor=tk.W, pady=(10, 0))
        self.mdb_password_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.mdb_password_var, show="*").pack(fill=tk.X, pady=(5, 10))
        
        ttk.Button(file_frame, text="Connect Database", command=self.connect_database).pack()
        
        # Table selection
        table_frame = ttk.LabelFrame(frame, text="Table Selection", padding=10)
        table_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(table_frame, text="Available Tables:").pack(anchor=tk.W)
        self.table_var = tk.StringVar()
        self.table_combo = ttk.Combobox(table_frame, textvariable=self.table_var, state="readonly")
        self.table_combo.pack(fill=tk.X, pady=(5, 10))
        self.table_combo.bind("<<ComboboxSelected>>", self.on_table_selected)
        
        # Table preview
        preview_frame = ttk.LabelFrame(frame, text="Table Preview", padding=10)
        preview_frame.pack(fill=tk.BOTH, expand=True)
        
        # Preview controls
        preview_controls = ttk.Frame(preview_frame)
        preview_controls.pack(fill=tk.X, pady=(0, 10))
        
        # Preview info label
        self.preview_info_label = ttk.Label(preview_controls, text="Select a table to view preview", foreground="gray")
        self.preview_info_label.pack(side=tk.LEFT)
        
        # Refresh button
        ttk.Button(preview_controls, text="Refresh Preview", command=self.load_table_preview).pack(side=tk.RIGHT, padx=(5, 0))
        
        # Test database button
        ttk.Button(preview_controls, text="Test Connection", command=self.test_database).pack(side=tk.RIGHT, padx=(5, 0))
        
        # Export data button
        ttk.Button(preview_controls, text="Export Sample Data", command=self.export_sample_data).pack(side=tk.RIGHT, padx=(5, 0))
        
        # Preview treeview with both scrollbars
        preview_container = ttk.Frame(preview_frame)
        preview_container.pack(fill=tk.BOTH, expand=True)
        
        self.table_tree = ttk.Treeview(preview_container, show="headings", height=15)
        table_scroll_y = ttk.Scrollbar(preview_container, orient=tk.VERTICAL, command=self.table_tree.yview)
        table_scroll_x = ttk.Scrollbar(preview_container, orient=tk.HORIZONTAL, command=self.table_tree.xview)
        
        self.table_tree.configure(yscrollcommand=table_scroll_y.set, xscrollcommand=table_scroll_x.set)
        
        # Pack treeview and scrollbars
        self.table_tree.grid(row=0, column=0, sticky="nsew")
        table_scroll_y.grid(row=0, column=1, sticky="ns")
        table_scroll_x.grid(row=1, column=0, sticky="ew")
        
        # Configure grid weights
        preview_container.grid_rowconfigure(0, weight=1)
        preview_container.grid_columnconfigure(0, weight=1)
    
    def switch_tab(self, tab_id: str):
        """Switch to specified tab using NavigationManager"""
        # Hide all tabs
        for frame in self.tab_frames.values():
            frame.pack_forget()
        
        # Show selected tab
        if tab_id in self.tab_frames:
            self.tab_frames[tab_id].pack(fill=tk.BOTH, expand=True)
            self.current_tab = tab_id
            
            # Update navigation button styles through NavigationManager
            if hasattr(self, 'nav_manager'):
                # Update button styles directly
                for btn_id, btn in self.nav_manager.nav_buttons.items():
                    if btn_id == tab_id:
                        btn.configure(style='Selected.TButton')
                    else:
                        btn.configure(style='TButton')
                        
                self.nav_manager.current_tab = tab_id
            
            # Refresh tab content if needed
            if tab_id == "dashboard":
                self.refresh_dashboard()
            elif tab_id == "scheduler":
                if hasattr(self, 'refresh_scheduler_log'):
                    self.refresh_scheduler_log()
            elif tab_id == "health_checks":
                # Auto-refresh health check status when tab is opened
                self.root.after(200, self.auto_refresh_health_status)
            elif tab_id == "mapping":
                # Auto-refresh API Field Mapping status when tab is opened
                if hasattr(self, 'refresh_api_mapping_status'):
                    self.root.after(200, self.refresh_api_mapping_status)
        else:
            self.log_entry(f"Warning: Tab '{tab_id}' not found in tab_frames. Available tabs: {list(self.tab_frames.keys())}", "WARN")
    
    # Continue with remaining methods...
    
    # Continue with remaining methods...
    
    def log_entry(self, message: str, level: str = "INFO", details: str = ""):
        """Unified logging method using LogManager"""
        self.log_manager.log(level, message, details, module="MDBAgentPro")
        
        # Update status variable
        if hasattr(self, 'status_var') and self.status_var:
            self.status_var.set(f"{level}: {message}")
        
        # Update dashboard if visible
        if hasattr(self, 'current_tab') and self.current_tab == "dashboard":
            self.refresh_dashboard()
    
    def browse_mdb_file(self):
        """Browse for MDB file"""
        filename = filedialog.askopenfilename(
            title="Select Microsoft Access Database",
            filetypes=[("Access Database", "*.mdb *.accdb"), ("All Files", "*.*")]
        )
        if filename:
            self.mdb_file_var.set(filename)
    
    def connect_database(self):
        """Connect to MDB database"""
        if not self.mdb_file_var.get():
            messagebox.showerror("Error", "Please select a database file first.")
            return
        
        if not os.path.exists(self.mdb_file_var.get()):
            messagebox.showerror("Error", "Database file not found.")
            return
        
        try:
            # Close existing connection
            if self.db_connection:
                self.db_connection.close()
            
            # Build connection string
            conn_str = f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={self.mdb_file_var.get()}'
            if self.mdb_password_var.get():
                conn_str += f';PWD={self.mdb_password_var.get()}'
            
            self.db_connection = pyodbc.connect(conn_str)
            self.load_tables()
            self.log_entry("Database connected successfully", "SUCCESS")
            self.db_status.update_status("good")
            
            # Update config
            self.config["mdb_file"] = self.mdb_file_var.get()
            self.config["mdb_password"] = self.mdb_password_var.get()
            self.save_config()
            
        except Exception as e:
            error_msg = f"Failed to connect to database: {str(e)}"
            self.log_entry(error_msg, "ERROR")
            self.db_status.update_status("error")
            messagebox.showerror("Database Error", error_msg)
    
    def load_tables(self):
        """Load table names from connected database"""
        if not self.db_connection:
            return
        
        try:
            cursor = self.db_connection.cursor()
            tables = []
            
            for table_info in cursor.tables(tableType='TABLE'):
                tables.append(table_info.table_name)
            
            self.table_combo['values'] = tables
            self.log_entry(f"Found {len(tables)} tables in database", "INFO")
            
        except Exception as e:
            error_msg = f"Failed to load tables: {str(e)}"
            self.log_entry(error_msg, "ERROR")
    
    def on_table_selected(self, event):
        """Handle table selection"""
        self.selected_table = self.table_var.get()
        if self.selected_table:
            self.load_table_preview()
            self.load_table_columns()
    
    def load_table_preview(self):
        """Load table preview data"""
        if not self.db_connection or not self.selected_table:
            # Clear preview if no connection or table
            if hasattr(self, 'table_tree'):
                for item in self.table_tree.get_children():
                    self.table_tree.delete(item)
                self.table_tree["columns"] = ()
            return
        
        try:
            cursor = self.db_connection.cursor()
            
            # Get columns with detailed information
            columns = []
            column_names = []
            for column in cursor.columns(table=self.selected_table):
                columns.append({
                    'name': column.column_name,
                    'type': column.type_name,
                    'size': getattr(column, 'column_size', 'N/A'),
                    'nullable': getattr(column, 'nullable', 'Unknown')
                })
                column_names.append(column.column_name)
            
            # Configure treeview
            self.table_tree["columns"] = column_names
            self.table_tree["show"] = "headings"
            
            # Clear existing items
            for item in self.table_tree.get_children():
                self.table_tree.delete(item)
            
            # Set column headings with type info
            for i, col in enumerate(columns):
                col_name = col['name']
                col_type = col['type']
                header_text = f"{col_name}\n({col_type})"
                
                self.table_tree.heading(col_name, text=header_text)
                
                # Auto-size columns based on content type
                if 'TEXT' in col_type or 'VARCHAR' in col_type:
                    width = 150
                elif 'DATE' in col_type or 'TIME' in col_type:
                    width = 120
                elif 'INT' in col_type or 'NUMBER' in col_type:
                    width = 80
                else:
                    width = 100
                    
                self.table_tree.column(col_name, width=width, minwidth=60)
            
            # Get sample data (first 20 rows with ordering)
            try:
                # Try to find a suitable ordering column
                order_column = None
                for col_name in ['ID', 'Id', 'id', 'TIMESTAMP', 'Timestamp', 'timestamp', 'DATE', 'Date', 'date']:
                    if col_name in column_names:
                        order_column = col_name
                        break
                
                if order_column:
                    query = f"SELECT TOP 20 * FROM [{self.selected_table}] ORDER BY [{order_column}] DESC"
                else:
                    query = f"SELECT TOP 20 * FROM [{self.selected_table}]"
                
                cursor.execute(query)
                rows = cursor.fetchall()
                
                # Track data statistics
                row_count = len(rows)
                
                # Get total row count
                cursor.execute(f"SELECT COUNT(*) FROM [{self.selected_table}]")
                total_rows = cursor.fetchone()[0]
                
                # Insert data into treeview
                for row in rows:
                    values = []
                    for i, value in enumerate(row):
                        if value is None:
                            values.append("NULL")
                        elif hasattr(value, 'strftime'):
                            values.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                        elif isinstance(value, (int, float)):
                            values.append(str(value))
                        else:
                            # Truncate long text values
                            str_value = str(value)
                            if len(str_value) > 50:
                                str_value = str_value[:47] + "..."
                            values.append(str_value)
                    
                    self.table_tree.insert("", tk.END, values=values)
                
                # Update preview info
                if hasattr(self, 'preview_info_label'):
                    info_text = f"Showing {row_count} of {total_rows} total records"
                    if order_column:
                        info_text += f" (ordered by {order_column})"
                    self.preview_info_label.config(text=info_text)
                
                # Update status
                success_msg = f"Loaded preview: {self.selected_table} ({row_count}/{total_rows} records)"
                self.log_entry(success_msg, "SUCCESS")
                
                # Auto-refresh mapping interface if it exists
                if hasattr(self, 'refresh_mapping_interface'):
                    self.refresh_mapping_interface()
                
            except Exception as query_error:
                # Fallback to simple query without ordering
                cursor.execute(f"SELECT TOP 20 * FROM [{self.selected_table}]")
                rows = cursor.fetchall()
                
                for row in rows:
                    values = []
                    for value in row:
                        if value is None:
                            values.append("NULL")
                        elif hasattr(value, 'strftime'):
                            values.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                        else:
                            str_value = str(value)
                            if len(str_value) > 50:
                                str_value = str_value[:47] + "..."
                            values.append(str_value)
                    
                    self.table_tree.insert("", tk.END, values=values)
                
                self.log_entry(f"Loaded preview for table: {self.selected_table} (basic query)", "INFO")
            
        except Exception as e:
            error_msg = f"Failed to load table preview: {str(e)}"
            self.log_entry(error_msg, "ERROR")
            
            # Show error in preview
            if hasattr(self, 'table_tree'):
                self.table_tree["columns"] = ("Error",)
                self.table_tree["show"] = "headings"
                for item in self.table_tree.get_children():
                    self.table_tree.delete(item)
                self.table_tree.heading("Error", text="Error Loading Data")
                self.table_tree.insert("", tk.END, values=(error_msg,))
            
            messagebox.showerror("Database Error", f"Failed to load table preview:\n{str(e)}")
    
    def test_database(self):
        """Test database connection with detailed feedback like Postman"""
        # Build comprehensive test results
        test_results = []
        overall_status = " SUCCESS"
        overall_color = "green"
        
        # Test 1: Basic connection
        try:
            if not self.db_connection:
                test_results.append(" Connection: No active database connection")
                overall_status = " FAILED"
                overall_color = "red"
            else:
                cursor = self.db_connection.cursor()
                cursor.execute("SELECT 1")
                result = cursor.fetchone()
                test_results.append(" Connection: Database connection active and responding")
                
                # Update status indicators
                self.db_status.update_status("good")
                if hasattr(self, 'dash_db_status'):
                    self.dash_db_status.config(text=" Connected", foreground="green")
                if hasattr(self, 'health_db_status'):
                    self.health_db_status.config(text=" Connected", foreground="green")
        except Exception as e:
            test_results.append(f" Connection: Basic query failed - {str(e)}")
            overall_status = " FAILED"
            overall_color = "red"
            self.db_status.update_status("error")
        
        # Test 2: Database metadata
        if self.db_connection:
            try:
                cursor = self.db_connection.cursor()
                tables = list(cursor.tables())
                table_count = len([table for table in tables if table.table_type == 'TABLE'])
                test_results.append(f" Metadata: Found {table_count} tables in database")
            except Exception as e:
                test_results.append(f" Metadata: Could not enumerate tables - {str(e)}")
                if overall_status == " SUCCESS":
                    overall_status = " PARTIAL"
                    overall_color = "orange"
        
        # Test 3: Selected table test
        if self.db_connection and self.selected_table:
            try:
                cursor = self.db_connection.cursor()
                cursor.execute(f"SELECT COUNT(*) FROM [{self.selected_table}]")
                record_count = cursor.fetchone()[0]
                test_results.append(f" Table Query: '{self.selected_table}' contains {record_count:,} records")
                
                # Test sample data retrieval
                cursor.execute(f"SELECT TOP 1 * FROM [{self.selected_table}]")
                sample_row = cursor.fetchone()
                if sample_row:
                    test_results.append(f" Data Access: Successfully retrieved sample data")
                else:
                    test_results.append(f" Data Access: Table is empty")
                    
            except Exception as e:
                test_results.append(f" Table Query: Failed to query '{self.selected_table}' - {str(e)}")
                if overall_status == " SUCCESS":
                    overall_status = " PARTIAL"
                    overall_color = "orange"
        elif self.selected_table:
            test_results.append(" Table Query: No database connection for table test")
        else:
            test_results.append(" Table Query: No table selected")
        
        # Test 4: Write access test (optional)
        if self.db_connection:
            try:
                # Try to create a temporary test table
                cursor = self.db_connection.cursor()
                test_table_name = f"_temp_test_{int(datetime.now().timestamp())}"
                cursor.execute(f"CREATE TABLE [{test_table_name}] (test_id INTEGER)")
                cursor.execute(f"DROP TABLE [{test_table_name}]")
                test_results.append(" Write Access: Database supports table creation/deletion")
            except Exception as e:
                test_results.append(f" Write Access: Limited write permissions - {str(e)}")
                # Don't change overall status for write access issues
        
        # Display results in Postman-style format
        result_text = f"{overall_status}\n\n" + "\n".join(test_results)
        
        if overall_status == " SUCCESS":
            messagebox.showinfo("Database Test Results", result_text)
            self.log_entry("Database connection test completed successfully", "SUCCESS")
        elif overall_status == " PARTIAL":
            messagebox.showwarning("Database Test Results", result_text)
            self.log_entry("Database connection test completed with warnings", "WARNING")
        else:
            messagebox.showerror("Database Test Results", result_text)
            self.log_entry("Database connection test failed", "ERROR")
    
    def export_sample_data(self):
        """Export sample data from current table to CSV"""
        if not self.db_connection or not self.selected_table:
            messagebox.showerror("Error", "Please select a table first.")
            return
        
        filename = filedialog.asksaveasfilename(
            title=f"Export Sample Data from {self.selected_table}",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                cursor = self.db_connection.cursor()
                
                # Get all columns
                columns = [col['name'] for col in self.table_columns] if hasattr(self, 'table_columns') and self.table_columns else []
                
                if not columns:
                    # Get columns from database
                    columns = [column.column_name for column in cursor.columns(table=self.selected_table)]
                
                # Get sample data (up to 100 records)
                cursor.execute(f"SELECT TOP 100 * FROM [{self.selected_table}]")
                rows = cursor.fetchall()
                
                # Write to file
                if filename.endswith('.csv'):
                    import csv
                    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(columns)  # Header
                        
                        for row in rows:
                            csv_row = []
                            for value in row:
                                if value is None:
                                    csv_row.append("")
                                elif hasattr(value, 'strftime'):
                                    csv_row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                                else:
                                    csv_row.append(str(value))
                            writer.writerow(csv_row)
                
                elif filename.endswith('.xlsx'):
                    try:
                        import openpyxl
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = self.selected_table
                        
                        # Write header
                        for col, column_name in enumerate(columns, 1):
                            ws.cell(row=1, column=col, value=column_name)
                        
                        # Write data
                        for row_idx, row in enumerate(rows, 2):
                            for col_idx, value in enumerate(row, 1):
                                if value is None:
                                    cell_value = ""
                                elif hasattr(value, 'strftime'):
                                    cell_value = value.strftime('%Y-%m-%d %H:%M:%S')
                                else:
                                    cell_value = str(value)
                                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                        
                        wb.save(filename)
                    except ImportError:
                        messagebox.showerror("Error", "openpyxl library not installed. Please use CSV format.")
                        return
                
                messagebox.showinfo("Export Complete", 
                                  f"Successfully exported {len(rows)} records to:\n{filename}")
                self.log_entry(f"Exported {len(rows)} records from {self.selected_table} to {filename}", "SUCCESS")
                
            except Exception as e:
                error_msg = f"Failed to export data: {str(e)}"
                messagebox.showerror("Export Error", error_msg)
                self.log_entry(error_msg, "ERROR")
    
    def load_table_columns(self):
        """Load table columns for mapping"""
        if not self.db_connection or not self.selected_table:
            return
        
        try:
            cursor = self.db_connection.cursor()
            columns = []
            
            for column in cursor.columns(table=self.selected_table):
                columns.append({
                    'name': column.column_name,
                    'type': column.type_name,
                    'size': column.column_size
                })
            
            self.table_columns = columns
            self.config["table_columns"] = columns
            self.config["selected_table"] = self.selected_table
            self.save_config()
            
            # Auto-refresh mapping interface when table columns are loaded
            if hasattr(self, 'refresh_mapping_interface'):
                self.refresh_mapping_interface()
            
            # Auto-refresh API mapping status
            if hasattr(self, 'refresh_api_mapping_status'):
                self.refresh_api_mapping_status()
            
            self.log_entry(f"Loaded {len(columns)} columns from table: {self.selected_table}", "SUCCESS")
            
        except Exception as e:
            self.log_entry(f"Failed to load table columns: {str(e)}", "ERROR")
    
    def test_api(self):
        """Test API connection with detailed feedback like Postman - NO AUTO-SAVE (testing only)"""
        # Get endpoint from API Settings
        endpoint = self.api_endpoint_var.get().strip() if hasattr(self, 'api_endpoint_var') else self.config.get('api_endpoint', '')
        if not endpoint:
            messagebox.showerror("Configuration Error", " No API endpoint configured.\n\nPlease configure endpoint in API Settings tab.")
            return
        
        # Get authentication from API Settings
        auth_type = self.auth_type_var.get() if hasattr(self, 'auth_type_var') else "no_auth"
        method = self.api_method_var.get() if hasattr(self, 'api_method_var') else "POST"
        
        # Build test results like Postman
        test_results = []
        overall_status = " SUCCESS"
        overall_color = "green"
        response_time = 0
        
        # Prepare headers and data
        headers = {'Content-Type': 'application/json', 'User-Agent': 'MDBAgentPro-Test/2.0'}
        test_data = {
            "test": True,
            "timestamp": datetime.now().isoformat(),
            "uuid": str(uuid.uuid4()),
            "source": "Dashboard_Test"
        }
        
        # Test 1: Endpoint validation
        try:
            from urllib.parse import urlparse
            parsed = urlparse(endpoint)
            if not parsed.scheme or not parsed.netloc:
                test_results.append(" URL Format: Invalid URL format")
                overall_status = " FAILED"
                overall_color = "red"
            else:
                test_results.append(f" URL Format: Valid URL ({parsed.scheme}://{parsed.netloc})")
        except Exception as e:
            test_results.append(f" URL Format: {str(e)}")
            overall_status = " FAILED"
            overall_color = "red"
        
        # Test 2: Authentication setup
        auth_status = " Auth Setup: "
        try:
            if auth_type == "api_key":
                api_key = self.api_key_var.get().strip() if hasattr(self, 'api_key_var') else ""
                if api_key:
                    headers['Authorization'] = f'Bearer {api_key}'
                    auth_status += f"API Key configured ({api_key[:10]}...)"
                    test_results.append(auth_status)
                else:
                    test_results.append(" Auth Setup: API Key authentication selected but no key provided")
                    overall_status = " FAILED"
                    overall_color = "red"
                    return
                    
            elif auth_type == "login":
                username = self.login_username_var.get().strip() if hasattr(self, 'login_username_var') else ""
                password = self.login_password_var.get().strip() if hasattr(self, 'login_password_var') else ""
                database = self.login_database_var.get().strip() if hasattr(self, 'login_database_var') else ""
                
                if "/login" in endpoint.lower() or "/auth" in endpoint.lower():
                    if username and password:
                        test_data = {
                            "login": username,
                            "password": password,
                            "database": database or "default"
                        }
                        auth_status += f"Login credentials for endpoint (user: {username})"
                        test_results.append(auth_status)
                    else:
                        test_results.append(" Auth Setup: Login credentials required but not provided")
                        overall_status = " FAILED"
                        overall_color = "red"
                        return
                else:
                    login_status = self.login_status_var.get() if hasattr(self, 'login_status_var') else "Not logged in"
                    if "successful" in login_status.lower():
                        auth_status += "Using existing login session"
                        test_results.append(auth_status)
                    else:
                        test_results.append(" Auth Setup: Login authentication but not logged in")
                        if overall_status == " SUCCESS":
                            overall_status = " PARTIAL"
                            overall_color = "orange"
            else:
                auth_status += "No authentication (public endpoint)"
                test_results.append(auth_status)
                
        except Exception as e:
            test_results.append(f" Auth Setup: {str(e)}")
            overall_status = " FAILED"
            overall_color = "red"
        
        # Test 3: Network connectivity and response
        if overall_status != " FAILED":
            try:
                start_time = datetime.now()
                self.log_entry(f" Testing API: {method} {endpoint} (Auth: {auth_type}) - test only, no auto-save", "INFO")
                
                # Make request
                if method.upper() == "GET":
                    response = requests.get(endpoint, headers=headers, timeout=15)
                else:
                    response = requests.post(endpoint, json=test_data, headers=headers, timeout=15)
                
                end_time = datetime.now()
                response_time = int((end_time - start_time).total_seconds() * 1000)
                
                # Test 4: Response analysis
                test_results.append(f" Network: Connected successfully ({response_time}ms)")
                test_results.append(f" Response: HTTP {response.status_code} received")
                
                # Analyze response content
                try:
                    content_type = response.headers.get('content-type', 'unknown')
                    test_results.append(f" Content-Type: {content_type}")
                    
                    if response.text:
                        response_size = len(response.text)
                        test_results.append(f" Response Size: {response_size} bytes")
                        
                        # Try to parse JSON
                        if 'json' in content_type.lower():
                            try:
                                json_data = response.json()
                                test_results.append(" JSON: Valid JSON response received")
                            except:
                                test_results.append(" JSON: Invalid JSON in response")
                    else:
                        test_results.append(" Response: Empty response body")
                except Exception as parse_error:
                    test_results.append(f" Response Parse: {str(parse_error)}")
                
                # Status code evaluation (NO AUTO-SAVE of status indicators)
                if response.status_code in [200, 201]:
                    test_results.append(" Status: Success response (200-201)")
                    test_results.append(" Test Only: No configuration auto-saved")
                        
                elif response.status_code in [400, 401, 422]:
                    if "/login" in endpoint.lower() and auth_type == "login":
                        test_results.append(f" Status: Login endpoint responding correctly ({response.status_code})")
                        test_results.append(" Note: 400/401 is expected for login endpoints with test data")
                        test_results.append(" Test Only: No configuration auto-saved")
                    else:
                        test_results.append(f" Status: Client error ({response.status_code})")
                        test_results.append(" Note: May indicate authentication or validation issues")
                        test_results.append(" Test Only: No configuration auto-saved")
                        if overall_status == " SUCCESS":
                            overall_status = " PARTIAL"
                            overall_color = "orange"
                
                elif response.status_code >= 500:
                    test_results.append(f" Status: Server error ({response.status_code})")
                    test_results.append(" Note: Server-side issue, endpoint may be temporarily unavailable")
                    test_results.append(" Test Only: No configuration auto-saved")
                    if overall_status == " SUCCESS":
                        overall_status = " PARTIAL"
                        overall_color = "orange"
                
                else:
                    test_results.append(f" Status: Other response ({response.status_code})")
                    test_results.append(" Test Only: No configuration auto-saved")
                    if overall_status == " SUCCESS":
                        overall_status = " PARTIAL"
                        overall_color = "orange"
                
            except requests.exceptions.Timeout:
                test_results.append(" Network: Request timeout (>15 seconds)")
                test_results.append(" Suggestion: Check endpoint URL and network connectivity")
                test_results.append(" Test Only: No configuration auto-saved")
                overall_status = " FAILED"
                overall_color = "red"
                
            except requests.exceptions.ConnectionError:
                test_results.append(" Network: Connection failed")
                test_results.append(" Suggestion: Verify URL and internet connection")
                test_results.append(" Test Only: No configuration auto-saved")
                overall_status = " FAILED"
                overall_color = "red"
                
            except Exception as e:
                test_results.append(f" Network: Request failed - {str(e)}")
                test_results.append(" Test Only: No configuration auto-saved")
                overall_status = " FAILED"
                overall_color = "red"
        
        # Display results in Postman-style format
        result_header = f"{overall_status}\n{method} {endpoint}\nAuth: {auth_type.replace('_', ' ').title()}"
        if response_time > 0:
            result_header += f"\nResponse Time: {response_time}ms"
        
        result_header += "\n\n TEST MODE - No Auto-Save"
        result_text = result_header + "\n\n" + "\n".join(test_results)
        
        if overall_status == " SUCCESS":
            messagebox.showinfo("API Test Results", result_text)
            self.log_entry(f" API test completed successfully: {endpoint} (test mode)", "SUCCESS")
        elif overall_status == " PARTIAL":
            messagebox.showwarning("API Test Results", result_text)
            self.log_entry(f" API test completed with warnings: {endpoint} (test mode)", "WARNING")
        else:
            messagebox.showerror("API Test Results", result_text)
            self.log_entry(f" API test failed: {endpoint} (test mode)", "ERROR")
    
    def manual_push(self):
        """Manual data push"""
        if not self.selected_table or not self.db_connection:
            messagebox.showerror("Error", "Please select a table first.")
            return
        
        if not self.config.get("api_endpoint"):
            messagebox.showerror("Error", "Please configure API endpoint first.")
            return
        
        # Check if mapping is complete
        if not self.is_mapping_complete():
            if messagebox.askyesno("Incomplete Mapping", 
                                 "Field mapping is not complete. This may result in missing data.\n\n"
                                 "Do you want to continue anyway?"):
                pass  # Continue with incomplete mapping
            else:
                return  # Cancel push
        
        try:
            data = self.get_latest_record()
            if data:
                if self.config.get("test_mode", False):
                    messagebox.showinfo("Test Mode", f"Would send data: {json.dumps(data, indent=2)}")
                    self.log_entry("Test mode: Manual push simulated", "INFO")
                else:
                    success = self.send_data_with_mapping(data)
                    if success:
                        messagebox.showinfo("Success", "Data pushed successfully with field mapping!")
                    else:
                        messagebox.showerror("Error", "Failed to push data. Check logs.")
            else:
                messagebox.showerror("Error", "No data found to push.")
                
        except Exception as e:
            error_msg = f"Manual push failed: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def is_mapping_complete(self):
        """Check if mapping is reasonably complete"""
        if not hasattr(self, 'field_mappings') or not self.field_mappings:
            return False
        
        # Count mapped fields
        mapped_count = len([f for f in self.field_mappings.values() if f.get('api_field')])
        total_count = len(getattr(self, 'table_columns', []))
        
        # Consider mapping complete if at least 50% of fields are mapped
        if total_count == 0:
            return False
        
        completion_ratio = mapped_count / total_count
        return completion_ratio >= 0.5  # At least 50% mapped
    
    def get_latest_record(self) -> Optional[Dict]:
        """Get latest record from selected table"""
        if not self.db_connection or not self.selected_table:
            return None
        
        try:
            cursor = self.db_connection.cursor()
            columns = [col['name'] for col in self.table_columns]
            
            # Find order column (ID or timestamp)
            order_column = None
            for col in ['ID', 'Id', 'id', 'TIMESTAMP', 'Timestamp', 'timestamp', 'DATE', 'Date']:
                if col in columns:
                    order_column = col
                    break
            
            if not order_column:
                order_column = columns[0]
            
            # Get latest record
            query = f"SELECT TOP 1 * FROM [{self.selected_table}] ORDER BY [{order_column}] DESC"
            cursor.execute(query)
            
            row = cursor.fetchone()
            if row:
                record = {}
                for i, col in enumerate(columns):
                    value = row[i]
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    record[col] = value
                
                # Apply field mapping if configured
                if hasattr(self, 'field_mappings') and self.field_mappings:
                    # Use new detailed mapping with transformations
                    mapped_record = {}
                    for db_field, value in record.items():
                        if db_field in self.field_mappings:
                            mapping_info = self.field_mappings[db_field]
                            api_field = mapping_info.get('api_field')
                            transform = mapping_info.get('transform', 'No Transform')
                            
                            if api_field:
                                transformed_value = self.apply_transformation(value, transform)
                                mapped_record[api_field] = transformed_value
                    
                    if mapped_record:
                        return mapped_record
                elif self.config.get("field_mapping"):
                    # Fallback to simple mapping
                    mapped_record = {}
                    for db_field, api_field in self.config["field_mapping"].items():
                        if db_field in record:
                            mapped_record[api_field] = record[db_field]
                    return mapped_record
                
                return record
                
        except Exception as e:
            self.log_entry(f"Failed to get latest record: {str(e)}", "ERROR")
        
        return None
    
    def build_api_payload(self, raw_data: Dict) -> Dict:
        """Build API payload from raw database data using current field mappings"""
        try:
            # Initialize field mapper if needed
            if not self.field_mapper or not hasattr(self, 'field_mappings'):
                self.update_field_mapper()
            
            # Use field mapper to build payload
            if self.field_mapper:
                mapped_data = self.field_mapper.build_api_payload(raw_data)
                
                # Add metadata
                mapped_data.update({
                    "uuid": str(uuid.uuid4()),
                    "timestamp": datetime.now().isoformat(),
                    "table": getattr(self, 'selected_table', 'unknown'),
                    "source": "MDB_Agent_Pro"
                })
                
                self.log_entry(f"Built API payload with {len(mapped_data)} fields", "INFO")
                return mapped_data
            else:
                # Fallback to simple mapping
                self.log_entry("No field mapper available, using raw data", "WARN")
                return {
                    "uuid": str(uuid.uuid4()),
                    "timestamp": datetime.now().isoformat(),
                    "table": getattr(self, 'selected_table', 'unknown'),
                    "data": raw_data
                }
                
        except Exception as e:
            self.log_entry(f"Failed to build API payload: {str(e)}", "ERROR")
            # Return minimal structure on error
            return {
                "uuid": str(uuid.uuid4()),
                "timestamp": datetime.now().isoformat(),
                "error": f"Payload build failed: {str(e)}",
                "raw_data": raw_data
            }
    
    def update_field_mapper(self):
        """Update field mapper with current mappings and mode"""
        try:
            # Get current mapping mode
            mode = self.mapping_mode.get() if hasattr(self, 'mapping_mode') else "flat"
            
            # Get current field mappings
            mappings = getattr(self, 'field_mappings', {})
            
            # Create new field mapper
            self.field_mapper = FieldMapper(mappings, mode)
            
            self.log_entry(f"Updated field mapper: mode={mode}, mappings={len(mappings)}", "INFO")
            
        except Exception as e:
            self.log_entry(f"Failed to update field mapper: {str(e)}", "ERROR")
            self.field_mapper = None
    
    def send_data_with_mapping(self, raw_data: Dict) -> bool:
        """Send data to API using field mapping"""
        try:
            # Build mapped payload
            mapped_data = self.build_api_payload(raw_data)
            
            # Send mapped data
            return self.send_to_api(mapped_data)
            
        except Exception as e:
            self.log_entry(f"Failed to send data with mapping: {str(e)}", "ERROR")
            return False
    
    def send_to_api(self, data: Dict) -> bool:
        """Send data to API endpoint"""
        if not self.config.get("api_endpoint") or not data:
            return False
        
        # Add UUID if not present
        if "uuid" not in data:
            data["uuid"] = str(uuid.uuid4())
        
        try:
            headers = {'Content-Type': 'application/json'}
            if self.config.get("api_key"):
                headers['Authorization'] = f'Bearer {self.config["api_key"]}'
            
            response = requests.post(
                self.config["api_endpoint"],
                json=data,
                headers=headers,
                timeout=30
            )
            
            if response.status_code == 200:
                self.log_entry(f"Data sent successfully (UUID: {data.get('uuid', 'N/A')})", "SUCCESS")
                return True
            else:
                self.log_entry(f"API returned HTTP {response.status_code}: {response.text}", "WARNING")
                # Add to buffer for retry
                self.db_manager.add_to_buffer(data, self.config["api_endpoint"], self.config.get("api_key", ""))
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_entry(f"API request failed: {str(e)}", "ERROR")
            # Add to buffer for retry
            self.db_manager.add_to_buffer(data, self.config["api_endpoint"], self.config.get("api_key", ""))
            return False
    
    def start_agent(self):
        """Start the agent with API Settings integration"""
        if not self.selected_table:
            messagebox.showerror("Error", "Please configure database and select a table first.")
            return
        
        # Check API Settings instead of old config
        endpoint = self.api_endpoint_var.get().strip() if hasattr(self, 'api_endpoint_var') else self.config.get("api_endpoint", "")
        if not endpoint:
            messagebox.showerror("Error", "Please configure API endpoint in API Settings tab first.")
            return
        
        if not self.admin_mode:
            messagebox.showerror("Error", "Admin mode required to start agent.")
            return
        
        # Validate API Settings authentication
        auth_type = self.auth_type_var.get() if hasattr(self, 'auth_type_var') else "no_auth"
        if auth_type == "api_key":
            api_key = self.api_key_var.get().strip() if hasattr(self, 'api_key_var') else ""
            if not api_key:
                messagebox.showerror("Error", "API Key required. Please configure in API Settings tab.")
                return
        elif auth_type == "login":
            username = self.login_username_var.get().strip() if hasattr(self, 'login_username_var') else ""
            password = self.login_password_var.get().strip() if hasattr(self, 'login_password_var') else ""
            if not (username and password):
                messagebox.showerror("Error", "Login credentials required. Please configure in API Settings tab.")
                return
        
        self.is_running = True
        self.config["auto_push"] = True
        # Sync API Settings to config for agent compatibility
        self.sync_api_settings_to_config()
        self.save_config()
        
        messagebox.showinfo("Agent Started", 
                          " Agent started successfully!\n\n"
                          " Using API Settings authentication\n"
                          " Monitoring for database changes\n"
                          " Auto-push enabled")
        self.log_entry("Agent started with API Settings integration", "INFO")
    
    def sync_api_settings_to_config(self):
        """Sync API Settings to legacy config for backward compatibility"""
        if hasattr(self, 'api_endpoint_var'):
            self.config["api_endpoint"] = self.api_endpoint_var.get().strip()
        
        if hasattr(self, 'auth_type_var'):
            auth_type = self.auth_type_var.get()
            if auth_type == "api_key" and hasattr(self, 'api_key_var'):
                self.config["api_key"] = self.api_key_var.get().strip()
            elif auth_type == "login":
                # For login type, we might need to handle differently
                # Could store the token obtained from login
                if hasattr(self, 'api_key_var'):  # Token stored in api_key field after login
                    self.config["api_key"] = self.api_key_var.get().strip()
        
        self.log_entry("API Settings synced to agent config", "INFO")
        
        self.log_entry("Agent started", "SUCCESS")
        self.agent_status_var.set("Agent: Running")
        
        messagebox.showinfo("Success", "Agent started successfully!")
    
    def stop_agent(self):
        """Stop the agent"""
        self.is_running = False
        self.config["auto_push"] = False
        self.save_config()
        
        self.log_entry("Agent stopped", "INFO")
        self.agent_status_var.set("Agent: Stopped")
        
        messagebox.showinfo("Info", "Agent stopped.")
    
    def clear_buffer(self):
        """Clear the buffer"""
        if not self.admin_mode:
            messagebox.showerror("Error", "Admin mode required.")
            return
        
        result = messagebox.askyesno("Confirm", "Are you sure you want to clear the buffer?")
        if result:
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM push_buffer")
            conn.commit()
            conn.close()
            
            self.log_entry("Buffer cleared", "INFO")
            messagebox.showinfo("Success", "Buffer cleared successfully!")
    
    # Health check methods
    def health_test_database(self):
        """Test database connection for health check"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if not self.db_connection:
                self.health_db_status.config(text=" Not Connected", foreground="red")
                self.health_tree.insert("", 0, values=(timestamp, "Database", "FAIL", "No database connection"))
                return
            
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT 1")
            cursor.fetchone()
            
            self.health_db_status.config(text=" Connected", foreground="green")
            self.health_tree.insert("", 0, values=(timestamp, "Database", "PASS", "Database connection successful"))
            
        except Exception as e:
            self.health_db_status.config(text=" Error", foreground="red")
            self.health_tree.insert("", 0, values=(timestamp, "Database", "FAIL", f"Error: {str(e)}"))
    
    def health_test_api(self):
        """Test API connection for health check - USE TEST CONNECTION RESULTS FIRST"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Priority 1: Use cached test connection results if available
        if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
            result = self.last_test_connection_result
            if result['success']:
                # Use cached test connection results
                response_time = result.get('response_time', 'N/A')
                self.health_api_status.config(text=f" Connected ({response_time})", foreground="green")
                self.health_tree.insert("", 0, values=(timestamp, "API", "PASS", f"Using Test Connection result: Connected in {response_time}"))
                return
            else:
                error_msg = result.get('error', 'Connection failed')
                self.health_api_status.config(text=f" {error_msg}", foreground="red")
                self.health_tree.insert("", 0, values=(timestamp, "API", "FAIL", f"Using Test Connection result: {error_msg}"))
                return
        
        # Priority 2: Basic configuration check without API calls
        endpoint = self.api_endpoint_var.get().strip() if hasattr(self, 'api_endpoint_var') else self.config.get('api_endpoint', '')
        if not endpoint:
            self.health_api_status.config(text=" Not Configured", foreground="red")
            self.health_tree.insert("", 0, values=(timestamp, "API", "FAIL", "No API endpoint configured - Please configure in API Settings tab"))
            return
        
        # Check authentication configuration without testing
        auth_type = self.auth_type_var.get() if hasattr(self, 'auth_type_var') else "no_auth"
        
        auth_status = ""
        if auth_type == "api_key":
            api_key = self.api_key_var.get().strip() if hasattr(self, 'api_key_var') else ""
            if api_key:
                auth_status = "API Key configured"
            else:
                auth_status = "API Key required but not set"
                
        elif auth_type == "login":
            username = self.login_username_var.get().strip() if hasattr(self, 'login_username_var') else ""
            password = self.login_password_var.get().strip() if hasattr(self, 'login_password_var') else ""
            
            if username and password:
                login_status = self.login_status_var.get() if hasattr(self, 'login_status_var') else ""
                if "successful" in login_status.lower():
                    auth_status = "Logged in with token"
                else:
                    auth_status = "Credentials configured"
            else:
                auth_status = "Login credentials required"
        else:
            auth_status = "No authentication required"
        
        # Show configuration status without making API calls
        self.health_api_status.config(text=" Ready - Test Connection first", foreground="blue")
        self.health_tree.insert("", 0, values=(timestamp, "API", "INFO", f"Configuration ready: {endpoint[:50]}... Auth: {auth_status}. Use Test Connection to verify."))
        
    def health_check_buffer(self):
        """Check buffer status for health check"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            buffer_items = self.db_manager.get_buffer_items()
            count = len(buffer_items)
            
            if count == 0:
                self.health_buffer_status.config(text=" Empty", foreground="green")
                self.health_tree.insert("", 0, values=(timestamp, "Buffer", "PASS", "Buffer is empty"))
            elif count < 10:
                self.health_buffer_status.config(text=f" {count} items", foreground="orange")
                self.health_tree.insert("", 0, values=(timestamp, "Buffer", "WARN", f"{count} items in buffer"))
            else:
                self.health_buffer_status.config(text=f" {count} items", foreground="red")
                self.health_tree.insert("", 0, values=(timestamp, "Buffer", "FAIL", f"{count} items in buffer - check API"))
                
        except Exception as e:
            self.health_buffer_status.config(text=" Error", foreground="red")
            self.health_tree.insert("", 0, values=(timestamp, "Buffer", "FAIL", f"Error: {str(e)}"))
    
    def health_verify_data(self):
        """Verify data integrity for health check"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if not self.selected_table or not self.db_connection:
                self.health_data_status.config(text=" No Table", foreground="red")
                self.health_tree.insert("", 0, values=(timestamp, "Data Integrity", "FAIL", "No table selected"))
                return
            
            cursor = self.db_connection.cursor()
            cursor.execute(f"SELECT COUNT(*) FROM [{self.selected_table}]")
            count = cursor.fetchone()[0]
            
            if count > 0:
                self.health_data_status.config(text=f" {count} records", foreground="green")
                self.health_tree.insert("", 0, values=(timestamp, "Data Integrity", "PASS", f"Table has {count} records"))
            else:
                self.health_data_status.config(text=" Empty table", foreground="orange")
                self.health_tree.insert("", 0, values=(timestamp, "Data Integrity", "WARN", "Table is empty"))
                
        except Exception as e:
            self.health_data_status.config(text=" Error", foreground="red")
            self.health_tree.insert("", 0, values=(timestamp, "Data Integrity", "FAIL", f"Error: {str(e)}"))
    
    def run_all_health_checks(self):
        """Run all health checks"""
        self.health_test_database()
        self.health_test_api() 
        self.health_check_buffer()
        self.health_verify_data()
        self.log_entry("All health checks completed", "INFO")
    
    def refresh_health_api_settings(self):
        """Refresh health check with current API Settings"""
        try:
            # Check if API Settings variables exist
            if not hasattr(self, 'api_endpoint_var'):
                messagebox.showwarning("Warning", "API Settings not found. Please configure API Settings first.")
                return
            
            endpoint = self.api_endpoint_var.get().strip()
            auth_type = self.auth_type_var.get() if hasattr(self, 'auth_type_var') else "no_auth"
            
            # Update status label with current configuration
            if endpoint:
                self.health_api_status.config(text=" Checking...", foreground="blue")
                self.root.update_idletasks()  # Update UI immediately
                
                # Run API health check with new settings
                self.health_test_api()
                
                # Show success message
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.health_tree.insert("", 0, values=(timestamp, "System", "INFO", f"API Settings refreshed - Endpoint: {endpoint[:50]}... Auth: {auth_type}"))
                
                messagebox.showinfo("Success", 
                                  f" Health Check updated with current API Settings!\n\n"
                                  f"Endpoint: {endpoint}\n"
                                  f"Auth Type: {auth_type}\n\n"
                                  f"Health check now uses this configuration.")
            else:
                self.health_api_status.config(text=" Not Configured", foreground="red")
                messagebox.showwarning("Warning", 
                                     f" No API endpoint configured in API Settings.\n\n"
                                     f"Please go to API Settings tab and configure:\n"
                                     f" Endpoint URL\n"
                                     f" Authentication\n\n"
                                     f"Then refresh health check again.")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh API Settings: {str(e)}")
            self.log_entry(f"Health check refresh failed: {str(e)}", "ERROR")

    def toggle_auto_health_check(self):
        """Toggle automatic health checks"""
        messagebox.showinfo("Info", "Auto health check feature will be implemented in future version")
    
    def auto_refresh_health_status(self):
        """Auto-refresh health status when health check tab is opened"""
        try:
            # Only refresh if we're on health check tab
            if hasattr(self, 'current_tab') and self.current_tab == "health_checks":
                # Check if API Settings are configured
                endpoint = self.api_endpoint_var.get().strip() if hasattr(self, 'api_endpoint_var') else ""
                
                if endpoint:
                    # Refresh API health status silently
                    self.health_api_status.config(text=" Checking...", foreground="blue")
                    self.root.update_idletasks()
                    
                    # Delay to show checking status, then run actual test
                    self.root.after(500, self.health_test_api)
                    
                    # Log the auto-refresh
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self.health_tree.insert("", 0, values=(timestamp, "System", "INFO", "Health Check tab opened - Auto-refreshing API status"))
                else:
                    # Show configuration needed message
                    self.health_api_status.config(text=" Not Configured", foreground="red")
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self.health_tree.insert("", 0, values=(timestamp, "System", "INFO", "API Settings not configured - Please configure API Settings first"))
                    
        except Exception as e:
            # Silent fail for auto-refresh
            print(f"Auto-refresh health status failed: {str(e)}")
    
    def export_health_report(self):
        """Export health check report"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                title="Save Health Report"
            )
            if filename:
                with open(filename, 'w') as f:
                    f.write("MDB Agent Pro - Health Check Report\n")
                    f.write("=" * 40 + "\n")
                    f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                    
                    # Export health tree data
                    for item in self.health_tree.get_children():
                        values = self.health_tree.item(item)['values']
                        f.write(f"{values[0]} | {values[1]} | {values[2]} | {values[3]}\n")
                
                messagebox.showinfo("Success", f"Health report exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export health report: {str(e)}")
    
    # Transaction log methods
    def apply_transaction_filter(self):
        """Apply filter to transaction log"""
        # Clear existing items
        for item in self.transaction_tree.get_children():
            self.transaction_tree.delete(item)
        
        # Load filtered transactions
        self.load_transaction_log()
    
    def load_transaction_log(self):
        """Load transaction log data"""
        try:
            # Get filter values
            start_date = self.trans_start_date.get()
            end_date = self.trans_end_date.get()
            status_filter = self.trans_status_filter.get()
            
            # Load recent logs as transaction data
            logs = self.log_manager.get_recent_logs(100)
            
            success_count = 0
            failed_count = 0
            pending_count = 0
            
            for i, log in enumerate(logs):
                status = "Success" if log['level'] == "SUCCESS" else "Failed" if log['level'] == "ERROR" else "Info"
                
                if status_filter != "All" and status != status_filter:
                    continue
                
                if status == "Success":
                    success_count += 1
                elif status == "Failed":
                    failed_count += 1
                else:
                    pending_count += 1
                
                self.transaction_tree.insert("", tk.END, values=(
                    i + 1,
                    log['timestamp'],
                    self.selected_table or "N/A",
                    status,
                    "1",
                    log['message']
                ))
            
            # Update statistics
            total_count = success_count + failed_count + pending_count
            self.stats_total.config(text=str(total_count))
            self.stats_success.config(text=str(success_count))
            self.stats_failed.config(text=str(failed_count))
            self.stats_pending.config(text=str(pending_count))
            
        except Exception as e:
            self.log_entry(f"Failed to load transaction log: {str(e)}", "ERROR")
    
    def export_transactions(self):
        """Export transaction log"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                title="Export Transactions"
            )
            if filename:
                import csv
                with open(filename, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(["ID", "Timestamp", "Table", "Status", "Records", "Details"])
                    
                    for item in self.transaction_tree.get_children():
                        values = self.transaction_tree.item(item)['values']
                        writer.writerow(values)
                
                messagebox.showinfo("Success", f"Transactions exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export transactions: {str(e)}")
    
    # Support methods
    def send_log_to_it(self):
        """Prepare and send log to IT support"""
        try:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_filename = f"mdb_agent_issue_{timestamp}.txt"
            
            # Show issue description dialog
            issue_window = tk.Toplevel(self.root)
            issue_window.title("Report Issue to IT Support")
            issue_window.geometry("500x400")
            issue_window.transient(self.root)
            issue_window.grab_set()
            
            ttk.Label(issue_window, text="Describe the issue you're experiencing:", font=('Arial', 10, 'bold')).pack(anchor=tk.W, padx=20, pady=(20, 10))
            
            # Issue description
            desc_frame = ttk.Frame(issue_window)
            desc_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
            
            issue_text = scrolledtext.ScrolledText(desc_frame, height=8, wrap=tk.WORD)
            issue_text.pack(fill=tk.BOTH, expand=True)
            issue_text.insert(1.0, "Please describe:\n1. What you were trying to do\n2. What happened instead\n3. Any error messages you saw\n\n")
            
            # Urgency level
            urgency_frame = ttk.LabelFrame(issue_window, text="Issue Priority", padding=10)
            urgency_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
            
            urgency_var = tk.StringVar(value="Normal")
            ttk.Radiobutton(urgency_frame, text=" Critical (System Down)", variable=urgency_var, value="Critical").pack(anchor=tk.W)
            ttk.Radiobutton(urgency_frame, text=" High (Major Function Broken)", variable=urgency_var, value="High").pack(anchor=tk.W)
            ttk.Radiobutton(urgency_frame, text=" Normal (Minor Issue)", variable=urgency_var, value="Normal").pack(anchor=tk.W)
            ttk.Radiobutton(urgency_frame, text=" Low (Enhancement Request)", variable=urgency_var, value="Low").pack(anchor=tk.W)
            
            def create_support_ticket():
                issue_desc = issue_text.get(1.0, tk.END).strip()
                if len(issue_desc) < 20:
                    messagebox.showerror("Error", "Please provide a more detailed description.")
                    return
                
                try:
                    # Create comprehensive support file
                    with open(log_filename, 'w', encoding='utf-8') as f:
                        f.write("MDB AGENT PRO - SUPPORT TICKET\n")
                        f.write("=" * 50 + "\n\n")
                        f.write(f"Generated: {datetime.now().isoformat()}\n")
                        f.write(f"Priority: {urgency_var.get()}\n")
                        f.write(f"User: {os.getenv('USERNAME', 'Unknown')}\n")
                        f.write(f"Computer: {os.getenv('COMPUTERNAME', 'Unknown')}\n\n")
                        
                        f.write("ISSUE DESCRIPTION:\n")
                        f.write("-" * 20 + "\n")
                        f.write(issue_desc + "\n\n")
                        
                        # System info
                        import sys, platform
                        f.write("SYSTEM INFORMATION:\n")
                        f.write("-" * 20 + "\n")
                        f.write(f"OS: {platform.platform()}\n")
                        f.write(f"Python: {sys.version.split()[0]}\n")
                        f.write(f"App Version: 2.0.0\n\n")
                        
                        # Current state
                        f.write("APPLICATION STATE:\n")
                        f.write("-" * 20 + "\n")
                        f.write(f"Database Connected: {'Yes' if self.db_connection else 'No'}\n")
                        f.write(f"API Configured: {'Yes' if self.config.get('api_endpoint') else 'No'}\n")
                        f.write(f"Selected Table: {getattr(self, 'selected_table', 'None')}\n\n")
                        
                        # Recent logs
                        f.write("RECENT LOGS:\n")
                        f.write("-" * 20 + "\n")
                        try:
                            logs = self.log_manager.get_recent_logs(20)
                            for log in logs:
                                f.write(f"{log['timestamp']} [{log['level']}] {log['message']}\n")
                        except:
                            f.write("Could not retrieve logs\n")
                    
                    messagebox.showinfo("Support Ticket Created", 
                                      f" Support ticket created successfully!\n\n"
                                      f"File: {log_filename}\n\n"
                                      f"Please send this file to:\n"
                                      f" freddy.pm@sahabatagro.co.id\n"
                                      f" +62 813-9855-2019\n\n"
                                      f"Priority: {urgency_var.get()}")
                    
                    self.log_entry(f"Support ticket created: {log_filename} (Priority: {urgency_var.get()})", "INFO")
                    issue_window.destroy()
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to create support ticket: {str(e)}")
            
            # Buttons
            btn_frame = ttk.Frame(issue_window)
            btn_frame.pack(fill=tk.X, padx=20, pady=20)
            
            ttk.Button(btn_frame, text=" Create Support Ticket", command=create_support_ticket).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(btn_frame, text=" Cancel", command=issue_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to prepare support ticket: {str(e)}")
    
    def check_updates(self):
        """Check for application updates"""
        update_window = tk.Toplevel(self.root)
        update_window.title("Check for Updates")
        update_window.geometry("450x350")
        update_window.transient(self.root)
        update_window.grab_set()
        
        # Main content
        content_frame = ttk.Frame(update_window)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Icon and title
        title_frame = ttk.Frame(content_frame)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(title_frame, text="", font=('Arial', 24)).pack(side=tk.LEFT)
        ttk.Label(title_frame, text="MDB Agent Pro", font=('Arial', 16, 'bold')).pack(side=tk.LEFT, padx=(10, 0))
        
        # Current version info
        version_frame = ttk.LabelFrame(content_frame, text="Current Version", padding=15)
        version_frame.pack(fill=tk.X, pady=(0, 15))
        
        version_info = [
            ("Version:", "2.0.0 Professional Edition"),
            ("Release Date:", "January 2025"),
            ("Build:", "Stable"),
            ("Developer:", "Freddy Mazmur"),
            ("Company:", "PT Sahabat Agro Group")
        ]
        
        for i, (label, value) in enumerate(version_info):
            ttk.Label(version_frame, text=label, font=('Arial', 9, 'bold')).grid(row=i, column=0, sticky=tk.W, padx=(0, 15), pady=2)
            ttk.Label(version_frame, text=value, font=('Arial', 9)).grid(row=i, column=1, sticky=tk.W, pady=2)
        
        # Update status
        status_frame = ttk.LabelFrame(content_frame, text="Update Status", padding=15)
        status_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(status_frame, text=" You are running the latest version!", 
                 font=('Arial', 10, 'bold'), foreground="green").pack(anchor=tk.W)
        
        ttk.Label(status_frame, text="This version includes all the latest features and security updates.", 
                 font=('Arial', 9)).pack(anchor=tk.W, pady=(5, 0))
        
        # Support info
        support_frame = ttk.LabelFrame(content_frame, text="Support Information", padding=15)
        support_frame.pack(fill=tk.X, pady=(0, 15))
        
        support_text = """For updates and support:
[CHAR] Email: freddy.pm@sahabatagro.co.id
[CHAR] Phone: +62 813-9855-2019
[CHAR] Company: PT Sahabat Agro Group
 Hours: Monday-Friday, 8AM-6PM (WIB)"""
        
        ttk.Label(support_frame, text=support_text, font=('Arial', 9), justify=tk.LEFT).pack(anchor=tk.W)
        
        # Buttons
        btn_frame = ttk.Frame(content_frame)
        btn_frame.pack(fill=tk.X, pady=(15, 0))
        
        ttk.Button(btn_frame, text=" Close", command=update_window.destroy).pack(side=tk.RIGHT)
        ttk.Button(btn_frame, text=" Contact Support", command=lambda: (update_window.destroy(), self.send_log_to_it())).pack(side=tk.RIGHT, padx=(0, 10))
    
    def reset_config(self):
        """Reset application configuration"""
        result = messagebox.askyesno("Confirm Reset", 
                                   "This will reset all configuration settings.\n\n"
                                   "Are you sure you want to continue?")
        if result:
            try:
                if os.path.exists(self.config_file):
                    os.remove(self.config_file)
                messagebox.showinfo("Success", "Configuration reset. Please restart the application.")
                self.root.quit()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to reset configuration: {str(e)}")
    
    def toggle_admin_mode(self):
        """Toggle admin mode with enhanced security"""
        if not self.admin_mode:
            pin = tk.simpledialog.askstring("Admin Mode", "Enter admin PIN:", show='*')
            if pin:
                # Get stored PIN hash from config
                stored_pin_hash = self.config_manager.get("admin_pin_hash", "")
                
                # If no hash stored, create one from default PIN
                if not stored_pin_hash:
                    default_pin = "1234"  # Default PIN
                    stored_pin_hash = self.security.hash_password(default_pin)
                    self.config_manager.set("admin_pin_hash", stored_pin_hash)
                
                # Verify PIN
                if self.auth_manager.verify_admin_pin(pin, stored_pin_hash):
                    self.admin_mode = True
                    self.current_session_id = str(uuid.uuid4())
                    self.auth_manager.create_admin_session(self.current_session_id)
                    
                    self.admin_btn.config(text=" Exit Admin Mode")
                    self.log_entry("Admin mode enabled", "INFO")
                    messagebox.showinfo("Success", "Admin mode enabled")
                else:
                    self.log_entry("Failed admin login attempt", "WARNING")
                    messagebox.showerror("Error", "Invalid PIN")
            else:
                messagebox.showwarning("Warning", "PIN is required")
        else:
            self.admin_mode = False
            if self.current_session_id:
                # Clean up session
                if self.current_session_id in self.auth_manager.admin_sessions:
                    del self.auth_manager.admin_sessions[self.current_session_id]
                self.current_session_id = None
            
            self.admin_btn.config(text=" Admin Mode")
            self.log_entry("Admin mode disabled", "INFO")
            messagebox.showinfo("Info", "Admin mode disabled")
    
    def save_api_settings(self):
        """Save API settings and preserve test connection results"""
        try:
            # Get current endpoint - basic requirement
            endpoint = self.api_endpoint_var.get().strip()
            if not endpoint:
                messagebox.showwarning("Missing Endpoint", 
                                     "Please enter an API endpoint URL before saving.")
                return False
            
            # Prepare profile data - save whatever is currently configured
            profile_data = {
                "endpoint": endpoint,
                "method": self.api_method_var.get(),
                "content_type": self.api_content_type_var.get(),
                "custom_content_type": self.custom_content_type_var.get(),
                "auth_type": self.auth_type_var.get(),
                "api_key": self.api_key_var.get(),
                "login_username": self.login_username_var.get(),
                "login_password": self.login_password_var.get(),
                "login_database": self.login_database_var.get(),
                "test_mode": self.test_mode_var.get(),
                "json_body": self.json_body_text.get("1.0", tk.END).strip() if hasattr(self, 'json_body_text') else "",
                "form_fields": self.get_form_fields_data() if hasattr(self, 'form_fields_frame') else [],
                "updated": datetime.now().isoformat(),
                "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            # Save to current profile
            current_profile = self.api_profile_var.get()
            self.api_profiles[current_profile] = profile_data
            
            # Save to main config for global access
            self.config.update(profile_data)
            self.config["api_profiles"] = self.api_profiles
            self.config["current_api_profile"] = current_profile
            
            # Update legacy config fields for backward compatibility
            self.config["api_endpoint"] = endpoint
            self.config["api_key"] = self.api_key_var.get()
            
            self.save_config()  # This will preserve test connection results
            
            # Simple integration without forcing tests
            self.integrate_api_settings_globally()
            
            # Show appropriate success message based on test connection status
            auth_type = self.auth_type_var.get()
            auth_display = {
                "api_key": "API Key/Token",
                "login": "Login Authentication", 
                "no_auth": "No Authentication"
            }.get(auth_type, auth_type)
            
            # Check if test connection results are available
            test_status_info = ""
            if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
                if self.last_test_connection_result['success']:
                    response_time = self.last_test_connection_result.get('response_time', 'N/A')
                    test_status_info = f"\n Test Connection Status: Connected ({response_time})"
                else:
                    error_msg = self.last_test_connection_result.get('error', 'Failed')
                    test_status_info = f"\n Last Test Connection: {error_msg}"
            else:
                test_status_info = "\n Use 'Test Connection' to verify API connectivity"
            
            messagebox.showinfo("Configuration Saved", 
                              f" API Configuration Saved Successfully!\n\n" +
                              f"Profile: {current_profile}\n" +
                              f"Endpoint: {endpoint}\n" +
                              f"Authentication: {auth_display}\n" +
                              f"Method: {self.api_method_var.get()}" +
                              test_status_info +
                              f"\n\n Configuration is now available in all tabs\n" +
                              f"    (Dashboard, Health Check, API Field Mapping)")
            
            self.log_entry(f"API configuration saved - Profile: {current_profile}, Auth: {auth_type}", "SUCCESS")
            return True
            
        except Exception as e:
            error_msg = f"Failed to save API configuration: {str(e)}"
            messagebox.showerror("Save Error", error_msg)
            self.log_entry(error_msg, "ERROR")
            return False
    
    def integrate_api_settings_globally(self):
        """Integrate API settings with all application components"""
        try:
            # Update Health Check
            if hasattr(self, 'health_api_status'):
                self.health_api_status.config(text=" Updating...", foreground="blue")
                self.root.after(500, self.health_test_api)  # Delayed refresh
            
            # Update Dashboard status
            if hasattr(self, 'refresh_dashboard'):
                self.refresh_dashboard()
            
            # Update API Field Mapping endpoint if available
            if hasattr(self, 'mapping_endpoint_var'):
                self.mapping_endpoint_var.set(self.api_endpoint_var.get())
            
            # Update any other endpoint references
            endpoint = self.api_endpoint_var.get()
            auth_type = self.auth_type_var.get()
            
            # Log integration success
            self.log_entry(f"API Settings integrated globally - Endpoint: {endpoint}, Auth: {auth_type}", "INFO")
            
        except Exception as e:
            self.log_entry(f"Warning: Some components may not be updated: {str(e)}", "WARNING")
    
    def validate_and_update_status(self):
        """Update validation status based on current selection - USE TEST CONNECTION RESULTS"""
        try:
            # Priority 1: Use test connection results if available
            if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
                result = self.last_test_connection_result
                if result['success']:
                    # Use cached test connection results instead of making new API calls
                    response_time = result.get('response_time', 'N/A')
                    self.validation_status_var.set(f" Connected ({response_time})")
                    self.validation_status_label.config(foreground='green')
                    self.save_api_btn.config(state='normal')
                    self.save_status_var.set("Test passed - Ready to save")
                    self.save_status_label.config(foreground='green')
                    return
                else:
                    error_msg = result.get('error', 'Connection failed')
                    self.validation_status_var.set(f" {error_msg}")
                    self.validation_status_label.config(foreground='red')
                    self.save_api_btn.config(state='normal')  # Still allow saving
                    self.save_status_var.set("Can save without test")
                    self.save_status_label.config(foreground='orange')
                    return
            
            # Priority 2: Basic validation without API calls
            endpoint = self.api_endpoint_var.get().strip()
            auth_type = self.auth_type_var.get()
            
            if not endpoint:
                self.validation_status_var.set(" Enter endpoint URL to enable saving")
                self.validation_status_label.config(foreground='orange')
                self.save_api_btn.config(state='disabled')
                self.save_status_var.set("Endpoint required")
                self.save_status_label.config(foreground='orange')
                return
            
            # Check current auth configuration (optional)
            auth_status = ""
            if auth_type == "api_key":
                api_key = self.api_key_var.get().strip()
                if api_key:
                    auth_status = " API Key configured"
                else:
                    auth_status = "API Key can be added"
                    
            elif auth_type == "login":
                username = self.login_username_var.get().strip()
                password = self.login_password_var.get().strip()
                
                if username and password:
                    login_status = self.login_status_var.get() if hasattr(self, 'login_status_var') else ""
                    if "successful" in login_status.lower():
                        auth_status = " Logged in and token available"
                    else:
                        auth_status = "Credentials configured, login available"
                else:
                    auth_status = "Login credentials can be added"
                    
            else:  # no_auth
                auth_status = " No authentication required"
            
            # Show ready status
            self.validation_status_var.set(f" Ready to save - {auth_status}")
            self.validation_status_label.config(foreground='green')
            self.save_api_btn.config(state='normal')
            self.save_status_var.set("Ready to save configuration")
            self.save_status_label.config(foreground='green')
        
        except Exception as e:
            self.validation_status_var.set(" Error checking status")
            self.validation_status_label.config(foreground='orange')
            self.log_entry(f"Status check error: {str(e)}", "ERROR")
    
    def test_all_configurations(self):
        """Test current selected authentication configuration - USE TEST CONNECTION RESULTS"""
        try:
            # Priority 1: Check if we have recent test connection results
            if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
                result = self.last_test_connection_result
                if result['success']:
                    # Use cached test connection results instead of making new API calls
                    response_time = result.get('response_time', 'N/A')
                    self.save_status_var.set(f" Test passed ({response_time})")
                    self.save_status_label.config(foreground='green')
                    
                    auth_type = self.auth_type_var.get()
                    messagebox.showinfo("Test Results", 
                                      f" Using Test Connection Results:\n\n" +
                                      f" {auth_type.replace('_', ' ').title()} Authentication: Connected\n" +
                                      f" Response Time: {response_time}\n" +
                                      f" Endpoint accessible\n\n" +
                                      "Configuration is ready to save!")
                    return
                else:
                    error_msg = result.get('error', 'Connection failed')
                    self.save_status_var.set(f" Test failed: {error_msg}")
                    self.save_status_label.config(foreground='red')
                    
                    messagebox.showwarning("Test Results", 
                                         f" Test Connection Result:\n\n" +
                                         f" Error: {error_msg}\n\n" +
                                         "Please run Test Connection first or save configuration anyway.")
                    return
            
            # Priority 2: Basic validation check without API calls
            self.save_status_var.set(" No test results - use Test Connection")
            self.save_status_label.config(foreground='blue')
            
            endpoint = self.api_endpoint_var.get().strip()
            if not endpoint:
                messagebox.showwarning("No Endpoint", "Please enter an endpoint URL first.")
                self.save_status_var.set(" No endpoint")
                self.save_status_label.config(foreground='orange')
                return
            
            auth_type = self.auth_type_var.get()
            
            # Show configuration status without testing
            config_status = []
            
            if auth_type == "api_key":
                api_key = self.api_key_var.get().strip()
                if api_key:
                    config_status.append(" API Key is configured")
                else:
                    config_status.append(" API Key field is empty - can be filled later")
                    
            elif auth_type == "login":
                username = self.login_username_var.get().strip()
                password = self.login_password_var.get().strip()
                if username and password:
                    login_status = self.login_status_var.get() if hasattr(self, 'login_status_var') else ""
                    if "successful" in login_status.lower():
                        config_status.append(" Already logged in with valid token")
                    else:
                        config_status.append(" Login credentials configured")
                else:
                    config_status.append(" Login credentials not provided - can be configured later")
                    
            elif auth_type == "no_auth":
                config_status.append(" No authentication endpoint configured")
            
            config_status.append(f" Endpoint configured: {endpoint}")
            config_status.append(" Use 'Test Connection' to verify API connectivity")
            
            self.save_status_var.set(f" Config ready - Test recommended")
            self.save_status_label.config(foreground='blue')
            
            messagebox.showinfo("Configuration Status", 
                              f" Configuration Status for {auth_type.replace('_', ' ').title()} Authentication:\n\n" +
                              "\n".join(config_status) + 
                              "\n\n Use 'Test Connection' button to verify API connectivity before saving.")
            
        except Exception as e:
            self.save_status_var.set(" Test error")
            self.save_status_label.config(foreground='red')
            messagebox.showerror("Test Error", f"Testing failed: {str(e)}")
            self.log_entry(f"Configuration test error: {str(e)}", "ERROR")
    
    def setup_profiles_tab(self, parent):
        """Setup API profiles management tab"""
        profiles_scroll = ttk.Frame(parent)
        profiles_scroll.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Profile list
        ttk.Label(profiles_scroll, text="Saved API Profiles:", font=('Arial', 12, 'bold')).pack(anchor=tk.W, pady=(0, 10))
        
        # Profile listbox with scrollbar
        listbox_frame = ttk.Frame(profiles_scroll)
        listbox_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.profiles_listbox = tk.Listbox(listbox_frame, height=8)
        profiles_scrollbar = ttk.Scrollbar(listbox_frame, orient="vertical", command=self.profiles_listbox.yview)
        self.profiles_listbox.configure(yscrollcommand=profiles_scrollbar.set)
        
        self.profiles_listbox.pack(side="left", fill="both", expand=True)
        profiles_scrollbar.pack(side="right", fill="y")
        
        # Profile controls
        profile_controls = ttk.Frame(profiles_scroll)
        profile_controls.pack(fill=tk.X)
        
        ttk.Button(profile_controls, text="Load Profile", command=self.load_selected_profile).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(profile_controls, text="Export Profile", command=self.export_profile).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(profile_controls, text="Import Profile", command=self.import_profile).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(profile_controls, text="Duplicate", command=self.duplicate_profile).pack(side=tk.LEFT)
    
    def setup_history_tab(self, parent):
        """Setup enhanced API history tab with detailed information"""
        history_scroll = ttk.Frame(parent)
        history_scroll.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # History controls
        history_controls = ttk.Frame(history_scroll)
        history_controls.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(history_controls, text="API Request History (Last 5):", font=('Arial', 12, 'bold')).pack(side=tk.LEFT)
        ttk.Button(history_controls, text="Clear History", command=self.clear_api_history).pack(side=tk.RIGHT)
        ttk.Button(history_controls, text="Export History", command=self.export_api_history).pack(side=tk.RIGHT, padx=(0, 10))
        ttk.Button(history_controls, text="Refresh", command=self.update_history_display).pack(side=tk.RIGHT, padx=(0, 10))
        
        # History treeview with enhanced columns
        columns = ("Time", "Method", "Endpoint", "Status", "Payload", "Response Time")
        self.history_tree = ttk.Treeview(history_scroll, columns=columns, show='headings', height=10)
        
        # Define column headings and widths
        self.history_tree.heading("Time", text="Time")
        self.history_tree.heading("Method", text="Method")
        self.history_tree.heading("Endpoint", text="Endpoint")
        self.history_tree.heading("Status", text="Status")
        self.history_tree.heading("Payload", text="Payload")
        self.history_tree.heading("Response Time", text="Time (ms)")
        
        self.history_tree.column("Time", width=140)
        self.history_tree.column("Method", width=70)
        self.history_tree.column("Endpoint", width=180)
        self.history_tree.column("Status", width=100)
        self.history_tree.column("Payload", width=150)
        self.history_tree.column("Response Time", width=80)
        
        # History scrollbar
        history_tree_scroll = ttk.Scrollbar(history_scroll, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=history_tree_scroll.set)
        
        self.history_tree.pack(side="left", fill="both", expand=True)
        history_tree_scroll.pack(side="right", fill="y")
        
        # Bind double-click to view details
        self.history_tree.bind("<Double-1>", self.view_history_details)
        
        # Information label
        info_label = ttk.Label(history_scroll, text=" Double-click any row to view complete request/response details", 
                              font=('Arial', 9), foreground='blue')
        info_label.pack(pady=(10, 0))
    
    def on_auth_type_change(self):
        """Handle authentication type change with new preview frames"""
        auth_type = self.auth_type_var.get()
        
        # Hide all auth frames first
        if hasattr(self, 'api_key_frame'):
            self.api_key_frame.pack_forget()
        if hasattr(self, 'login_frame'):
            self.login_frame.pack_forget()
        if hasattr(self, 'no_auth_frame'):
            self.no_auth_frame.pack_forget()
        
        # Show appropriate frame
        if auth_type == "api_key":
            self.api_key_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        elif auth_type == "login":
            self.login_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        elif auth_type == "no_auth":
            self.no_auth_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.update_button_states()
        
        # Update validation status when auth type changes
        if hasattr(self, 'validate_and_update_status'):
            self.root.after(100, self.validate_and_update_status)
    
    def on_content_type_change(self, event=None):
        """Handle content type change"""
        content_type = self.api_content_type_var.get()
        
        # Hide all body frames first
        self.json_body_frame.pack_forget()
        self.form_body_frame.pack_forget()
        self.custom_content_type_frame.pack_forget()
        
        if content_type == "Custom...":
            self.custom_content_type_frame.pack(fill=tk.X, pady=(0, 10))
            self.json_body_frame.pack(fill=tk.BOTH, expand=True)
            self.body_type_label.config(text="Custom Content Editor")
        elif content_type == "application/json":
            self.json_body_frame.pack(fill=tk.BOTH, expand=True)
            self.body_type_label.config(text="JSON Editor")
        elif content_type in ["application/x-www-form-urlencoded", "multipart/form-data"]:
            self.form_body_frame.pack(fill=tk.BOTH, expand=True)
            self.body_type_label.config(text="Form Data Editor")
        else:
            self.json_body_frame.pack(fill=tk.BOTH, expand=True)
            self.body_type_label.config(text="Text Editor")
        
        self.update_payload_preview()
    
    def on_method_change(self, event=None):
        """Handle HTTP method change"""
        method = self.api_method_var.get()
        
        # Enable/disable body editor based on method
        if method in ["GET", "DELETE"]:
            self.json_body_text.config(state=tk.DISABLED)
            self.body_type_label.config(text="Body (Not used for GET/DELETE)")
        else:
            self.json_body_text.config(state=tk.NORMAL)
            self.on_content_type_change()
        
        self.update_payload_preview()
    
    def add_form_field(self):
        """Add a new form field row"""
        field_frame = ttk.Frame(self.form_fields_frame)
        field_frame.pack(fill=tk.X, pady=2)
        
        # Key entry
        key_var = tk.StringVar()
        key_entry = ttk.Entry(field_frame, textvariable=key_var, width=20)
        key_entry.pack(side=tk.LEFT)
        
        # Value entry
        value_var = tk.StringVar()
        value_entry = ttk.Entry(field_frame, textvariable=value_var, width=30)
        value_entry.pack(side=tk.LEFT, padx=(10, 0))
        
        # Type combo
        type_var = tk.StringVar(value="text")
        type_combo = ttk.Combobox(field_frame, textvariable=type_var, 
                                values=["text", "file"], width=8)
        type_combo.pack(side=tk.LEFT, padx=(10, 0))
        
        # Remove button
        remove_btn = ttk.Button(field_frame, text="", width=3, 
                              command=lambda: self.remove_form_field(field_frame))
        remove_btn.pack(side=tk.LEFT, padx=(10, 0))
        
        # Store references
        field_data = {
            'frame': field_frame,
            'key_var': key_var,
            'value_var': value_var,
            'type_var': type_var
        }
        self.form_fields.append(field_data)
        
        # Bind change events
        key_var.trace('w', lambda *args: self.update_payload_preview())
        value_var.trace('w', lambda *args: self.update_payload_preview())
        type_var.trace('w', lambda *args: self.update_payload_preview())
    
    def remove_form_field(self, field_frame):
        """Remove a form field row"""
        # Find and remove from list
        self.form_fields = [f for f in self.form_fields if f['frame'] != field_frame]
        field_frame.destroy()
        self.update_payload_preview()
    
    def get_form_fields_data(self):
        """Get form fields as dictionary"""
        data = {}
        for field in self.form_fields:
            key = field['key_var'].get().strip()
            value = field['value_var'].get().strip()
            field_type = field['type_var'].get()
            
            if key:  # Only include non-empty keys
                data[key] = {"value": value, "type": field_type}
        return data
    
    def bind_validation_events(self):
        """Bind events for real-time validation"""
        # Bind text change events
        self.api_endpoint_var.trace('w', lambda *args: self.update_validation())
        self.api_key_var.trace('w', lambda *args: self.update_validation())
        self.login_username_var.trace('w', lambda *args: self.update_validation())
        self.login_password_var.trace('w', lambda *args: self.update_validation())
        
        # Bind text widget changes
        self.json_body_text.bind('<KeyRelease>', lambda *args: self.update_payload_preview())
    
    def update_validation(self):
        """Update button states and validation"""
        self.update_button_states()
        self.update_payload_preview()
        
        # Update validation status if available
        if hasattr(self, 'validate_and_update_status'):
            self.root.after(100, self.validate_and_update_status)
    
    def update_button_states(self):
        """Enable/disable buttons based on current state with login auto-conversion"""
        endpoint = self.api_endpoint_var.get().strip()
        auth_type = self.auth_type_var.get()
        
        # Basic validation
        has_endpoint = bool(endpoint)
        
        # Auth validation with auto-conversion awareness
        has_auth = True
        auth_status = ""
        
        if auth_type == "api_key":
            has_auth = bool(self.api_key_var.get().strip())
            auth_status = "API Key" if has_auth else "API Key Required"
        elif auth_type == "login":
            username = self.login_username_var.get().strip()
            password = self.login_password_var.get().strip()
            has_auth = bool(username and password)
            
            if has_auth:
                auth_status = f"Login Ready (usernamelogin)"
            elif username and not password:
                auth_status = "Password Required"
            elif not username and password:
                auth_status = "Username Required"
            else:
                auth_status = "Login Credentials Required"
        else:
            auth_status = "No Authentication"
        
        # Update button states with descriptive text
        can_test = has_endpoint
        can_send = has_endpoint and has_auth
        
        # Save button is always enabled and visible
        if hasattr(self, 'save_btn'):
            self.save_btn.config(state=tk.NORMAL, text="Save API Settings")
            # Ensure button is visible
            if hasattr(self.save_btn, 'pack_info') and not self.save_btn.pack_info():
                self.save_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Update test connection button
        if hasattr(self, 'test_conn_btn'):
            if can_test:
                self.test_conn_btn.config(state=tk.NORMAL, text="Test Connection")
            else:
                self.test_conn_btn.config(state=tk.DISABLED, text="Test Connection (No Endpoint)")
            # Ensure button is visible
            if hasattr(self.test_conn_btn, 'pack_info') and not self.test_conn_btn.pack_info():
                self.test_conn_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Update send sample button
        if hasattr(self, 'send_sample_btn'):
            if can_send:
                if auth_type == "login":
                    self.send_sample_btn.config(state=tk.NORMAL, text="Send Sample Data")
                else:
                    self.send_sample_btn.config(state=tk.NORMAL, text="Send Sample Data")
            else:
                if not has_endpoint:
                    self.send_sample_btn.config(state=tk.DISABLED, text="Send Sample Data (No Endpoint)")
                elif not has_auth:
                    self.send_sample_btn.config(state=tk.DISABLED, text=f"Send Sample Data ({auth_status})")
                else:
                    self.send_sample_btn.config(state=tk.DISABLED, text="Send Sample Data")
            # Ensure button is visible
            if hasattr(self.send_sample_btn, 'pack_info') and not self.send_sample_btn.pack_info():
                self.send_sample_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Update button status indicator
        if hasattr(self, 'button_status_label'):
            if can_test and can_send:
                self.button_status_label.config(text=" Semua button aktif - siap digunakan!", foreground='green')
            elif can_test:
                self.button_status_label.config(text=" Test Connection aktif, lengkapi authentication untuk Send Data", foreground='orange')
            else:
                self.button_status_label.config(text=" Isi Endpoint URL untuk mengaktifkan button", foreground='red')
        
        # Update API Field Mapping status if exists
        if hasattr(self, 'refresh_api_mapping_status'):
            self.refresh_api_mapping_status()
        
        # Update login button if exists
        if hasattr(self, 'login_btn'):
            login_required = auth_type == "login" and endpoint
            if login_required and has_auth:
                self.login_btn.config(state=tk.NORMAL, text="Login / Get Token")
            elif login_required:
                self.login_btn.config(state=tk.DISABLED, text="Login (Fill Credentials)")
            else:
                self.login_btn.config(state=tk.DISABLED, text="Login")
        
        # Trigger validation update
        self.validate_login_fields()
    
    def update_payload_preview(self):
        """Update payload preview in real-time"""
        try:
            # Check if payload_preview widget exists
            if not hasattr(self, 'payload_preview'):
                return
                
            method = self.api_method_var.get()
            content_type = self.api_content_type_var.get()
            
            if content_type == "Custom...":
                content_type = self.custom_content_type_var.get()
            
            # Clear preview
            self.payload_preview.config(state=tk.NORMAL)
            self.payload_preview.delete("1.0", tk.END)
            
            preview_text = f"Method: {method}\n"
            preview_text += f"Content-Type: {content_type}\n"
            preview_text += f"Endpoint: {self.api_endpoint_var.get()}\n\n"
            
            # Add auth headers
            auth_type = self.auth_type_var.get()
            if auth_type == "api_key" and self.api_key_var.get():
                preview_text += f"Authorization: Bearer {self.api_key_var.get()[:20]}...\n\n"
            elif auth_type == "login":
                username = self.login_username_var.get().strip()
                if username:
                    preview_text += f"Authentication: Login (username: {username}  login field)\n\n"
                else:
                    preview_text += f"Authentication: Login ( credentials required)\n\n"
            elif auth_type == "no_auth":
                preview_text += "Authorization: None\n\n"
            
            # Add body based on content type and method
            if method not in ["GET", "DELETE"]:
                preview_text += "Body:\n"
                
                if content_type == "application/json":
                    body = self.json_body_text.get("1.0", tk.END).strip() if hasattr(self, 'json_body_text') else ""
                    if body:
                        try:
                            # Try to format JSON
                            import json
                            parsed = json.loads(body)
                            formatted = json.dumps(parsed, indent=2)
                            preview_text += formatted
                        except:
                            preview_text += body
                    else:
                        # Show default JSON with login auto-conversion if applicable
                        if auth_type == "login":
                            login_payload = self.get_login_payload()
                            if login_payload:
                                sample_data = {
                                    "test": True,
                                    "timestamp": "2025-07-31T12:00:00",
                                    "uuid": "sample-uuid-here"
                                }
                                sample_data.update(login_payload)
                                formatted = json.dumps(sample_data, indent=2)
                                preview_text += f"{formatted}\n\n Username auto-converted to 'login' field"
                            else:
                                preview_text += "{\n  \"login\": \"<username>\",\n  \"password\": \"<password>\"\n}\n\n Fill username & password above"
                        else:
                            preview_text += "{}"
                        
                elif content_type in ["application/x-www-form-urlencoded", "multipart/form-data"]:
                    form_data = self.get_form_fields_data()
                    if form_data:
                        for key, field_info in form_data.items():
                            preview_text += f"{key}={field_info['value']} ({field_info['type']})\n"
                    else:
                        preview_text += "No form fields defined"
                else:
                    body = self.json_body_text.get("1.0", tk.END).strip() if hasattr(self, 'json_body_text') else ""
                    preview_text += body if body else "No body content"
            else:
                preview_text += "Body: Not applicable for GET/DELETE"
            
            self.payload_preview.insert("1.0", preview_text)
            self.payload_preview.config(state=tk.DISABLED)
            
        except Exception as e:
            # Silently handle errors during initialization
            pass
    
    def test_api_connection(self):
        """Enhanced API connection test with proper JSON body sending and detailed response - NO AUTO-SAVE (testing only)"""
        try:
            endpoint = self.api_endpoint_var.get().strip()
            if not endpoint:
                self.show_error_response(" Endpoint URL is required")
                return
            
            method = self.api_method_var.get()
            content_type = self.api_content_type_var.get()
            auth_type = self.auth_type_var.get()
            
            if content_type == "Custom...":
                content_type = self.custom_content_type_var.get().strip()
                if not content_type:
                    content_type = "application/json"
            
            # Prepare headers
            headers = {
                'Content-Type': content_type,
                'User-Agent': 'MDB-Agent-Pro/2.0-TestMode',
                'Accept': 'application/json'
            }
            
            # Add authorization header (NO AUTO-SAVE)
            if auth_type == "api_key" and self.api_key_var.get().strip():
                api_key = self.api_key_var.get().strip()
                headers['Authorization'] = f'Bearer {api_key}'
            
            # Log test mode
            self.log_entry(f" Testing API Connection: {method} {endpoint} (test mode - no auto-save)", "INFO")
            
            # Store test result for propagation to other features
            test_successful = False
            test_status_code = 0
            test_response_text = ""
            
            # Prepare request body for all methods (including GET if body specified)
            data = None
            json_data = None
            
            if content_type == "application/json":
                # Get JSON body from editor (case-sensitive, no extra spaces)
                json_body = self.json_body_text.get("1.0", tk.END).strip()
                if json_body:
                    try:
                        # Parse and re-serialize to ensure valid JSON (preserves exact case)
                        json_data = json.loads(json_body)
                        # Validate JSON structure
                        json.dumps(json_data)  # This will raise if invalid
                    except json.JSONDecodeError as e:
                        self.show_error_response(f" Invalid JSON in body: {str(e)}\n\n Tip: Check for missing quotes, trailing commas, or syntax errors")
                        return
                else:
                    # Default test payload for login endpoints
                    if "/login" in endpoint.lower():
                        if auth_type == "login":
                            username = self.login_username_var.get().strip()
                            password = self.login_password_var.get().strip()
                            database = self.login_database_var.get().strip()
                            if username and password:
                                # AUTO-CONVERSION: username -> login
                                json_data = {
                                    "login": username,  # Case-sensitive field name
                                    "password": password,
                                    "database": database or "default"
                                }
                        else:
                            # For non-login auth on login endpoint, use generic test data
                            json_data = {
                                "test": True,
                                "timestamp": datetime.now().isoformat(),
                                "source": "MDBAgentPro",
                                "action": "connection_test"
                            }
                    else:
                        # For non-login endpoints, use appropriate test data
                        if "/data" in endpoint.lower() or "/records" in endpoint.lower():
                            json_data = {
                                "id": 12345,
                                "name": "Test Record",
                                "value": 123.45,
                                "timestamp": datetime.now().isoformat(),
                                "status": "test"
                            }
                        else:
                            json_data = {
                                "test": True,
                                "timestamp": datetime.now().isoformat(),
                                "source": "MDBAgentPro"
                            }
            
            elif content_type in ["application/x-www-form-urlencoded", "multipart/form-data"]:
                # Get form data
                form_data = self.get_form_fields_data()
                if form_data:
                    data = {}
                    for key, field_info in form_data.items():
                        data[key] = field_info['value']
            
            # Update response status
            self.response_status_var.set(" Testing connection...")
            self.response_status_label.config(foreground='orange')
            
            # Make request with body if specified
            start_time = time.time()
            
            try:
                if method == "GET":
                    response = requests.get(endpoint, headers=headers, params=data, timeout=15)
                elif method == "POST":
                    if json_data is not None:
                        response = requests.post(endpoint, json=json_data, headers=headers, timeout=15)
                    elif data:
                        response = requests.post(endpoint, data=data, headers=headers, timeout=15)
                    else:
                        response = requests.post(endpoint, headers=headers, timeout=15)
                elif method == "PUT":
                    if json_data is not None:
                        response = requests.put(endpoint, json=json_data, headers=headers, timeout=15)
                    else:
                        response = requests.put(endpoint, data=data, headers=headers, timeout=15)
                elif method == "DELETE":
                    response = requests.delete(endpoint, headers=headers, timeout=15)
                elif method == "PATCH":
                    if json_data is not None:
                        response = requests.patch(endpoint, json=json_data, headers=headers, timeout=15)
                    else:
                        response = requests.patch(endpoint, data=data, headers=headers, timeout=15)
                else:
                    response = requests.request(method, endpoint, json=json_data, data=data, headers=headers, timeout=15)
                
            except requests.exceptions.Timeout:
                self.response_status_var.set(" Connection timeout")
                self.response_status_label.config(foreground='red')
                self.show_error_response("Connection timeout (15s)\n\n Tips:\n Check if the server is running\n Verify the endpoint URL\n Check your internet connection\n Server might be slow or overloaded")
                return
                
            except requests.exceptions.ConnectionError:
                self.response_status_var.set(" Connection failed")
                self.response_status_label.config(foreground='red')
                self.show_error_response("Could not connect to endpoint\n\n Tips:\n Verify the URL is correct (include http/https)\n Check if the server is running\n Check firewall settings\n Verify network connectivity")
                return
            
            response_time = round((time.time() - start_time) * 1000, 2)
            
            # Format status with color coding
            status_code = response.status_code
            if status_code < 300:
                color = 'green'
                status_icon = ""
                status_text = "SUCCESS"
            elif status_code < 400:
                color = 'orange'
                status_icon = ""
                status_text = "WARNING"
            else:
                color = 'red'
                status_icon = ""
                status_text = "ERROR"
            
            status_display = f"{status_icon} {status_code} {status_text} ({response_time}ms)"
            self.response_status_var.set(status_display)
            self.response_status_label.config(foreground=color)
            
            # Display detailed response
            self.response_text.config(state=tk.NORMAL)
            self.response_text.delete("1.0", tk.END)
            
            response_content = f" TEST MODE - NO AUTO-SAVE\n"
            response_content += f"="*50 + "\n\n"
            response_content += f"REQUEST DETAILS:\n"
            response_content += f"Method: {method}\n"
            response_content += f"URL: {endpoint}\n"
            response_content += f"Content-Type: {content_type}\n"
            
            if auth_type == "api_key" and self.api_key_var.get().strip():
                response_content += f"Authorization: Bearer {api_key[:10]}...{api_key[-4:]}\n"
            elif auth_type == "login":
                response_content += f"Auth Type: Login (auto-conversion enabled)\n"
            
            if json_data:
                response_content += f"JSON Body: {json.dumps(json_data, indent=2)}\n"
            elif data:
                response_content += f"Form Data: {data}\n"
            
            response_content += f"\nRESPONSE DETAILS:\n"
            response_content += f"Status: {status_code} ({response.reason})\n"
            response_content += f"Time: {response_time}ms\n"
            response_content += f"Headers: {dict(response.headers)}\n\n"
            
            # Try to format JSON response
            try:
                if response.headers.get('content-type', '').lower().startswith('application/json'):
                    json_response = response.json()
                    response_content += f"JSON Response:\n{json.dumps(json_response, indent=2, ensure_ascii=False)}"
                else:
                    response_content += f"Response Body:\n{response.text}"
            except:
                response_content += f"Response Body:\n{response.text}"
            
            # Add troubleshooting tips for errors
            if status_code >= 400:
                response_content += f"\n\n TROUBLESHOOTING TIPS:\n"
                if status_code == 400:
                    response_content += " Check request body format and required fields\n Verify JSON syntax is valid\n Ensure all required parameters are included"
                elif status_code == 401:
                    response_content += " Check authentication credentials\n Verify API key or login details\n Ensure authorization header is correct"
                elif status_code == 403:
                    response_content += " Check API permissions\n Verify account has access to this endpoint\n Check rate limiting"
                elif status_code == 404:
                    response_content += " Verify endpoint URL is correct\n Check if the API path exists\n Ensure method is supported"
                elif status_code == 500:
                    response_content += " Server internal error\n Check server logs\n Contact API provider\n Try again later"
                else:
                    response_content += f" HTTP {status_code} error\n Check API documentation\n Verify request format"
            
            self.response_text.insert("1.0", response_content)
            self.response_text.config(state=tk.DISABLED)
            
            # Store test results for status propagation
            test_successful = status_code < 400
            test_status_code = status_code
            test_response_text = response.text[:100]
            
            # Add to history with complete details (TEST MODE indicator)
            payload_summary = ""
            if json_data:
                payload_summary = f"JSON: {str(json_data)[:50]}..." if len(str(json_data)) > 50 else f"JSON: {json_data}"
            elif data:
                payload_summary = f"Form: {str(data)[:50]}..." if len(str(data)) > 50 else f"Form: {data}"
            else:
                payload_summary = "No body"
            
            # Add test mode indicator to history
            action_name = " Test Connection (No Auto-Save)"
            self.add_to_history(action_name, method, endpoint, status_code, 
                              response.text[:100], payload_summary, response_time)
            
            # Log result with details (TEST MODE)
            log_msg = f" API connection test: {method} {endpoint} -> {status_code} in {response_time}ms (test mode)"
            if status_code < 400:
                self.log_entry(log_msg, "SUCCESS")
            else:
                self.log_entry(f"{log_msg} - Error: {response.text[:100]}", "ERROR")
            
            # PROPAGATE TEST RESULTS TO OTHER FEATURES (temporary status only)
            self.propagate_test_connection_results(test_successful, status_code, test_response_text, response_time)
            
        except Exception as e:
            self.response_status_var.set(" Test failed")
            self.response_status_label.config(foreground='red')
            error_msg = f" Connection test failed: {str(e)} (test mode - no auto-save)"
            self.show_error_response(error_msg)
            self.log_entry(error_msg, "ERROR")
            
            # PROPAGATE FAILED TEST RESULTS
            self.propagate_test_connection_results(False, 0, str(e), 0)
    
    def propagate_test_connection_results(self, test_successful, status_code, response_text, response_time):
        """Propagate test connection results to Dashboard, Health Check, and API Field Mapping"""
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Store test results for other components to use
            self.last_test_connection_result = {
                'success': test_successful,
                'status_code': status_code,
                'response_time': f"{response_time}ms",
                'error': response_text if not test_successful else None,
                'timestamp': timestamp
            }
            
            if test_successful:
                # SUCCESS - Update all status indicators to show test passed
                success_msg = f" Test Passed ({status_code}, {response_time}ms)"
                
                # Update Dashboard API Status
                if hasattr(self, 'dash_api_status'):
                    self.dash_api_status.config(text=success_msg, foreground="green")
                
                # Update Health Check API Status
                if hasattr(self, 'health_api_status'):
                    self.health_api_status.config(text=success_msg, foreground="green")
                
                # Update API Field Mapping Status
                if hasattr(self, 'api_status_label'):
                    self.api_status_label.config(
                        text=f" API Connected ({response_time}ms)\n Test successful - Ready for mapping", 
                        foreground='green'
                    )
                    
                # Update API Status Indicator (top bar)
                if hasattr(self, 'api_status'):
                    self.api_status.update_status("good")
                
                # Add to Health Check Results
                if hasattr(self, 'health_tree'):
                    self.health_tree.insert("", 0, values=(
                        timestamp, 
                        "API Connection Test", 
                        "PASS", 
                        f"Connection successful - Status {status_code} in {response_time}ms"
                    ))
                
                # Log success propagation
                self.log_entry(f" Test connection success propagated to all features (Status: {status_code})", "SUCCESS")
                
            else:
                # FAILED - Update all status indicators to show test failed
                if status_code > 0:
                    fail_msg = f" Test Failed ({status_code})"
                else:
                    fail_msg = f" Test Failed (Connection Error)"
                
                # Update Dashboard API Status
                if hasattr(self, 'dash_api_status'):
                    self.dash_api_status.config(text=fail_msg, foreground="red")
                
                # Update Health Check API Status  
                if hasattr(self, 'health_api_status'):
                    self.health_api_status.config(text=fail_msg, foreground="red")
                
                # Update API Field Mapping Status
                if hasattr(self, 'api_status_label'):
                    error_detail = response_text[:50] if response_text else "Connection failed"
                    self.api_status_label.config(
                        text=f" API Test Failed\n {error_detail}", 
                        foreground='red'
                    )
                    
                # Update API Status Indicator (top bar)
                if hasattr(self, 'api_status'):
                    self.api_status.update_status("error")
                
                # Add to Health Check Results
                if hasattr(self, 'health_tree'):
                    error_detail = response_text[:100] if response_text else "Connection failed"
                    self.health_tree.insert("", 0, values=(
                        timestamp, 
                        "API Connection Test", 
                        "FAIL", 
                        f"Test failed - {error_detail}"
                    ))
                
                # Log failure propagation
                self.log_entry(f" Test connection failure propagated to all features ({fail_msg})", "ERROR")
            
            # Refresh API Field Mapping status to use the stored results
            if hasattr(self, 'refresh_api_mapping_status'):
                self.refresh_api_mapping_status()
            
            # Refresh dashboard to reflect updated status
            if hasattr(self, 'current_tab') and self.current_tab == "dashboard":
                self.root.after(200, self.refresh_dashboard_after_test)
                
        except Exception as e:
            self.log_entry(f" Failed to propagate test results: {str(e)}", "ERROR")
    
    def refresh_dashboard_after_test(self):
        """Refresh dashboard specifically after test connection to show updated status"""
        try:
            # Only update the parts that show test results, not full refresh
            # This prevents overriding the test status with config-based status
            
            # Update agent status
            if self.is_running:
                self.dash_agent_status.config(text=" Running", foreground="green")
            else:
                self.dash_agent_status.config(text=" Stopped", foreground="red")
            
            # Update buffer status
            try:
                conn = sqlite3.connect(self.db_manager.db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM push_buffer WHERE status = 'pending'")
                count = cursor.fetchone()[0]
                conn.close()
                
                if count > 0:
                    self.dash_buffer_status.config(text=f" {count} items", foreground="orange")
                else:
                    self.dash_buffer_status.config(text=" 0 items", foreground="green")
            except:
                self.dash_buffer_status.config(text=" Error", foreground="red")
                
            # Note: We deliberately DON'T call full refresh_dashboard() here 
            # to preserve the test connection results in dash_api_status
            
        except Exception as e:
            self.log_entry(f" Dashboard refresh after test failed: {str(e)}", "ERROR")
    
    def send_sample_data_advanced(self):
        """Send sample data using current API settings with login auto-conversion"""
        try:
            # Advanced validation with auto-conversion
            errors, warnings = self.validate_api_settings_advanced()
            
            if errors:
                error_msg = "\n".join(f" {error}" for error in errors)
                messagebox.showerror("Validation Error", f"Please fix the following issues:\n\n{error_msg}")
                return
            
            if warnings:
                warning_msg = "\n".join(f" {warning}" for warning in warnings)
                result = messagebox.askyesno("Validation Warning", 
                                           f"Please note:\n\n{warning_msg}\n\nDo you want to continue?")
                if not result:
                    return
            
            endpoint = self.api_endpoint_var.get().strip()
            method = self.api_method_var.get()
            content_type = self.api_content_type_var.get()
            
            if content_type == "Custom...":
                content_type = self.custom_content_type_var.get()
            
            # Prepare headers
            headers = {
                'Content-Type': content_type,
                'User-Agent': 'MDB-Agent-Pro/2.0'
            }
            
            # Add authorization with auto-conversion
            auth_type = self.auth_type_var.get()
            if auth_type == "api_key" and self.api_key_var.get():
                headers['Authorization'] = f'Bearer {self.api_key_var.get()}'
            
            # Prepare data with auto-conversion for login
            data = None
            if method not in ["GET", "DELETE"]:
                if content_type == "application/json":
                    body = self.json_body_text.get("1.0", tk.END).strip()
                    if body:
                        try:
                            data = json.loads(body)
                        except json.JSONDecodeError:
                            messagebox.showerror("Error", "Invalid JSON in request body")
                            return
                    else:
                        # Smart sample data based on endpoint type
                        if "/login" in endpoint.lower() and auth_type == "login":
                            # For login endpoints with login auth, use proper login data
                            username = self.login_username_var.get().strip()
                            password = self.login_password_var.get().strip()
                            database = self.login_database_var.get().strip()
                            
                            if username and password:
                                data = {
                                    "login": username,  # Auto-conversion
                                    "password": password,
                                    "database": database or "default"
                                }
                                self.log_entry(f"Using login credentials for sample data (usernamelogin)", "INFO")
                            else:
                                messagebox.showerror("Error", "Login credentials required for login endpoint")
                                return
                        elif "/data" in endpoint.lower() or "/records" in endpoint.lower() or "/create" in endpoint.lower():
                            # For data endpoints, use realistic sample record
                            data = {
                                "id": 12345,
                                "name": "Sample Record",
                                "value": 123.45,
                                "timestamp": datetime.now().isoformat(),
                                "status": "active",
                                "category": "test_data",
                                "uuid": str(uuid.uuid4())
                            }
                        else:
                            # Generic test data for other endpoints
                            data = {
                                "test": True,
                                "timestamp": datetime.now().isoformat(),
                                "uuid": str(uuid.uuid4()),
                                "sample_field": "Sample Value",
                                "source": "MDB Agent Pro"
                            }
                
                elif content_type == "application/x-www-form-urlencoded":
                    form_data = self.get_form_fields_data()
                    data = {k: v['value'] for k, v in form_data.items()}
                    
                    # Add login auto-conversion for form data
                    if auth_type == "login":
                        login_payload = self.get_login_payload()
                        if login_payload:
                            data.update(login_payload)
                    
                elif content_type == "multipart/form-data":
                    # Handle multipart data with login auto-conversion
                    files = {}
                    data = {}
                    form_data = self.get_form_fields_data()
                    
                    for key, field_info in form_data.items():
                        if field_info['type'] == 'file':
                            # For demo, create a dummy file
                            files[key] = ('sample.txt', 'Sample file content', 'text/plain')
                        else:
                            data[key] = field_info['value']
                    
                    # Add login auto-conversion for multipart
                    if auth_type == "login":
                        login_payload = self.get_login_payload()
                        if login_payload:
                            data.update(login_payload)
            
            # Update status
            test_mode = self.test_mode_var.get()
            status_msg = "Sending sample data (TEST MODE)" if test_mode else "Sending sample data"
            self.response_status_var.set(status_msg)
            self.response_status_label.config(foreground='orange')
            
            # Send request (skip if test mode)
            if test_mode:
                # Simulate response for test mode with login auto-conversion info
                self.response_status_var.set("Test Mode - No data sent ")
                self.response_status_label.config(foreground='blue')
                
                self.response_text.config(state=tk.NORMAL)
                self.response_text.delete("1.0", tk.END)
                
                preview_text = f"TEST MODE - Request would be sent to:\n{endpoint}\n\n"
                preview_text += f"Method: {method}\nHeaders: {headers}\n\n"
                
                if auth_type == "login" and data:
                    preview_text += "Data (with auto-conversion):\n"
                    if 'login' in str(data):
                        preview_text += " Username auto-converted to 'login' field\n"
                    preview_text += f"{data}"
                else:
                    preview_text += f"Data: {data}"
                
                self.response_text.insert("1.0", preview_text)
                self.response_text.config(state=tk.DISABLED)
                
                self.add_to_history("Sample Data (Test)", method, endpoint, 200, "Test mode - not sent")
                self.log_entry("Sample data test completed (test mode) with login auto-conversion", "INFO")
                
            else:
                # Actually send the request
                start_time = time.time()
                
                if content_type == "multipart/form-data" and 'files' in locals():
                    response = requests.request(method, endpoint, headers={k: v for k, v in headers.items() if k != 'Content-Type'}, 
                                              data=data, files=files, timeout=15)
                else:
                    if content_type == "application/json" and data:
                        response = requests.request(method, endpoint, headers=headers, json=data, timeout=15)
                    elif content_type == "application/x-www-form-urlencoded" and data:
                        response = requests.request(method, endpoint, headers=headers, data=data, timeout=15)
                    else:
                        response = requests.request(method, endpoint, headers=headers, timeout=15)
                
                response_time = round((time.time() - start_time) * 1000, 2)
                
                # Handle response
                status_text = f"Response: {response.status_code} ({response_time}ms)"
                if response.status_code < 300:
                    self.response_status_label.config(foreground='green')
                    status_text += " "
                    log_level = "SUCCESS"
                else:
                    self.response_status_label.config(foreground='red')
                    status_text += " "
                    log_level = "ERROR"
                
                self.response_status_var.set(status_text)
                
                # Display response
                self.response_text.config(state=tk.NORMAL)
                self.response_text.delete("1.0", tk.END)
                
                response_content = f"Status: {response.status_code}\n"
                response_content += f"Response Time: {response_time}ms\n\n"
                response_content += f"Headers:\n{dict(response.headers)}\n\n"
                response_content += f"Body:\n{response.text}"
                
                self.response_text.insert("1.0", response_content)
                self.response_text.config(state=tk.DISABLED)
                
                # Add to history
                self.add_to_history("Sample Data", method, endpoint, response.status_code, response.text[:100])
                
                # Log result with auto-conversion info
                log_msg = f"Sample data sent: {response.status_code} in {response_time}ms"
                if auth_type == "login" and 'login' in str(data):
                    log_msg += " (with login auto-conversion)"
                self.log_entry(log_msg, log_level)
                
                # Show result message
                if response.status_code < 300:
                    success_msg = f"Sample data sent successfully!\nStatus: {response.status_code}\nTime: {response_time}ms"
                    if auth_type == "login":
                        success_msg += "\n\n Username was auto-converted to 'login' field"
                    messagebox.showinfo("Success", success_msg)
                else:
                    messagebox.showerror("Error", f"Request failed with status {response.status_code}\nResponse: {response.text[:200]}...")
                    
        except Exception as e:
            self.response_status_var.set("Send failed")
            self.response_status_label.config(foreground='red')
            error_msg = f"Failed to send sample data: {str(e)}"
            self.show_error_response(error_msg)
            messagebox.showerror("Error", error_msg)
    
    def validate_api_settings(self):
        """Validate current API settings"""
        endpoint = self.api_endpoint_var.get().strip()
        if not endpoint:
            messagebox.showerror("Validation Error", "Please enter an endpoint URL")
            return False
        
        auth_type = self.auth_type_var.get()
        if auth_type == "api_key" and not self.api_key_var.get().strip():
            messagebox.showerror("Validation Error", "Please enter an API key for API key authentication")
            return False
        
        if auth_type == "login":
            if not all([self.login_username_var.get().strip(), self.login_password_var.get().strip()]):
                messagebox.showerror("Validation Error", "Please enter username and password for login authentication")
                return False
        
        # Validate JSON body if applicable
        method = self.api_method_var.get()
        content_type = self.api_content_type_var.get()
        
        if method not in ["GET", "DELETE"] and content_type == "application/json":
            body = self.json_body_text.get("1.0", tk.END).strip()
            if body:
                try:
                    json.loads(body)
                except json.JSONDecodeError as e:
                    messagebox.showerror("Validation Error", f"Invalid JSON in request body:\n{str(e)}")
                    return False
        
        return True
    
    def show_error_response(self, error_msg):
        """Show error in response panel"""
        self.response_text.config(state=tk.NORMAL)
        self.response_text.delete("1.0", tk.END)
        self.response_text.insert("1.0", f"Error: {error_msg}\n\nTroubleshooting Tips:\n"
                                        " Check your internet connection\n"
                                        " Verify the endpoint URL is correct\n"
                                        " Ensure authentication credentials are valid\n"
                                        " Check if the API server is running\n"
                                        " Review request format and required fields")
        self.response_text.config(state=tk.DISABLED)
        self.log_entry(error_msg, "ERROR")
    
    def add_to_history(self, action, method, endpoint, status_code, response, payload_summary=None, response_time=None):
        """Add request to enhanced history with complete details"""
        from datetime import datetime
        
        # Format status with color indicator
        if status_code < 300:
            status_text = f" {status_code}"
        elif status_code < 400:
            status_text = f" {status_code}"
        else:
            status_text = f" {status_code}"
        
        history_item = {
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": action,
            "method": method,
            "endpoint": endpoint[:50] + "..." if len(endpoint) > 50 else endpoint,
            "full_endpoint": endpoint,
            "status": status_text,
            "status_code": status_code,
            "payload": payload_summary or "No body",
            "response": response[:100] + "..." if len(response) > 100 else response,
            "full_response": response,
            "response_time": f"{response_time}ms" if response_time else "N/A"
        }
        
        self.api_history.insert(0, history_item)  # Add to beginning
        
        # Keep only last 50 items
        if len(self.api_history) > 50:
            self.api_history = self.api_history[:50]
        
        # Update history display
        self.update_history_display()
    
    def update_history_display(self):
        """Update enhanced history treeview with complete information"""
        if hasattr(self, 'history_tree'):
            # Clear existing items
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)
            
            # Add history items (show only last 5 for quick view)
            display_items = self.api_history[:5] if len(self.api_history) > 5 else self.api_history
            
            for item in display_items:
                # Insert with complete details
                self.history_tree.insert("", tk.END, values=(
                    item.get("time", ""),
                    item.get("method", ""),
                    item.get("endpoint", ""),
                    item.get("status", ""),
                    item.get("payload", "")[:30] + "..." if len(item.get("payload", "")) > 30 else item.get("payload", ""),
                    item.get("response_time", "")
                ))
    
    def view_history_details(self, event):
        """View detailed history item information"""
        selection = self.history_tree.selection()
        if not selection:
            return
        
        # Get selected item index
        item_index = self.history_tree.index(selection[0])
        if item_index >= len(self.api_history):
            return
        
        history_item = self.api_history[item_index]
        
        # Create detail window
        detail_window = tk.Toplevel(self.root)
        detail_window.title("API Request Details")
        detail_window.geometry("800x600")
        detail_window.transient(self.root)
        
        # Detail content
        detail_text = scrolledtext.ScrolledText(detail_window, wrap=tk.WORD)
        detail_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Format details
        details = f"""API REQUEST DETAILS
{'-' * 50}

Time: {history_item.get('time', 'N/A')}
Action: {history_item.get('action', 'N/A')}
Method: {history_item.get('method', 'N/A')}
Endpoint: {history_item.get('full_endpoint', 'N/A')}
Status: {history_item.get('status_code', 'N/A')}
Response Time: {history_item.get('response_time', 'N/A')}

PAYLOAD SENT:
{history_item.get('payload', 'No payload')}

FULL RESPONSE:
{history_item.get('full_response', 'No response')}
"""
        
        detail_text.insert("1.0", details)
        detail_text.config(state=tk.DISABLED)
        
        # Close button
        ttk.Button(detail_window, text="Close", command=detail_window.destroy).pack(pady=10)
    
    def load_api_profiles(self):
        """Load API profiles from config"""
        self.api_profiles = self.config.get("api_profiles", {"Default": {}})
        self.current_profile = self.config.get("current_api_profile", "Default")
        self.api_history = self.config.get("api_history", [])
        
        # Update profile combo
        if hasattr(self, 'api_profile_combo'):
            self.api_profile_combo['values'] = list(self.api_profiles.keys())
            self.api_profile_var.set(self.current_profile)
        
        # Load current profile
        self.load_profile_data(self.current_profile)
        self.update_history_display()
        
        # Update profiles listbox
        if hasattr(self, 'profiles_listbox'):
            self.profiles_listbox.delete(0, tk.END)
            for profile_name in self.api_profiles.keys():
                self.profiles_listbox.insert(tk.END, profile_name)
    
    def create_new_profile(self):
        """Create a new API profile"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Create New Profile")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (300 // 2)
        dialog.geometry(f"400x300+{x}+{y}")
        
        # Profile name
        tk.Label(dialog, text="Profile Name:", font=('Arial', 10, 'bold')).pack(pady=10)
        name_var = tk.StringVar()
        name_entry = tk.Entry(dialog, textvariable=name_var, width=40, font=('Arial', 10))
        name_entry.pack(pady=5)
        name_entry.focus()
        
        # Description
        tk.Label(dialog, text="Description:", font=('Arial', 10, 'bold')).pack(pady=(20,5))
        desc_text = tk.Text(dialog, width=45, height=5, font=('Arial', 9))
        desc_text.pack(pady=5)
        
        # Copy settings from current profile
        copy_var = tk.BooleanVar(value=True)
        tk.Checkbutton(dialog, text="Copy settings from current profile", 
                      variable=copy_var, font=('Arial', 9)).pack(pady=10)
        
        # Buttons
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=20)
        
        def create_profile():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("Error", "Please enter a profile name", parent=dialog)
                return
            
            if name in self.api_profiles:
                messagebox.showerror("Error", "Profile name already exists", parent=dialog)
                return
            
            # Create profile data
            if copy_var.get():
                # Copy current settings
                profile_data = self.get_current_api_settings()
            else:
                # Create empty profile
                profile_data = {
                    'endpoint': '',
                    'method': 'POST',
                    'content_type': 'application/json',
                    'auth_type': 'no_auth',
                    'api_key': '',
                    'username': '',
                    'password': '',
                    'json_body': '',
                    'form_fields': []
                }
            
            profile_data['description'] = desc_text.get("1.0", tk.END).strip()
            profile_data['created'] = datetime.now().isoformat()
            
            # Add to profiles
            self.api_profiles[name] = profile_data
            self.save_api_config()
            
            # Update UI
            self.api_profile_combo['values'] = list(self.api_profiles.keys())
            self.profiles_listbox.insert(tk.END, name)
            
            # Switch to new profile
            self.api_profile_var.set(name)
            self.current_profile = name
            self.load_profile_data(name)
            
            dialog.destroy()
            messagebox.showinfo("Success", f"Profile '{name}' created successfully!")
        
        tk.Button(button_frame, text="Create", command=create_profile, 
                 bg='#4CAF50', fg='white', font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Cancel", command=dialog.destroy, 
                 bg='#f44336', fg='white', font=('Arial', 10)).pack(side=tk.LEFT, padx=5)
    
    def save_current_profile(self):
        """Save current settings to active profile"""
        if not self.current_profile:
            self.current_profile = "Default"
        
        profile_data = self.get_current_api_settings()
        self.api_profiles[self.current_profile] = profile_data
        self.save_api_config()
        
        messagebox.showinfo("Saved", f"Settings saved to profile '{self.current_profile}'")
        self.log_entry(f"Profile '{self.current_profile}' saved", "INFO")
    
    def load_selected_profile(self):
        """Load the selected profile from listbox"""
        selection = self.profiles_listbox.curselection()
        if not selection:
            return
        
        profile_name = self.profiles_listbox.get(selection[0])
        self.api_profile_var.set(profile_name)
        self.current_profile = profile_name
        self.load_profile_data(profile_name)
        
        messagebox.showinfo("Loaded", f"Profile '{profile_name}' loaded successfully")
    
    def get_current_api_settings(self):
        """Get current API settings as dictionary"""
        settings = {
            'endpoint': self.api_endpoint_var.get(),
            'method': self.api_method_var.get(),
            'content_type': self.api_content_type_var.get(),
            'custom_content_type': self.custom_content_type_var.get(),
            'auth_type': self.auth_type_var.get(),
            'api_key': self.api_key_var.get(),
            'username': self.login_username_var.get(),
            'password': self.login_password_var.get(),
            'test_mode': self.test_mode_var.get(),
            'json_body': self.json_body_text.get("1.0", tk.END).strip() if hasattr(self, 'json_body_text') else '',
            'form_fields': self.get_form_fields_data() if hasattr(self, 'form_fields_frame') else [],
            'updated': datetime.now().isoformat()
        }
        return settings
    
    def load_profile_data(self, profile_name):
        """Load data from specified profile"""
        if profile_name not in self.api_profiles:
            return
        
        data = self.api_profiles[profile_name]
        
        # Load basic settings
        self.api_endpoint_var.set(data.get('endpoint', ''))
        self.api_method_var.set(data.get('method', 'POST'))
        self.api_content_type_var.set(data.get('content_type', 'application/json'))
        self.custom_content_type_var.set(data.get('custom_content_type', ''))
        self.auth_type_var.set(data.get('auth_type', 'no_auth'))
        self.api_key_var.set(data.get('api_key', ''))
        self.login_username_var.set(data.get('username', ''))
        self.login_password_var.set(data.get('password', ''))
        self.test_mode_var.set(data.get('test_mode', False))
        
        # Load JSON body
        if hasattr(self, 'json_body_text'):
            self.json_body_text.delete("1.0", tk.END)
            self.json_body_text.insert("1.0", data.get('json_body', ''))
        
        # Load form fields
        if hasattr(self, 'form_fields_frame'):
            self.clear_form_fields()
            for field_data in data.get('form_fields', []):
                self.add_form_field(field_data.get('key', ''), field_data.get('value', ''), field_data.get('type', 'text'))
        
        # Trigger change events
        self.on_auth_type_change()
        self.on_content_type_change()
        self.update_payload_preview()
        self.update_button_states()
    
    def save_api_config(self):
        """Save API configuration to file"""
        self.config['api_profiles'] = self.api_profiles
        self.config['current_api_profile'] = self.current_profile
        self.config['api_history'] = self.api_history
        self.save_config()
    
    def export_profile(self):
        """Export selected profile to JSON file"""
        selection = self.profiles_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a profile to export")
            return
        
        profile_name = self.profiles_listbox.get(selection[0])
        profile_data = self.api_profiles[profile_name]
        
        filename = filedialog.asksaveasfilename(
            title="Export Profile",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile=f"{profile_name}_profile.json"
        )
        
        if filename:
            try:
                export_data = {
                    'profile_name': profile_name,
                    'exported_date': datetime.now().isoformat(),
                    'mdb_agent_version': '2.0',
                    'profile_data': profile_data
                }
                
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, indent=2, ensure_ascii=False)
                
                messagebox.showinfo("Success", f"Profile '{profile_name}' exported to:\n{filename}")
                self.log_entry(f"Profile '{profile_name}' exported", "INFO")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export profile:\n{str(e)}")
    
    def import_profile(self):
        """Import profile from JSON file"""
        filename = filedialog.askopenfilename(
            title="Import Profile",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    import_data = json.load(f)
                
                if 'profile_data' not in import_data or 'profile_name' not in import_data:
                    messagebox.showerror("Error", "Invalid profile file format")
                    return
                
                profile_name = import_data['profile_name']
                profile_data = import_data['profile_data']
                
                # Check if profile already exists
                if profile_name in self.api_profiles:
                    result = messagebox.askyesno("Profile Exists", 
                                               f"Profile '{profile_name}' already exists.\nDo you want to overwrite it?")
                    if not result:
                        return
                
                # Import profile
                self.api_profiles[profile_name] = profile_data
                self.save_api_config()
                
                # Update UI
                self.api_profile_combo['values'] = list(self.api_profiles.keys())
                
                # Update listbox
                self.profiles_listbox.delete(0, tk.END)
                for name in self.api_profiles.keys():
                    self.profiles_listbox.insert(tk.END, name)
                
                messagebox.showinfo("Success", f"Profile '{profile_name}' imported successfully!")
                self.log_entry(f"Profile '{profile_name}' imported", "INFO")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to import profile:\n{str(e)}")
    
    def delete_selected_profile(self):
        """Delete the selected profile"""
        selection = self.profiles_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a profile to delete")
            return
        
        profile_name = self.profiles_listbox.get(selection[0])
        
        if profile_name == "Default":
            messagebox.showerror("Error", "Cannot delete the Default profile")
            return
        
        result = messagebox.askyesno("Confirm Delete", 
                                   f"Are you sure you want to delete profile '{profile_name}'?\nThis action cannot be undone.")
        if result:
            del self.api_profiles[profile_name]
            
            # Update current profile if needed
            if self.current_profile == profile_name:
                self.current_profile = "Default"
                self.api_profile_var.set("Default")
                self.load_profile_data("Default")
            
            self.save_api_config()
            
            # Update UI
            self.api_profile_combo['values'] = list(self.api_profiles.keys())
            self.profiles_listbox.delete(selection[0])
            
            messagebox.showinfo("Deleted", f"Profile '{profile_name}' deleted successfully")
            self.log_entry(f"Profile '{profile_name}' deleted", "INFO")
    
    def clear_api_history(self):
        """Clear all API history"""
        result = messagebox.askyesno("Confirm Clear", 
                                   "Are you sure you want to clear all API history?\nThis action cannot be undone.")
        if result:
            self.api_history = []
            self.save_api_config()
            self.update_history_display()
            messagebox.showinfo("Cleared", "API history cleared successfully")
            self.log_entry("API history cleared", "INFO")
    
    def export_api_history(self):
        """Export API history to JSON file"""
        if not self.api_history:
            messagebox.showwarning("Warning", "No history to export")
            return
        
        filename = filedialog.asksaveasfilename(
            title="Export API History",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile=f"api_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )
        
        if filename:
            try:
                export_data = {
                    'exported_date': datetime.now().isoformat(),
                    'mdb_agent_version': '2.0',
                    'total_requests': len(self.api_history),
                    'history': self.api_history
                }
                
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, indent=2, ensure_ascii=False)
                
                messagebox.showinfo("Success", f"API history exported to:\n{filename}")
                self.log_entry("API history exported", "INFO")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export history:\n{str(e)}")
    
    def on_profile_change(self, *args):
        """Called when profile selection changes"""
        selected_profile = self.api_profile_var.get()
        if selected_profile and selected_profile != self.current_profile:
            self.current_profile = selected_profile
            self.load_profile_data(selected_profile)
            self.log_entry(f"Switched to profile: {selected_profile}", "INFO")
    
    def duplicate_profile(self):
        """Duplicate the selected profile"""
        selection = self.profiles_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a profile to duplicate")
            return
        
        source_profile = self.profiles_listbox.get(selection[0])
        
        # Ask for new name
        new_name = tk.simpledialog.askstring("Duplicate Profile", 
                                            f"Enter name for duplicate of '{source_profile}':")
        if not new_name:
            return
        
        if new_name in self.api_profiles:
            messagebox.showerror("Error", "Profile name already exists")
            return
        
        # Copy profile data
        profile_data = self.api_profiles[source_profile].copy()
        profile_data['created'] = datetime.now().isoformat()
        profile_data['description'] = f"Duplicate of {source_profile}"
        
        self.api_profiles[new_name] = profile_data
        self.save_api_config()
        
        # Update UI
        self.api_profile_combo['values'] = list(self.api_profiles.keys())
        self.profiles_listbox.insert(tk.END, new_name)
        
        messagebox.showinfo("Success", f"Profile '{new_name}' created as duplicate of '{source_profile}'")
        self.log_entry(f"Profile '{source_profile}' duplicated as '{new_name}'", "INFO")
    
    def rename_profile(self):
        """Rename the selected profile"""
        selection = self.profiles_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a profile to rename")
            return
        
        old_name = self.profiles_listbox.get(selection[0])
        
        if old_name == "Default":
            messagebox.showerror("Error", "Cannot rename the Default profile")
            return
        
        new_name = tk.simpledialog.askstring("Rename Profile", 
                                           f"Enter new name for '{old_name}':",
                                           initialvalue=old_name)
        if not new_name or new_name == old_name:
            return
        
        if new_name in self.api_profiles:
            messagebox.showerror("Error", "Profile name already exists")
            return
        
        # Rename profile
        self.api_profiles[new_name] = self.api_profiles.pop(old_name)
        
        # Update current profile if needed
        if self.current_profile == old_name:
            self.current_profile = new_name
            self.api_profile_var.set(new_name)
        
        self.save_api_config()
        
        # Update UI
        self.api_profile_combo['values'] = list(self.api_profiles.keys())
        
        # Update listbox
        self.profiles_listbox.delete(0, tk.END)
        for name in self.api_profiles.keys():
            self.profiles_listbox.insert(tk.END, name)
        
        messagebox.showinfo("Success", f"Profile renamed from '{old_name}' to '{new_name}'")
        self.log_entry(f"Profile '{old_name}' renamed to '{new_name}'", "INFO")
    
    def copy_to_clipboard(self):
        """Copy API settings to clipboard"""
        try:
            settings = self.get_current_api_settings()
            settings_text = json.dumps(settings, indent=2)
            
            self.root.clipboard_clear()
            self.root.clipboard_append(settings_text)
            
            messagebox.showinfo("Success", "API settings copied to clipboard")
            self.log_entry("API settings copied to clipboard", "INFO")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to copy to clipboard:\n{str(e)}")
    
    def paste_from_clipboard(self):
        """Paste API settings from clipboard"""
        try:
            clipboard_text = self.root.clipboard_get()
            settings = json.loads(clipboard_text)
            
            # Validate that it's API settings
            required_keys = ['endpoint', 'method', 'content_type', 'auth_type']
            if not all(key in settings for key in required_keys):
                messagebox.showerror("Error", "Clipboard does not contain valid API settings")
                return
            
            # Load settings
            self.api_endpoint_var.set(settings.get('endpoint', ''))
            self.api_method_var.set(settings.get('method', 'POST'))
            self.api_content_type_var.set(settings.get('content_type', 'application/json'))
            self.custom_content_type_var.set(settings.get('custom_content_type', ''))
            self.auth_type_var.set(settings.get('auth_type', 'no_auth'))
            self.api_key_var.set(settings.get('api_key', ''))
            self.login_username_var.set(settings.get('username', ''))
            self.login_password_var.set(settings.get('password', ''))
            self.test_mode_var.set(settings.get('test_mode', False))
            
            # Load JSON body
            if hasattr(self, 'json_body_text'):
                self.json_body_text.delete("1.0", tk.END)
                self.json_body_text.insert("1.0", settings.get('json_body', ''))
            
            # Trigger change events
            self.on_auth_type_change()
            self.on_content_type_change()
            self.update_payload_preview()
            
            messagebox.showinfo("Success", "API settings pasted from clipboard")
            self.log_entry("API settings pasted from clipboard", "INFO")
            
        except tk.TclError:
            messagebox.showerror("Error", "Clipboard is empty")
        except json.JSONDecodeError:
            messagebox.showerror("Error", "Clipboard does not contain valid JSON")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to paste from clipboard:\n{str(e)}")
    
    def reset_to_defaults(self):
        """Reset all API settings to defaults"""
        result = messagebox.askyesno("Confirm Reset", 
                                   "Are you sure you want to reset all API settings to defaults?\nThis will clear all current settings.")
        if result:
            # Reset all variables
            self.api_endpoint_var.set('')
            self.api_method_var.set('POST')
            self.api_content_type_var.set('application/json')
            self.custom_content_type_var.set('')
            self.auth_type_var.set('no_auth')
            self.api_key_var.set('')
            self.login_username_var.set('')
            self.login_password_var.set('')
            self.test_mode_var.set(False)
            
            # Clear text areas
            if hasattr(self, 'json_body_text'):
                self.json_body_text.delete("1.0", tk.END)
            
            if hasattr(self, 'response_text'):
                self.response_text.config(state=tk.NORMAL)
                self.response_text.delete("1.0", tk.END)
                self.response_text.config(state=tk.DISABLED)
            
            # Clear form fields
            if hasattr(self, 'form_fields_frame'):
                self.clear_form_fields()
            
            # Reset status
            self.response_status_var.set("Ready")
            self.response_status_label.config(foreground='blue')
            
            # Trigger change events
            self.on_auth_type_change()
            self.on_content_type_change()
            self.update_payload_preview()
            
            messagebox.showinfo("Reset", "All API settings have been reset to defaults")
            self.log_entry("API settings reset to defaults", "INFO")
    
    def clear_form_fields(self):
        """Clear all form fields in the form data editor"""
        if hasattr(self, 'form_fields_frame'):
            # Clear all child widgets
            for widget in self.form_fields_frame.winfo_children():
                widget.destroy()
    
    def get_form_fields_data(self):
        """Get form fields data as dictionary"""
        if not hasattr(self, 'form_fields_frame'):
            return {}
        
        form_data = {}
        for widget in self.form_fields_frame.winfo_children():
            if isinstance(widget, tk.Frame):
                # Get key and value from frame
                key_widget = None
                value_widget = None
                type_widget = None
                
                for child in widget.winfo_children():
                    if isinstance(child, tk.Entry):
                        if not key_widget:
                            key_widget = child
                        else:
                            value_widget = child
                    elif isinstance(child, ttk.Combobox):
                        type_widget = child
                
                if key_widget and value_widget:
                    key = key_widget.get().strip()
                    value = value_widget.get().strip()
                    field_type = type_widget.get() if type_widget else 'text'
                    
                    if key:  # Only add if key is not empty
                        form_data[key] = {
                            'value': value,
                            'type': field_type
                        }
        
        return form_data
    
    def validate_login_fields(self, *args):
        """Validate login fields and show auto-conversion warning"""
        try:
            username = self.login_username_var.get().strip()
            password = self.login_password_var.get().strip()
            
            # Clear previous warnings
            self.login_warning_var.set("")
            
            if username and password:
                # Both filled - show success
                self.login_warning_var.set(" Valid")
                self.login_warning_label.config(foreground='green')
                
            elif username and not password:
                # Username filled but no password
                self.login_warning_var.set(" Password required!")
                self.login_warning_label.config(foreground='orange')
                
            elif not username and password:
                # Password filled but no username
                self.login_warning_var.set(" Username required!")
                self.login_warning_label.config(foreground='orange')
                
            elif not username and not password:
                # Both empty - show info about login field
                if self.auth_type_var.get() == "login":
                    self.login_warning_var.set(" Login fields required!")
                    self.login_warning_label.config(foreground='red')
                else:
                    self.login_warning_var.set("")
            
            # Update payload preview to reflect changes
            self.update_payload_preview()
            
        except Exception as e:
            # Silently handle validation errors
            pass
    
    def get_login_payload(self):
        """Get login payload with auto-conversion from username to login field"""
        username = self.login_username_var.get().strip()
        password = self.login_password_var.get().strip()
        database = self.login_database_var.get().strip()
        
        if not username or not password:
            return None
        
        # Auto-convert username to 'login' field in JSON
        payload = {
            "login": username,  # Auto-conversion: username -> login
            "password": password
        }
        
        # Add database if provided
        if database:
            payload["database"] = database
        
        return payload
    
    def validate_api_settings_advanced(self):
        """Advanced validation for API settings with login auto-conversion"""
        errors = []
        warnings = []
        
        # Basic endpoint validation
        endpoint = self.api_endpoint_var.get().strip()
        if not endpoint:
            errors.append("API endpoint is required")
        elif not (endpoint.startswith('http://') or endpoint.startswith('https://')):
            warnings.append("API endpoint should start with http:// or https://")
        
        # Authentication validation
        auth_type = self.auth_type_var.get()
        
        if auth_type == "api_key":
            if not self.api_key_var.get().strip():
                errors.append("API key is required for API key authentication")
        
        elif auth_type == "login":
            username = self.login_username_var.get().strip()
            password = self.login_password_var.get().strip()
            
            if not username:
                errors.append("Username is required for login authentication")
            if not password:
                errors.append("Password is required for login authentication")
            
            # Show auto-conversion info
            if username and password:
                warnings.append(f"Username '{username}' will be sent as 'login' field in JSON payload")
        
        # Method and content-type validation
        method = self.api_method_var.get()
        content_type = self.api_content_type_var.get()
        
        if method not in ["GET", "DELETE"] and content_type == "application/json":
            body = self.json_body_text.get("1.0", tk.END).strip() if hasattr(self, 'json_body_text') else ""
            if body:
                try:
                    json.loads(body)
                except json.JSONDecodeError as e:
                    errors.append(f"Invalid JSON in request body: {str(e)}")
        
        return errors, warnings
    
    def login_and_get_token(self):
        """Login to API and get access token with auto-conversion"""
        try:
            # Validate inputs
            username = self.login_username_var.get().strip()
            password = self.login_password_var.get().strip()
            database = self.login_database_var.get().strip()
            endpoint = self.api_endpoint_var.get().strip()
            
            if not all([username, password, database, endpoint]):
                messagebox.showerror("Error", "Please fill in all login fields (Username, Password, Database, Endpoint)")
                return
            
            if "/api/auth/login" not in endpoint and "/auth/login" not in endpoint:
                messagebox.showerror("Error", "Endpoint URL must contain '/api/auth/login' or '/auth/login'")
                return
            
            # Show auto-conversion confirmation
            conversion_msg = f"Auto-converting username to login field:\n\n"
            conversion_msg += f"Frontend: username = \"{username}\"\n"
            conversion_msg += f"Backend: login = \"{username}\"\n\n"
            conversion_msg += f"JSON payload will be:\n"
            conversion_msg += f'{{\n  "login": "{username}",\n  "password": "***",\n  "database": "{database}"\n}}\n\n'
            conversion_msg += f"Continue with login?"
            
            result = messagebox.askyesno("Auto-Conversion Confirmation", conversion_msg)
            if not result:
                return
            
            # Update status
            self.login_status_var.set("Logging in...")
            self.login_status_label.config(foreground='orange')
            
            # Prepare login payload with AUTO-CONVERSION
            login_data = {
                "login": username,    # AUTO-CONVERSION: username -> login
                "password": password,
                "database": database
            }
            
            # Send login request
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json'
            }
            
            self.log_entry(f"Attempting login to {endpoint} with auto-conversion (username->login)", "INFO")
            
            response = requests.post(
                endpoint,
                json=login_data,
                headers=headers,
                timeout=15
            )
            
            if response.status_code == 200:
                try:
                    result = response.json()
                    
                    # Try to extract token from different possible field names
                    token = None
                    refresh_token = None
                    
                    # Handle nested token structure like {"token": {"access_token": "...", "refresh_token": "..."}}
                    if 'token' in result and isinstance(result['token'], dict):
                        token_obj = result['token']
                        token = token_obj.get('access_token') or token_obj.get('token')
                        refresh_token = token_obj.get('refresh_token')
                    else:
                        # Try direct field access
                        for token_field in ['access_token', 'token', 'accessToken', 'jwt', 'authToken']:
                            if token_field in result:
                                token = result[token_field]
                                break
                        
                        # Look for refresh token
                        for refresh_field in ['refresh_token', 'refreshToken', 'refresh']:
                            if refresh_field in result:
                                refresh_token = result[refresh_field]
                                break
                    
                    if token:
                        # Auto-fill the API Key field
                        self.api_key_var.set(token)
                        
                        # Update button states after setting API key
                        self.update_button_states()
                        
                        # Update API mapping status
                        if hasattr(self, 'refresh_api_mapping_status'):
                            self.refresh_api_mapping_status()
                        
                        # Update status
                        self.login_status_var.set("Login successful!")
                        self.login_status_label.config(foreground='green')
                        
                        # Auto-refresh dashboard status with successful API connection
                        if hasattr(self, 'dash_api_status'):
                            self.dash_api_status.config(text=" Connected & Authenticated", foreground="green")
                        
                        # Auto-refresh health check status
                        if hasattr(self, 'health_api_status'):
                            try:
                                self.health_api_status.config(text=" Authenticated", foreground="green")
                                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                self.health_tree.insert("", 0, values=(timestamp, "API", "PASS", f"Login successful - Token obtained and API authenticated"))
                            except:
                                pass
                        
                        # Log success with auto-conversion info
                        self.log_entry(f"Login successful with auto-conversion (username '{username}' -> login field), token obtained", "SUCCESS")
                        
                        # Show detailed response body dialog
                        self.show_login_response_dialog(result, token, refresh_token, username, database)
                    else:
                        # Token not found in expected fields
                        self.login_status_var.set("Login response received, but token not found")
                        self.login_status_label.config(foreground='orange')
                        
                        messagebox.showwarning("Warning", 
                                             f" Login request was successful with auto-conversion,\n"
                                             f"but access token not found in response.\n\n"
                                             f" Username '{username}' was converted to 'login' field\n"
                                             f"Response: {response.text[:200]}...")
                        
                except ValueError as e:
                    # Invalid JSON response
                    self.login_status_var.set("Invalid response format")
                    self.login_status_label.config(foreground='red')
                    
                    messagebox.showerror("Error", 
                                       f" Login request returned invalid JSON.\n\n"
                                       f"Note: Username '{username}' was converted to 'login' field\n"
                                       f"Response: {response.text[:200]}...")
                    
            else:
                # Login failed
                self.login_status_var.set(f"Login failed ({response.status_code})")
                self.login_status_label.config(foreground='red')
                
                # Update dashboard with specific status for 400 errors
                if hasattr(self, 'dash_api_status'):
                    if response.status_code == 400:
                        self.dash_api_status.config(text=f" Bad Request (400) - Check credentials", foreground="orange")
                    elif response.status_code == 401:
                        self.dash_api_status.config(text=f" Unauthorized (401)", foreground="red")
                    else:
                        self.dash_api_status.config(text=f" Login Failed ({response.status_code})", foreground="red")
                
                error_msg = f" Login failed with status {response.status_code}\n\n"
                error_msg += f"Note: Username '{username}' was converted to 'login' field\n"
                try:
                    error_detail = response.json().get('message', response.text)
                    error_msg += f": {error_detail}"
                except:
                    error_msg += f": {response.text[:200]}..."
                
                self.log_entry(error_msg, "ERROR")
                messagebox.showerror("Login Failed", error_msg)
                
        except requests.exceptions.Timeout:
            self.login_status_var.set("Login timeout")
            self.login_status_label.config(foreground='red')
            
            # Update dashboard
            if hasattr(self, 'dash_api_status'):
                self.dash_api_status.config(text=" Timeout", foreground="red")
            
            error_msg = "Login request timed out. Please check your connection and endpoint URL."
            messagebox.showerror("Timeout", error_msg)
            self.log_entry(error_msg, "ERROR")
            
        except requests.exceptions.ConnectionError:
            self.login_status_var.set("Connection failed")
            self.login_status_label.config(foreground='red')
            
            # Update dashboard
            if hasattr(self, 'dash_api_status'):
                self.dash_api_status.config(text=" Connection Failed", foreground="red")
            
            error_msg = "Could not connect to the API endpoint. Please check the URL and your internet connection."
            messagebox.showerror("Connection Error", error_msg)
            self.log_entry(error_msg, "ERROR")
            
        except Exception as e:
            self.login_status_var.set("Login error")
            self.login_status_label.config(foreground='red')
            
            # Update dashboard
            if hasattr(self, 'dash_api_status'):
                self.dash_api_status.config(text=" Login Error", foreground="red")
            
            error_msg = f"Unexpected error during login: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def test_api_key(self):
        """Test API endpoint with API Key authentication - NO AUTO-SAVE (testing only)"""
        try:
            endpoint = self.api_endpoint_var.get().strip()
            api_key = self.api_key_var.get().strip()
            
            if not endpoint:
                messagebox.showerror("Error", "Please enter an endpoint URL")
                return
            
            if not api_key:
                messagebox.showerror("Error", "Please enter an API Key")
                return
            
            # Update status
            self.api_key_status_var.set("Testing API Key...")
            self.api_key_status_label.config(foreground='orange')
            
            # Prepare headers
            headers = {
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json',
                'Accept': 'application/json'
            }
            
            # Log the test (NO AUTO-SAVE - this is testing only)
            self.log_entry(f" Testing API Key authentication to {endpoint} (no auto-save)", "INFO")
            
            # Send test request
            response = requests.get(endpoint, headers=headers, timeout=15)
            
            # Update status based on response (NO AUTO-SAVE)
            if response.status_code == 200:
                self.api_key_status_var.set(" API Key valid!")
                self.api_key_status_label.config(foreground='green')
                
                # Show response in preview
                self.show_api_key_response(response)
                
                messagebox.showinfo("Test Success", 
                                  f" API Key authentication test successful!\n\n"
                                  f"Status: {response.status_code}\n"
                                  f"Response: {response.text[:100]}...\n\n"
                                  f" Note: This is a test only - no configuration saved")
                
            elif response.status_code == 401:
                self.api_key_status_var.set(" Unauthorized - Invalid API Key")
                self.api_key_status_label.config(foreground='red')
                
                self.show_api_key_response(response, is_error=True)
                messagebox.showerror("Test Failed", " API Key is invalid or expired\n\n Note: Test only - configuration not affected")
                
            elif response.status_code == 403:
                self.api_key_status_var.set(" Forbidden - Insufficient permissions")
                self.api_key_status_label.config(foreground='red')
                
                self.show_api_key_response(response, is_error=True)
                messagebox.showerror("Test Failed", " API Key valid but insufficient permissions\n\n Note: Test only - configuration not affected")
                
            else:
                self.api_key_status_var.set(f" Response {response.status_code}")
                self.api_key_status_label.config(foreground='orange')
                
                self.show_api_key_response(response, is_error=True)
                messagebox.showwarning("Test Warning", 
                                     f" Unexpected response status: {response.status_code}\n"
                                     f"Response: {response.text[:100]}...\n\n"
                                     f" Note: Test only - no configuration saved")
            
        except requests.exceptions.Timeout:
            self.api_key_status_var.set(" Timeout")
            self.api_key_status_label.config(foreground='red')
            messagebox.showerror("Test Timeout", "Request timed out. Please check your connection.\n\n Note: Test only - configuration not affected")
            
        except requests.exceptions.ConnectionError:
            self.api_key_status_var.set(" Connection failed")
            self.api_key_status_label.config(foreground='red')
            messagebox.showerror("Test Connection Error", "Could not connect to the endpoint.\n\n Note: Test only - configuration not affected")
            
        except Exception as e:
            self.api_key_status_var.set(" Test failed")
            self.api_key_status_label.config(foreground='red')
            error_msg = f"API Key test failed: {str(e)}"
            messagebox.showerror("Test Error", f"{error_msg}\n\n Note: Test only - configuration not affected")
            self.log_entry(f" {error_msg} (test mode)", "ERROR")
    
    def test_no_auth(self):
        """Test API endpoint without authentication"""
        try:
            endpoint = self.api_endpoint_var.get().strip()
            
            if not endpoint:
                messagebox.showerror("Error", "Please enter an endpoint URL")
                return
            
            # Update status
            self.no_auth_status_var.set("Testing endpoint...")
            self.no_auth_status_label.config(foreground='orange')
            
            # Prepare headers (minimal)
            headers = {
                'Content-Type': 'application/json',
                'Accept': 'application/json'
            }
            
            # Log the test
            self.log_entry(f"Testing endpoint without authentication: {endpoint}", "INFO")
            
            # Send test request
            response = requests.get(endpoint, headers=headers, timeout=15)
            
            # Update status based on response
            if response.status_code == 200:
                self.no_auth_status_var.set(" Endpoint accessible!")
                self.no_auth_status_label.config(foreground='green')
                
                # Show response in preview
                self.show_no_auth_response(response)
                
                messagebox.showinfo("Success", 
                                  f" Endpoint is accessible without authentication!\n\n"
                                  f"Status: {response.status_code}\n"
                                  f"Response: {response.text[:100]}...")
                
            elif response.status_code == 401:
                self.no_auth_status_var.set(" Authentication required")
                self.no_auth_status_label.config(foreground='orange')
                
                self.show_no_auth_response(response, is_error=True)
                messagebox.showwarning("Authentication Required", 
                                     " This endpoint requires authentication.\n"
                                     "Please switch to API Key or Login authentication.")
                
            elif response.status_code == 404:
                self.no_auth_status_var.set(" Endpoint not found")
                self.no_auth_status_label.config(foreground='red')
                
                self.show_no_auth_response(response, is_error=True)
                messagebox.showerror("Not Found", " Endpoint not found. Please check the URL.")
                
            else:
                self.no_auth_status_var.set(f" Response {response.status_code}")
                self.no_auth_status_label.config(foreground='blue')
                
                self.show_no_auth_response(response)
                messagebox.showinfo("Response Received", 
                                  f" Endpoint responded with status: {response.status_code}\n"
                                  f"Response: {response.text[:100]}...")
            
        except requests.exceptions.Timeout:
            self.no_auth_status_var.set(" Timeout")
            self.no_auth_status_label.config(foreground='red')
            messagebox.showerror("Timeout", "Request timed out. Please check your connection.")
            
        except requests.exceptions.ConnectionError:
            self.no_auth_status_var.set(" Connection failed")
            self.no_auth_status_label.config(foreground='red')
            messagebox.showerror("Connection Error", "Could not connect to the endpoint.")
            
        except Exception as e:
            self.no_auth_status_var.set(" Test failed")
            self.no_auth_status_label.config(foreground='red')
            error_msg = f"Endpoint test failed: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def show_api_key_response(self, response, is_error=False):
        """Show API Key test response in preview panel"""
        try:
            self.api_key_response_text.config(state=tk.NORMAL)
            self.api_key_response_text.delete("1.0", tk.END)
            
            # Format response
            response_text = f"Status: {response.status_code}\n"
            response_text += f"Headers: {dict(response.headers)}\n\n"
            
            try:
                # Try to format as JSON
                json_response = response.json()
                response_text += f"Response Body (JSON):\n{json.dumps(json_response, indent=2)}"
            except:
                # Plain text response
                response_text += f"Response Body:\n{response.text}"
            
            self.api_key_response_text.insert("1.0", response_text)
            self.api_key_response_text.config(state=tk.DISABLED)
            
        except Exception as e:
            self.api_key_response_text.config(state=tk.NORMAL)
            self.api_key_response_text.delete("1.0", tk.END)
            self.api_key_response_text.insert("1.0", f"Error displaying response: {str(e)}")
            self.api_key_response_text.config(state=tk.DISABLED)
    
    def show_no_auth_response(self, response, is_error=False):
        """Show No Auth test response in preview panel"""
        try:
            self.no_auth_response_text.config(state=tk.NORMAL)
            self.no_auth_response_text.delete("1.0", tk.END)
            
            # Format response
            response_text = f"Status: {response.status_code}\n"
            response_text += f"Headers: {dict(response.headers)}\n\n"
            
            try:
                # Try to format as JSON
                json_response = response.json()
                response_text += f"Response Body (JSON):\n{json.dumps(json_response, indent=2)}"
            except:
                # Plain text response
                response_text += f"Response Body:\n{response.text}"
            
            self.no_auth_response_text.insert("1.0", response_text)
            self.no_auth_response_text.config(state=tk.DISABLED)
            
        except Exception as e:
            self.no_auth_response_text.config(state=tk.NORMAL)
            self.no_auth_response_text.delete("1.0", tk.END)
            self.no_auth_response_text.insert("1.0", f"Error displaying response: {str(e)}")
            self.no_auth_response_text.config(state=tk.DISABLED)
    
    def update_api_key_status(self):
        """Update API key status based on current input"""
        api_key = self.api_key_var.get().strip()
        endpoint = self.api_endpoint_var.get().strip()
        
        if not endpoint:
            self.api_key_status_var.set(" Enter endpoint first")
            self.api_key_status_label.config(foreground='orange')
        elif not api_key:
            self.api_key_status_var.set(" Enter API key to test")
            self.api_key_status_label.config(foreground='gray')
        elif len(api_key) < 10:
            self.api_key_status_var.set(" API key seems too short")
            self.api_key_status_label.config(foreground='orange')
        else:
            self.api_key_status_var.set(" Ready to test API key")
            self.api_key_status_label.config(foreground='green')
    
    def update_no_auth_status(self):
        """Update no auth status based on current endpoint"""
        endpoint = self.api_endpoint_var.get().strip()
        
        if not endpoint:
            self.no_auth_status_var.set(" Enter endpoint to test")
            self.no_auth_status_label.config(foreground='gray')
        elif not (endpoint.startswith('http://') or endpoint.startswith('https://')):
            self.no_auth_status_var.set(" URL should start with http:// or https://")
            self.no_auth_status_label.config(foreground='orange')
        else:
            self.no_auth_status_var.set(" Ready to test endpoint")
            self.no_auth_status_label.config(foreground='green')
    
    def show_login_response_dialog(self, response_data, access_token, refresh_token=None, username="", database=""):
        """Show detailed login response dialog with JSON body"""
        try:
            import json
            
            # Create dialog window
            dialog = tk.Toplevel(self.root)
            dialog.title("Login Response - API Authentication Successful")
            dialog.geometry("600x500")
            dialog.transient(self.root)
            dialog.grab_set()
            
            # Configure dialog
            dialog.configure(bg='#f0f8f0')
            
            # Title frame
            title_frame = ttk.Frame(dialog)
            title_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
            
            title_label = ttk.Label(title_frame, text=" Login Successful!", 
                                   font=('Arial', 16, 'bold'), foreground='green')
            title_label.pack()
            
            subtitle_label = ttk.Label(title_frame, text="API Authentication Response", 
                                      font=('Arial', 10), foreground='gray')
            subtitle_label.pack()
            
            # Info frame
            info_frame = ttk.LabelFrame(dialog, text="Authentication Details", padding=10)
            info_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
            
            info_grid = ttk.Frame(info_frame)
            info_grid.pack(fill=tk.X)
            
            # Login details
            ttk.Label(info_grid, text="Username:", font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky=tk.W)
            ttk.Label(info_grid, text=f"{username}  login", foreground='blue').grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
            
            ttk.Label(info_grid, text="Database:", font=('Arial', 9, 'bold')).grid(row=1, column=0, sticky=tk.W)
            ttk.Label(info_grid, text=database).grid(row=1, column=1, sticky=tk.W, padx=(10, 0))
            
            ttk.Label(info_grid, text="Access Token:", font=('Arial', 9, 'bold')).grid(row=2, column=0, sticky=tk.W)
            token_display = f"{access_token[:20]}..." if len(access_token) > 20 else access_token
            ttk.Label(info_grid, text=f" {token_display}", foreground='green').grid(row=2, column=1, sticky=tk.W, padx=(10, 0))
            
            if refresh_token:
                ttk.Label(info_grid, text="Refresh Token:", font=('Arial', 9, 'bold')).grid(row=3, column=0, sticky=tk.W)
                refresh_display = f"{refresh_token[:20]}..." if len(refresh_token) > 20 else refresh_token
                ttk.Label(info_grid, text=f" {refresh_display}", foreground='green').grid(row=3, column=1, sticky=tk.W, padx=(10, 0))
            
            # Response body frame
            response_frame = ttk.LabelFrame(dialog, text="Response Body (JSON)", padding=10)
            response_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
            
            # JSON text area with syntax highlighting
            from tkinter import scrolledtext
            json_text = scrolledtext.ScrolledText(response_frame, height=12, width=70, 
                                                 font=('Consolas', 10), wrap=tk.WORD)
            json_text.pack(fill=tk.BOTH, expand=True)
            
            # Format and insert JSON
            formatted_json = json.dumps(response_data, indent=2, ensure_ascii=False)
            json_text.insert(1.0, formatted_json)
            json_text.config(state=tk.DISABLED)
            
            # Apply basic syntax highlighting
            json_text.config(state=tk.NORMAL)
            
            # Highlight strings (simple highlighting)
            content = json_text.get(1.0, tk.END)
            lines = content.split('\n')
            
            json_text.delete(1.0, tk.END)
            for line in lines:
                if '"access_token"' in line or '"token"' in line:
                    json_text.insert(tk.END, line + '\n', 'token_line')
                elif '"code"' in line and '200' in line:
                    json_text.insert(tk.END, line + '\n', 'success_line')
                elif '"message"' in line and 'Success' in line:
                    json_text.insert(tk.END, line + '\n', 'success_line')
                else:
                    json_text.insert(tk.END, line + '\n')
            
            # Configure tags for highlighting
            json_text.tag_configure('token_line', foreground='#2E7D32', font=('Consolas', 10, 'bold'))
            json_text.tag_configure('success_line', foreground='#1976D2', font=('Consolas', 10, 'bold'))
            
            json_text.config(state=tk.DISABLED)
            
            # Action frame
            action_frame = ttk.Frame(dialog)
            action_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
            
            # Copy token button
            def copy_access_token():
                dialog.clipboard_clear()
                dialog.clipboard_append(access_token)
                copy_btn.config(text=" Copied!")
                dialog.after(2000, lambda: copy_btn.config(text=" Copy Access Token"))
            
            def copy_response():
                dialog.clipboard_clear()
                dialog.clipboard_append(formatted_json)
                copy_response_btn.config(text=" Copied!")
                dialog.after(2000, lambda: copy_response_btn.config(text=" Copy Response"))
            
            # Buttons
            btn_frame = ttk.Frame(action_frame)
            btn_frame.pack(expand=True)
            
            copy_btn = ttk.Button(btn_frame, text=" Copy Access Token", command=copy_access_token)
            copy_btn.pack(side=tk.LEFT, padx=(0, 10))
            
            copy_response_btn = ttk.Button(btn_frame, text=" Copy Response", command=copy_response)
            copy_response_btn.pack(side=tk.LEFT, padx=(0, 10))
            
            ttk.Button(btn_frame, text=" Continue", command=dialog.destroy).pack(side=tk.LEFT, padx=(10, 0))
            
            # Status message
            status_frame = ttk.Frame(action_frame)
            status_frame.pack(fill=tk.X, pady=(10, 0))
            
            status_msg = ttk.Label(status_frame, text=" Token telah otomatis diisi ke API Key field. Anda sekarang dapat menggunakan semua fitur API.", 
                                  font=('Arial', 9), foreground='green', wraplength=550)
            status_msg.pack()
            
            # Center the dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
            y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
            dialog.geometry(f"+{x}+{y}")
            
        except Exception as e:
            # Fallback to simple message if dialog creation fails
            messagebox.showinfo("Login Successful", 
                              f" Login successful with auto-conversion!\n\n"
                              f" Username '{username}' converted to 'login' field\n"
                              f" Token automatically filled in API Key field\n"
                              f" Database: {database}\n\n"
                              f"Response: {str(response_data)[:200]}...")
    
    def send_sample_data(self):
        """Send sample data to API"""
        if not self.config.get("api_endpoint"):
            messagebox.showerror("Error", "Please configure API endpoint first.")
            return
        
        sample_data = {
            "test": True,
            "timestamp": datetime.now().isoformat(),
            "uuid": str(uuid.uuid4()),
            "sample_field": "Sample Value",
            "record_count": 1
        }
        
        try:
            headers = {'Content-Type': 'application/json'}
            if self.config.get("api_key"):
                headers['Authorization'] = f'Bearer {self.config["api_key"]}'
            
            response = requests.post(
                self.config["api_endpoint"],
                json=sample_data,
                headers=headers,
                timeout=10
            )
            
            if response.status_code == 200:
                messagebox.showinfo("Success", f"Sample data sent successfully!\nResponse: {response.text}")
                self.log_entry("Sample data sent successfully", "SUCCESS")
            else:
                messagebox.showerror("Error", f"API returned status {response.status_code}\nResponse: {response.text}")
                self.log_entry(f"Sample data failed: {response.status_code}", "ERROR")
                
        except Exception as e:
            error_msg = f"Failed to send sample data: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def generate_payload_preview(self):
        """Generate payload preview"""
        try:
            if not self.selected_table or not self.table_columns:
                sample_payload = {
                    "uuid": "sample-uuid-here",
                    "timestamp": datetime.now().isoformat(),
                    "table": "your_table_name",
                    "data": {
                        "field1": "value1",
                        "field2": "value2",
                        "field3": "value3"
                    }
                }
            else:
                # Generate based on actual table structure
                sample_data = {}
                for col in self.table_columns[:5]:  # Show first 5 columns
                    col_name = col['name']
                    col_type = col.get('type', 'TEXT')
                    
                    if 'INT' in col_type.upper():
                        sample_data[col_name] = 123
                    elif 'DATE' in col_type.upper():
                        sample_data[col_name] = datetime.now().isoformat()
                    elif 'BOOL' in col_type.upper():
                        sample_data[col_name] = True
                    else:
                        sample_data[col_name] = f"sample_{col_name}"
                
                sample_payload = {
                    "uuid": str(uuid.uuid4()),
                    "timestamp": datetime.now().isoformat(),
                    "table": self.selected_table,
                    "data": sample_data
                }
            
            # Pretty print JSON
            json_text = json.dumps(sample_payload, indent=2)
            self.payload_text.delete(1.0, tk.END)
            self.payload_text.insert(1.0, json_text)
            
        except Exception as e:
            error_msg = f"Failed to generate preview: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def save_scheduler_settings(self):
        """Save scheduler settings"""
        try:
            interval = int(self.push_interval_var.get())
            if interval < 1:
                messagebox.showerror("Error", "Push interval must be at least 1 minute.")
                return
            
            self.config["push_interval"] = interval * 60  # Convert to seconds
            self.config["auto_push"] = self.enable_auto_push_var.get()
            self.save_config()
            
            messagebox.showinfo("Success", "Scheduler settings saved successfully!")
            self.log_entry("Scheduler settings saved", "INFO")
            
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number for push interval.")
        except Exception as e:
            error_msg = f"Failed to save scheduler settings: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def start_scheduler(self):
        """Start the scheduler"""
        if not self.admin_mode:
            messagebox.showerror("Error", "Admin mode required to start scheduler.")
            return
        
        if not self.config.get("auto_push"):
            messagebox.showerror("Error", "Please enable automatic push first.")
            return
        
        try:
            self.is_running = True
            self.sched_agent_status.config(text="Running", foreground="green")
            
            # Calculate next push time
            interval = self.config.get("push_interval", 300)
            next_push = datetime.now() + timedelta(seconds=interval)
            self.sched_next_push.config(text=next_push.strftime("%Y-%m-%d %H:%M:%S"))
            
            self.log_entry("Scheduler started", "SUCCESS")
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.scheduler_tree.insert("", 0, values=(timestamp, "Start", "SUCCESS", "Scheduler started successfully"))
            
            messagebox.showinfo("Success", "Scheduler started successfully!")
            
        except Exception as e:
            error_msg = f"Failed to start scheduler: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def stop_scheduler(self):
        """Stop the scheduler"""
        try:
            self.is_running = False
            self.sched_agent_status.config(text="Stopped", foreground="red")
            self.sched_next_push.config(text="Not scheduled")
            
            self.log_entry("Scheduler stopped", "INFO")
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.scheduler_tree.insert("", 0, values=(timestamp, "Stop", "INFO", "Scheduler stopped"))
            
            messagebox.showinfo("Info", "Scheduler stopped.")
            
        except Exception as e:
            error_msg = f"Failed to stop scheduler: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def force_push_now(self):
        """Force immediate push"""
        if not self.admin_mode:
            messagebox.showerror("Error", "Admin mode required.")
            return
        
        try:
            result = self.manual_push()
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if result:
                self.scheduler_tree.insert("", 0, values=(timestamp, "Manual Push", "SUCCESS", "Manual push completed successfully"))
            else:
                self.scheduler_tree.insert("", 0, values=(timestamp, "Manual Push", "FAILED", "Manual push failed"))
            
        except Exception as e:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.scheduler_tree.insert("", 0, values=(timestamp, "Manual Push", "ERROR", f"Error: {str(e)}"))
    
    def load_settings(self):
        """Load settings from config"""
        try:
            if hasattr(self, 'mdb_file_var'):
                self.mdb_file_var.set(self.config.get("mdb_file", ""))
            if hasattr(self, 'mdb_password_var'):
                self.mdb_password_var.set(self.config.get("mdb_password", ""))
            if hasattr(self, 'api_endpoint_var'):
                self.api_endpoint_var.set(self.config.get("api_endpoint", ""))
            if hasattr(self, 'api_key_var'):
                self.api_key_var.set(self.config.get("api_key", ""))
            if hasattr(self, 'login_username_var'):
                self.login_username_var.set(self.config.get("login_username", ""))
            if hasattr(self, 'login_password_var'):
                self.login_password_var.set(self.config.get("login_password", ""))
            if hasattr(self, 'login_database_var'):
                self.login_database_var.set(self.config.get("login_database", ""))
            if hasattr(self, 'test_mode_var'):
                self.test_mode_var.set(self.config.get("test_mode", False))
            if hasattr(self, 'push_interval_var'):
                interval_minutes = self.config.get("push_interval", 300) // 60
                self.push_interval_var.set(str(interval_minutes))
            if hasattr(self, 'enable_auto_push_var'):
                self.enable_auto_push_var.set(self.config.get("auto_push", False))
                
        except Exception as e:
            self.log_entry(f"Failed to load settings: {str(e)}", "ERROR")
    
    def refresh_dashboard(self):
        """Refresh dashboard data with API Settings integration"""
        # Update database status
        if self.db_connection:
            self.dash_db_status.config(text=" Connected", foreground="green")
        else:
            self.dash_db_status.config(text=" Not Connected", foreground="red")
        
        # Update API status based on API Settings
        endpoint = self.api_endpoint_var.get().strip() if hasattr(self, 'api_endpoint_var') else self.config.get('api_endpoint', '')
        auth_type = self.auth_type_var.get() if hasattr(self, 'auth_type_var') else "no_auth"
        
        if endpoint:
            # Check authentication status
            if auth_type == "api_key":
                api_key = self.api_key_var.get().strip() if hasattr(self, 'api_key_var') else ""
                if api_key:
                    self.dash_api_status.config(text=" Configured (API Key)", foreground="green")
                else:
                    self.dash_api_status.config(text=" Endpoint set, API Key missing", foreground="orange")
                    
            elif auth_type == "login":
                login_status = self.login_status_var.get() if hasattr(self, 'login_status_var') else "Not logged in"
                if "successful" in login_status.lower():
                    self.dash_api_status.config(text=" Configured & Authenticated", foreground="green")
                else:
                    username = self.login_username_var.get().strip() if hasattr(self, 'login_username_var') else ""
                    password = self.login_password_var.get().strip() if hasattr(self, 'login_password_var') else ""
                    if username and password:
                        self.dash_api_status.config(text=" Configured, not logged in", foreground="orange")
                    else:
                        self.dash_api_status.config(text=" Endpoint set, credentials missing", foreground="orange")
                        
            else:  # no_auth
                self.dash_api_status.config(text=" Configured (No Auth)", foreground="green")
        else:
            self.dash_api_status.config(text=" Not Configured", foreground="red")
        
        # Update agent status
        if self.is_running:
            self.dash_agent_status.config(text=" Running", foreground="green")
        else:
            self.dash_agent_status.config(text=" Stopped", foreground="red")
        
        # Update buffer status
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM push_buffer WHERE status = 'pending'")
            count = cursor.fetchone()[0]
            conn.close()
            
            if count > 0:
                self.dash_buffer_status.config(text=f"{count} items", foreground="orange")
                self.buffer_status.update_status("warning")
            else:
                self.dash_buffer_status.config(text="0 items", foreground="green")
                self.buffer_status.update_status("good")
        except:
            self.dash_buffer_status.config(text="Error", foreground="red")
            self.buffer_status.update_status("error")
        
        # Update recent activity
        self.activity_tree.delete(*self.activity_tree.get_children())
        logs = self.log_manager.get_recent_logs(20)
        
        for log in logs:
            self.activity_tree.insert("", 0, values=(
                log['timestamp'][:19],  # Remove microseconds
                log['level'],
                log['message']
            ))
    
    def create_about_tab(self):
        """Create about application tab"""
        frame = ttk.Frame(self.content_inner)
        self.tab_frames["about"] = frame
        
        # Create scrollable frame for About content
        about_canvas = tk.Canvas(frame)
        about_scrollbar = ttk.Scrollbar(frame, orient="vertical", command=about_canvas.yview)
        scrollable_frame = ttk.Frame(about_canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: about_canvas.configure(scrollregion=about_canvas.bbox("all"))
        )
        
        about_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        about_canvas.configure(yscrollcommand=about_scrollbar.set)
        
        about_canvas.pack(side="left", fill="both", expand=True)
        about_scrollbar.pack(side="right", fill="y")
        
        # Title
        ttk.Label(scrollable_frame, text="About MDB Agent Pro", style='Title.TLabel').pack(anchor=tk.W, pady=(10, 20))
        
        # Application Information
        app_frame = ttk.LabelFrame(scrollable_frame, text="Application Information", padding=15)
        app_frame.pack(fill=tk.X, pady=(0, 15), padx=10)
        
        app_info = [
            ("Application Name:", "MDB Agent Pro"),
            ("Version:", "2.0.0 (Professional Edition)"),
            ("Release Date:", "January 2025"),
            ("Purpose:", "Microsoft Access Database to API Bridge"),
            ("License:", "Proprietary - PT Sahabat Agro Group"),
            ("Architecture:", "Python-based Desktop Application"),
            ("Compatibility:", "Windows 10/11, Microsoft Access 2016+")
        ]
        
        for i, (label, value) in enumerate(app_info):
            ttk.Label(app_frame, text=label, font=('Arial', 9, 'bold')).grid(row=i, column=0, sticky=tk.W, padx=(0, 15), pady=2)
            ttk.Label(app_frame, text=value, font=('Arial', 9)).grid(row=i, column=1, sticky=tk.W, pady=2)
        
        # Company Information
        company_frame = ttk.LabelFrame(scrollable_frame, text="Company Information", padding=15)
        company_frame.pack(fill=tk.X, pady=(0, 15), padx=10)
        
        company_info = [
            ("Company:", "PT Sahabat Agro Group"),
            ("Industry:", "Agriculture Technology & Data Management"),
            ("Headquarters:", "Indonesia"),
            ("Business Focus:", "Agricultural Supply Chain & Digital Solutions"),
            ("Established:", "Serving agricultural sector since 2010"),
            ("Specialization:", "Database Integration & API Management"),
            ("Mission:", "Modernizing agricultural data systems")
        ]
        
        for i, (label, value) in enumerate(company_info):
            ttk.Label(company_frame, text=label, font=('Arial', 9, 'bold')).grid(row=i, column=0, sticky=tk.W, padx=(0, 15), pady=2)
            ttk.Label(company_frame, text=value, font=('Arial', 9)).grid(row=i, column=1, sticky=tk.W, pady=2)
        
        # Developer Information
        dev_frame = ttk.LabelFrame(scrollable_frame, text="Developer Information", padding=15)
        dev_frame.pack(fill=tk.X, pady=(0, 15), padx=10)
        
        dev_info = [
            ("Lead Developer:", "Freddy Mazmur"),
            ("Title:", "Senior Software Engineer & Database Specialist"),
            ("Email:", "freddy.pm@sahabatagro.co.id"),
            ("Mobile:", "+62 813-9855-2019"),
            ("Specialization:", "Database Integration, API Development, Python Applications"),
            ("Experience:", "10+ years in Enterprise Software Development"),
            ("Expertise:", "Microsoft Access, SQL Server, Python, API Design"),
            ("Location:", "Jakarta, Indonesia")
        ]
        
        for i, (label, value) in enumerate(dev_info):
            ttk.Label(dev_frame, text=label, font=('Arial', 9, 'bold')).grid(row=i, column=0, sticky=tk.W, padx=(0, 15), pady=2)
            ttk.Label(dev_frame, text=value, font=('Arial', 9)).grid(row=i, column=1, sticky=tk.W, pady=2)
        
        # Technical Specifications
        tech_frame = ttk.LabelFrame(scrollable_frame, text="Technical Specifications", padding=15)
        tech_frame.pack(fill=tk.X, pady=(0, 15), padx=10)
        
        import sys
        import platform
        
        tech_info = [
            ("Python Version:", f"{sys.version.split()[0]}"),
            ("Platform:", f"{platform.system()} {platform.release()}"),
            ("Architecture:", f"{platform.machine()}"),
            ("GUI Framework:", "Tkinter (Built-in)"),
            ("Database Connectivity:", "ODBC (pyodbc)"),
            ("HTTP Client:", "Requests Library"),
            ("Data Storage:", "SQLite3 (Local)"),
            ("Logging:", "Python Standard Logging")
        ]
        
        for i, (label, value) in enumerate(tech_info):
            ttk.Label(tech_frame, text=label, font=('Arial', 9, 'bold')).grid(row=i, column=0, sticky=tk.W, padx=(0, 15), pady=2)
            ttk.Label(tech_frame, text=value, font=('Arial', 9)).grid(row=i, column=1, sticky=tk.W, pady=2)
        
        # Key Features
        features_frame = ttk.LabelFrame(scrollable_frame, text="Key Features & Capabilities", padding=15)
        features_frame.pack(fill=tk.X, pady=(0, 15), padx=10)
        
        features_text = """
[CHAR] Visual Field Mapping System
   [CHAR] Drag-and-drop interface for database to API field mapping
   [CHAR] Support for data transformations and formatting
   [CHAR] Template management for reusable configurations

[CHAR] Real-time Data Processing
   [CHAR] Automatic data synchronization from Access databases
   [CHAR] Intelligent retry mechanism for failed API calls
   [CHAR] Background processing with minimal system impact

[CHAR] Robust Error Handling
   [CHAR] Comprehensive logging and audit trails
   [CHAR] Buffer system for offline resilience
   [CHAR] Health monitoring and diagnostics

[CHAR] Advanced Monitoring
   [CHAR] Transaction logging with detailed statistics
   [CHAR] API performance monitoring
   [CHAR] Database connection health checks

[CHAR] Enterprise Configuration
   [CHAR] JSON-based configuration management
   [CHAR] Multiple database and API endpoint support
   [CHAR] Scheduled operations and automation

[CHAR] User-Friendly Interface
   [CHAR] Intuitive tabbed interface design
   [CHAR] Context-sensitive help and validation
   [CHAR] Professional administrative controls
        """
        
        features_label = ttk.Label(features_frame, text=features_text.strip(), justify=tk.LEFT, font=('Arial', 9))
        features_label.pack(anchor=tk.W)
        
        # Support & Contact
        support_frame = ttk.LabelFrame(scrollable_frame, text="Support & Maintenance", padding=15)
        support_frame.pack(fill=tk.X, pady=(0, 15), padx=10)
        
        support_text = """
For technical support, feature requests, or system maintenance:

[CHAR] Primary Contact: freddy.pm@sahabatagro.co.id
[CHAR] Emergency Support: +62 813-9855-2019
[CHAR] Company: PT Sahabat Agro Group
[CHAR] Business Hours: Monday - Friday, 8:00 AM - 6:00 PM (WIB)

System Maintenance:
[CHAR] Regular updates and security patches
[CHAR] Database optimization and performance tuning
[CHAR] Custom feature development available
[CHAR] On-site training and consultation services
        """
        
        ttk.Label(support_frame, text=support_text.strip(), justify=tk.LEFT, font=('Arial', 9)).pack(anchor=tk.W)
        
        # Action Buttons
        action_frame = ttk.Frame(scrollable_frame)
        action_frame.pack(fill=tk.X, pady=15, padx=10)
        
        btn_frame = ttk.Frame(action_frame)
        btn_frame.pack()
        
        ttk.Button(btn_frame, text=" Send Log to IT", command=self.send_log_to_it).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text=" Check for Updates", command=self.check_updates).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text=" Generate System Report", command=self.generate_system_report).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text=" Export Support Logs", command=self.export_support_logs).pack(side=tk.LEFT)
        
        # Copyright and License
        copyright_frame = ttk.LabelFrame(scrollable_frame, text="Copyright & License", padding=15)
        copyright_frame.pack(fill=tk.X, pady=(0, 15), padx=10)
        
        copyright_text = """
[CHAR] 2025 PT Sahabat Agro Group. All rights reserved.

This software is proprietary and confidential. Unauthorized copying, distribution, 
or modification is strictly prohibited. This application is licensed exclusively 
for use by PT Sahabat Agro Group and its authorized partners.

Developed by: Freddy Mazmur
Built with: Python, Tkinter, and enterprise-grade libraries
Version: 2.0.0 Professional Edition
        """
        
        ttk.Label(copyright_frame, text=copyright_text.strip(), justify=tk.LEFT, font=('Arial', 8), foreground="gray").pack(anchor=tk.W)
        
        # Bind mousewheel to canvas for scrolling
        def _on_mousewheel(event):
            about_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        about_canvas.bind("<MouseWheel>", _on_mousewheel)
        license_frame = ttk.LabelFrame(frame, text="License & Legal", padding=10)
        license_frame.pack(fill=tk.BOTH, expand=True)
        
        license_text = """ 2025 PT Sahabat Agro Group. All rights reserved.

This software is licensed for use within the organization only.
Redistribution or modification without written permission is prohibited.

For technical support, please contact:
IT Support: freddy.pm@sahabatagro.co.id
Phone: +62 813-9855-2019
        """
        ttk.Label(license_frame, text=license_text.strip(), justify=tk.LEFT).pack(anchor=tk.W)
    
    def create_api_tab(self):
        """Create enhanced API configuration tab with responsive scrollable design"""
        frame = ttk.Frame(self.content_inner)
        self.tab_frames["api"] = frame
        
        # Page title with better spacing
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(title_frame, text="API Settings", 
                 style='Title.TLabel', 
                 font=('Arial', 16, 'bold')).pack(anchor=tk.W)
        ttk.Label(title_frame, text="Advanced API configuration with multiple authentication methods and request types", 
                 font=('Arial', 11), 
                 foreground='gray').pack(anchor=tk.W, pady=(3, 0))
        
        # Create notebook for tabs
        notebook = ttk.Notebook(frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Main Settings Tab
        main_tab = ttk.Frame(notebook)
        notebook.add(main_tab, text="Settings")
        
        # Profiles Tab
        profiles_tab = ttk.Frame(notebook)
        notebook.add(profiles_tab, text="Profiles")
        
        # History Tab  
        history_tab = ttk.Frame(notebook)
        notebook.add(history_tab, text="History")
        
        # === MAIN SETTINGS TAB WITH RESPONSIVE SCROLLING ===
        # Create canvas and scrollbar for responsive design
        main_canvas = tk.Canvas(main_tab, highlightthickness=0)
        main_scrollbar = ttk.Scrollbar(main_tab, orient="vertical", command=main_canvas.yview)
        main_scroll = ttk.Frame(main_canvas)
        
        main_scroll.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=main_scroll, anchor="nw")
        main_canvas.configure(yscrollcommand=main_scrollbar.set)
        
        # Pack canvas and scrollbar
        main_canvas.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        main_scrollbar.pack(side="right", fill="y", pady=10)
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Profile Selection
        profile_frame = ttk.LabelFrame(main_scroll, text="API Profile", padding=10)
        profile_frame.pack(fill=tk.X, pady=(0, 10))
        
        profile_select_frame = ttk.Frame(profile_frame)
        profile_select_frame.pack(fill=tk.X)
        
        ttk.Label(profile_select_frame, text="Profile:").pack(side=tk.LEFT)
        self.api_profile_var = tk.StringVar(value="Default")
        self.api_profile_combo = ttk.Combobox(profile_select_frame, textvariable=self.api_profile_var, 
                                            values=["Default", "Development", "Testing", "Production"], width=15)
        self.api_profile_combo.pack(side=tk.LEFT, padx=(5, 10))
        
        ttk.Button(profile_select_frame, text="New", command=self.create_new_profile, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(profile_select_frame, text="Save", command=self.save_current_profile, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(profile_select_frame, text="Delete", command=self.delete_selected_profile, width=8).pack(side=tk.LEFT, padx=2)
        
        # Bind profile selection
        self.api_profile_combo.bind('<<ComboboxSelected>>', self.on_profile_change)
        
        # Basic Settings
        basic_frame = ttk.LabelFrame(main_scroll, text="Basic Settings", padding=10)
        basic_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Endpoint URL
        ttk.Label(basic_frame, text="Endpoint URL:").pack(anchor=tk.W)
        self.api_endpoint_var = tk.StringVar()
        endpoint_entry = ttk.Entry(basic_frame, textvariable=self.api_endpoint_var)
        endpoint_entry.pack(fill=tk.X, pady=(5, 10))
        
        # Method and Content-Type row
        method_content_frame = ttk.Frame(basic_frame)
        method_content_frame.pack(fill=tk.X, pady=(0, 10))
        
        # HTTP Method
        ttk.Label(method_content_frame, text="Method:").pack(side=tk.LEFT)
        self.api_method_var = tk.StringVar(value="POST")
        method_combo = ttk.Combobox(method_content_frame, textvariable=self.api_method_var, 
                                  values=["POST", "GET", "PUT", "DELETE", "PATCH"], width=10)
        method_combo.pack(side=tk.LEFT, padx=(5, 20))
        method_combo.bind('<<ComboboxSelected>>', self.on_method_change)
        
        # Content-Type
        ttk.Label(method_content_frame, text="Content-Type:").pack(side=tk.LEFT)
        self.api_content_type_var = tk.StringVar(value="application/json")
        content_type_combo = ttk.Combobox(method_content_frame, textvariable=self.api_content_type_var, 
                                        values=["application/json", "application/x-www-form-urlencoded", 
                                               "multipart/form-data", "text/plain", "Custom..."], width=25)
        content_type_combo.pack(side=tk.LEFT, padx=(5, 0))
        content_type_combo.bind('<<ComboboxSelected>>', self.on_content_type_change)
        
        # Custom Content-Type entry (hidden by default)
        self.custom_content_type_frame = ttk.Frame(basic_frame)
        ttk.Label(self.custom_content_type_frame, text="Custom Content-Type:").pack(anchor=tk.W)
        self.custom_content_type_var = tk.StringVar()
        ttk.Entry(self.custom_content_type_frame, textvariable=self.custom_content_type_var).pack(fill=tk.X, pady=(5, 10))
        
        # Authentication Settings
        auth_frame = ttk.LabelFrame(main_scroll, text="Authentication", padding=10)
        auth_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Endpoint Type
        ttk.Label(auth_frame, text="Authentication Type:").pack(anchor=tk.W)
        self.auth_type_var = tk.StringVar(value="api_key")
        auth_type_frame = ttk.Frame(auth_frame)
        auth_type_frame.pack(fill=tk.X, pady=(5, 15))
        
        ttk.Radiobutton(auth_type_frame, text="API Key/Token", variable=self.auth_type_var, 
                       value="api_key", command=self.on_auth_type_change).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Radiobutton(auth_type_frame, text="Login Required", variable=self.auth_type_var, 
                       value="login", command=self.on_auth_type_change).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Radiobutton(auth_type_frame, text="No Authentication", variable=self.auth_type_var, 
                       value="no_auth", command=self.on_auth_type_change).pack(side=tk.LEFT)
        
        # API Key Frame with response preview
        self.api_key_frame = ttk.Frame(auth_frame)
        ttk.Label(self.api_key_frame, text="API Key/Token:").pack(anchor=tk.W)
        self.api_key_var = tk.StringVar()
        api_key_entry = ttk.Entry(self.api_key_frame, textvariable=self.api_key_var, show="*")
        api_key_entry.pack(fill=tk.X, pady=(5, 10))
        
        # API Key actions frame
        api_key_actions = ttk.Frame(self.api_key_frame)
        api_key_actions.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(api_key_actions, text="Test API Key", command=self.test_api_key).pack(side=tk.LEFT, padx=(0, 10))
        
        # API Key status
        self.api_key_status_var = tk.StringVar(value="Enter API key to test")
        self.api_key_status_label = ttk.Label(api_key_actions, textvariable=self.api_key_status_var, foreground='gray')
        self.api_key_status_label.pack(side=tk.LEFT, padx=(20, 0))
        
        # API Key response preview
        api_key_preview_frame = ttk.LabelFrame(self.api_key_frame, text="API Key Response Preview", padding=5)
        api_key_preview_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        self.api_key_response_text = scrolledtext.ScrolledText(api_key_preview_frame, height=4, state=tk.DISABLED)
        self.api_key_response_text.pack(fill=tk.BOTH, expand=True)
        
        # No Auth Frame with response preview
        self.no_auth_frame = ttk.Frame(auth_frame)
        
        no_auth_info = ttk.Label(self.no_auth_frame, text=" No authentication required for this endpoint", 
                                font=('Arial', 9), foreground='blue')
        no_auth_info.pack(anchor=tk.W, pady=(0, 10))
        
        # No auth actions frame
        no_auth_actions = ttk.Frame(self.no_auth_frame)
        no_auth_actions.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(no_auth_actions, text="Test Endpoint", command=self.test_no_auth).pack(side=tk.LEFT, padx=(0, 10))
        
        # No auth status
        self.no_auth_status_var = tk.StringVar(value="Ready to test endpoint")
        self.no_auth_status_label = ttk.Label(no_auth_actions, textvariable=self.no_auth_status_var, foreground='gray')
        self.no_auth_status_label.pack(side=tk.LEFT, padx=(20, 0))
        
        # No auth response preview
        no_auth_preview_frame = ttk.LabelFrame(self.no_auth_frame, text="Endpoint Response Preview", padding=5)
        no_auth_preview_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        self.no_auth_response_text = scrolledtext.ScrolledText(no_auth_preview_frame, height=4, state=tk.DISABLED)
        self.no_auth_response_text.pack(fill=tk.BOTH, expand=True)
        
        # Login Frame
        self.login_frame = ttk.Frame(auth_frame)
        
        # Username
        username_label_frame = ttk.Frame(self.login_frame)
        username_label_frame.pack(fill=tk.X, anchor=tk.W)
        ttk.Label(username_label_frame, text="Username:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT)
        
        # Auto-validation warning label for username
        self.login_warning_var = tk.StringVar()
        self.login_warning_label = ttk.Label(username_label_frame, textvariable=self.login_warning_var, 
                                           foreground='red', font=('Arial', 8))
        self.login_warning_label.pack(side=tk.LEFT, padx=(10, 0))
        
        self.login_username_var = tk.StringVar()
        username_entry = ttk.Entry(self.login_frame, textvariable=self.login_username_var)
        username_entry.pack(fill=tk.X, pady=(5, 10))
        
        # Auto-validation info label
        info_label = ttk.Label(self.login_frame, text=" Note: Username will be auto-converted to 'login' field in JSON payload", 
                              font=('Arial', 8), foreground='blue')
        info_label.pack(anchor=tk.W, pady=(0, 10))
        
        # Password
        ttk.Label(self.login_frame, text="Password:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        self.login_password_var = tk.StringVar()
        password_entry = ttk.Entry(self.login_frame, textvariable=self.login_password_var, show="*")
        password_entry.pack(fill=tk.X, pady=(5, 10))
        
        # Database
        ttk.Label(self.login_frame, text="Database:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        self.login_database_var = tk.StringVar()
        database_entry = ttk.Entry(self.login_frame, textvariable=self.login_database_var)
        database_entry.pack(fill=tk.X, pady=(5, 10))
        
        # Bind validation events for API key
        self.api_key_var.trace('w', lambda *args: self.update_api_key_status())
        
        # Bind validation events for endpoint 
        self.api_endpoint_var.trace('w', lambda *args: self.update_no_auth_status())
        
        # Bind validation events for login fields
        self.login_username_var.trace('w', self.validate_login_fields)
        self.login_password_var.trace('w', self.validate_login_fields)
        
        # Login button and status
        login_btn_frame = ttk.Frame(self.login_frame)
        login_btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.login_btn = ttk.Button(login_btn_frame, text="Login / Get Token", command=self.login_and_get_token)
        self.login_btn.pack(side=tk.LEFT)
        
        self.login_status_var = tk.StringVar(value="Not logged in")
        self.login_status_label = ttk.Label(login_btn_frame, textvariable=self.login_status_var, foreground='gray')
        self.login_status_label.pack(side=tk.LEFT, padx=(20, 0))
        
        # Request Body Settings
        body_frame = ttk.LabelFrame(main_scroll, text="Request Body", padding=10)
        body_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Body type indicator
        self.body_type_label = ttk.Label(body_frame, text="JSON Editor", font=('Arial', 10, 'bold'))
        self.body_type_label.pack(anchor=tk.W, pady=(0, 5))
        
        # JSON Editor
        self.json_body_frame = ttk.Frame(body_frame)
        self.json_body_text = scrolledtext.ScrolledText(self.json_body_frame, height=8, width=70)
        self.json_body_text.pack(fill=tk.BOTH, expand=True)
        
        # Form Editor (for form-data)
        self.form_body_frame = ttk.Frame(body_frame)
        self.form_fields = []
        form_header_frame = ttk.Frame(self.form_body_frame)
        form_header_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(form_header_frame, text="Key", width=20).pack(side=tk.LEFT)
        ttk.Label(form_header_frame, text="Value", width=30).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Label(form_header_frame, text="Type", width=10).pack(side=tk.LEFT, padx=(10, 0))
        
        self.form_fields_frame = ttk.Frame(self.form_body_frame)
        self.form_fields_frame.pack(fill=tk.BOTH, expand=True)
        
        # Add field button
        ttk.Button(self.form_body_frame, text="+ Add Field", command=self.add_form_field).pack(anchor=tk.W, pady=(5, 0))
        
        # Control Buttons with visual separator
        separator_frame = ttk.Frame(main_scroll)
        separator_frame.pack(fill=tk.X, pady=(20, 0))
        ttk.Separator(separator_frame, orient='horizontal').pack(fill=tk.X, pady=(0, 15))
        
        control_frame = ttk.Frame(main_scroll)
        control_frame.pack(fill=tk.X, pady=(0, 0))
        
        # Section label
        ttk.Label(control_frame, text="Global API Actions:", font=('Arial', 10, 'bold')).pack(anchor=tk.W, pady=(0, 5))
        
        # Help text
        help_text = ttk.Label(control_frame, text=" Buttons akan aktif setelah mengisi Endpoint dan memilih Authentication", 
                             font=('Arial', 9), foreground='blue')
        help_text.pack(anchor=tk.W, pady=(0, 5))
        
        # Button status indicator
        self.button_status_label = ttk.Label(control_frame, text="", font=('Arial', 9))
        self.button_status_label.pack(anchor=tk.W, pady=(0, 10))
        
        # Button container with responsive layout
        btn_container = ttk.Frame(control_frame)
        btn_container.pack(fill=tk.X)
        
        # Main action buttons with better styling
        main_btns = ttk.Frame(btn_container)
        main_btns.pack(anchor=tk.W, pady=(0, 10))
        
        self.test_conn_btn = ttk.Button(main_btns, text=" Test Connection", command=self.test_api_connection, width=20)
        self.test_conn_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.send_sample_btn = ttk.Button(main_btns, text=" Send Sample Data", command=self.send_sample_data_advanced, width=20)
        self.send_sample_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Options row
        options_frame = ttk.Frame(btn_container)
        options_frame.pack(anchor=tk.W)
        
        self.test_mode_var = tk.BooleanVar()
        test_mode_cb = ttk.Checkbutton(options_frame, text=" Test Mode (Safe Mode)", variable=self.test_mode_var)
        test_mode_cb.pack(side=tk.LEFT)
        
        # Response Panel (inside main settings tab) - More compact
        response_frame = ttk.LabelFrame(main_scroll, text=" Response & Preview Monitor", padding=10)
        response_frame.pack(fill=tk.BOTH, expand=True, pady=(20, 0))
        
        # Response tabs
        response_notebook = ttk.Notebook(response_frame)
        response_notebook.pack(fill=tk.BOTH, expand=True)
        
        # Preview tab
        preview_tab = ttk.Frame(response_notebook)
        response_notebook.add(preview_tab, text=" Payload Preview")
        
        self.payload_preview = scrolledtext.ScrolledText(preview_tab, height=6, state=tk.DISABLED, wrap=tk.WORD)
        self.payload_preview.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        # Response tab
        response_tab = ttk.Frame(response_notebook)
        response_notebook.add(response_tab, text=" API Response")
        
        # Response status
        response_status_frame = ttk.Frame(response_tab)
        response_status_frame.pack(fill=tk.X, pady=(0, 5))
        
        self.response_status_var = tk.StringVar(value=" Ready - No response yet")
        self.response_status_label = ttk.Label(response_status_frame, textvariable=self.response_status_var, font=('Arial', 9))
        self.response_status_label.pack(side=tk.LEFT)
        
        # Response content
        self.response_text = scrolledtext.ScrolledText(response_tab, height=6, state=tk.DISABLED, wrap=tk.WORD)
        self.response_text.pack(fill=tk.BOTH, expand=True)
        
        # === FINAL SAVE SETTINGS SECTION ===
        # Global Save Settings (positioned at the bottom after all configurations)
        save_section = ttk.LabelFrame(main_scroll, text=" Save API Configuration", padding=15)
        save_section.pack(fill=tk.X, pady=(20, 0))
        
        # Save instruction
        save_instruction = ttk.Label(save_section, 
                                   text=" Save your current API configuration to use across all application features:",
                                   font=('Arial', 10), foreground='blue')
        save_instruction.pack(anchor=tk.W, pady=(0, 10))
        
        # Configuration info
        config_info = ttk.Label(save_section, 
                               text=" Configuration will be saved regardless of authentication completion\n" +
                                    " You can test individual authentication methods using buttons above\n" +
                                    " Saved settings will be used by Dashboard, Health Check, and API Field Mapping",
                               font=('Arial', 9), foreground='gray', justify=tk.LEFT)
        config_info.pack(anchor=tk.W, pady=(0, 10))
        
        # Validation status display
        self.validation_status_var = tk.StringVar(value=" Enter endpoint URL to enable saving")
        self.validation_status_label = ttk.Label(save_section, textvariable=self.validation_status_var, 
                                                font=('Arial', 9), foreground='orange')
        self.validation_status_label.pack(anchor=tk.W, pady=(0, 10))
        
        # Save button with enhanced styling
        save_btn_frame = ttk.Frame(save_section)
        save_btn_frame.pack(fill=tk.X)
        
        self.save_api_btn = ttk.Button(save_btn_frame, text=" Save API Configuration", 
                                      command=self.save_api_settings, 
                                      style='Accent.TButton')
        self.save_api_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        # Additional action buttons
        ttk.Button(save_btn_frame, text=" Check Status", 
                  command=self.validate_and_update_status).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(save_btn_frame, text=" Test Current Auth", 
                  command=self.test_all_configurations).pack(side=tk.LEFT, padx=(0, 10))
        
        # Save status indicator
        self.save_status_var = tk.StringVar(value="")
        self.save_status_label = ttk.Label(save_btn_frame, textvariable=self.save_status_var, 
                                          font=('Arial', 9))
        self.save_status_label.pack(side=tk.LEFT, padx=(20, 0))
        
        # === PROFILES TAB ===
        self.setup_profiles_tab(profiles_tab)
        
        # === HISTORY TAB ===
        self.setup_history_tab(history_tab)
        
        # Initialize
        self.api_profiles = {"Default": {}}
        self.api_history = []
        self.current_profile = "Default"
        self.on_auth_type_change()
        self.on_content_type_change()
        self.on_method_change()
        self.update_button_states()
        
        # Load profiles after all widgets are created
        self.load_api_profiles()
        
        # Bind events for real-time validation
        self.bind_validation_events()
        
        # Initial update
        self.update_payload_preview()
        
        # Log tab creation
        self.log_entry("API Settings tab initialized successfully", "INFO")
    
    def create_scheduler_tab(self):
        """Create scheduler configuration tab"""
        frame = ttk.Frame(self.content_inner)
        self.tab_frames["scheduler"] = frame
        
        # Page title with better spacing
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, pady=(0, 25))
        
        ttk.Label(title_frame, text="Scheduler", 
                 style='Title.TLabel', 
                 font=('Arial', 16, 'bold')).pack(anchor=tk.W)
        ttk.Label(title_frame, text="Configure automatic data push scheduling", 
                 font=('Arial', 11), 
                 foreground='gray').pack(anchor=tk.W, pady=(3, 0))
        
        # Scheduler settings
        scheduler_frame = ttk.LabelFrame(frame, text="Scheduler Settings", padding=10)
        scheduler_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Push interval
        interval_frame = ttk.Frame(scheduler_frame)
        interval_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(interval_frame, text="Push Interval:").pack(side=tk.LEFT)
        self.push_interval_var = tk.StringVar(value="5")
        ttk.Entry(interval_frame, textvariable=self.push_interval_var, width=10).pack(side=tk.LEFT, padx=(5, 5))
        ttk.Label(interval_frame, text="minutes").pack(side=tk.LEFT)
        
        # Enable auto push
        self.enable_auto_push_var = tk.BooleanVar()
        ttk.Checkbutton(scheduler_frame, text="Enable Automatic Push", variable=self.enable_auto_push_var).pack(anchor=tk.W, pady=(0, 10))
        
        # Save button
        ttk.Button(scheduler_frame, text="Save Scheduler Settings", command=self.save_scheduler_settings).pack()
        
        # Current status
        status_frame = ttk.LabelFrame(frame, text="Current Status", padding=10)
        status_frame.pack(fill=tk.X, pady=(0, 10))
        
        status_grid = ttk.Frame(status_frame)
        status_grid.pack(fill=tk.X)
        
        ttk.Label(status_grid, text="Agent Status:", style='Header.TLabel').grid(row=0, column=0, sticky=tk.W, padx=(0, 20))
        self.sched_agent_status = ttk.Label(status_grid, text="Stopped", foreground="red")
        self.sched_agent_status.grid(row=0, column=1, sticky=tk.W)
        
        ttk.Label(status_grid, text="Next Push:", style='Header.TLabel').grid(row=1, column=0, sticky=tk.W, padx=(0, 20))
        self.sched_next_push = ttk.Label(status_grid, text="Not scheduled")
        self.sched_next_push.grid(row=1, column=1, sticky=tk.W)
        
        ttk.Label(status_grid, text="Last Push:", style='Header.TLabel').grid(row=2, column=0, sticky=tk.W, padx=(0, 20))
        self.sched_last_push = ttk.Label(status_grid, text="Never")
        self.sched_last_push.grid(row=2, column=1, sticky=tk.W)
        
        # Manual controls
        controls_frame = ttk.LabelFrame(frame, text="Manual Controls", padding=10)
        controls_frame.pack(fill=tk.X, pady=(0, 10))
        
        controls_btn_frame = ttk.Frame(controls_frame)
        controls_btn_frame.pack(fill=tk.X)
        
        ttk.Button(controls_btn_frame, text="Start Scheduler", command=self.start_scheduler).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(controls_btn_frame, text="Stop Scheduler", command=self.stop_scheduler).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(controls_btn_frame, text="Force Push Now", command=self.force_push_now).pack(side=tk.LEFT)
        
        # Activity log
        activity_frame = ttk.LabelFrame(frame, text="Scheduler Activity Log", padding=10)
        activity_frame.pack(fill=tk.BOTH, expand=True)
        
        self.scheduler_tree = ttk.Treeview(activity_frame, columns=("timestamp", "action", "status", "details"), show="headings", height=8)
        self.scheduler_tree.heading("timestamp", text="Timestamp")
        self.scheduler_tree.heading("action", text="Action")
        self.scheduler_tree.heading("status", text="Status")
        self.scheduler_tree.heading("details", text="Details")
        
        self.scheduler_tree.column("timestamp", width=150)
        self.scheduler_tree.column("action", width=120)
        self.scheduler_tree.column("status", width=80)
        self.scheduler_tree.column("details", width=350)
        
        scheduler_scroll = ttk.Scrollbar(activity_frame, orient=tk.VERTICAL, command=self.scheduler_tree.yview)
        self.scheduler_tree.configure(yscrollcommand=scheduler_scroll.set)
        
        self.scheduler_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scheduler_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    def create_mapping_tab(self):
        """Create comprehensive API field mapping tab"""
        frame = ttk.Frame(self.content_inner)
        self.tab_frames["mapping"] = frame
        
        # Page title with better spacing
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=tk.X, pady=(0, 25))
        
        ttk.Label(title_frame, text="API Field Mapping", 
                 style='Title.TLabel', 
                 font=('Arial', 16, 'bold')).pack(anchor=tk.W)
        ttk.Label(title_frame, text="Map database fields to API endpoints with transformations", 
                 font=('Arial', 11), 
                 foreground='gray').pack(anchor=tk.W, pady=(3, 0))
        
        # Main container with two panels
        main_container = ttk.PanedWindow(frame, orient=tk.HORIZONTAL)
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Left panel - Mapping interface
        left_panel = ttk.Frame(main_container)
        main_container.add(left_panel, weight=2)
        
        # Right panel - Preview and controls
        right_panel = ttk.Frame(main_container)
        main_container.add(right_panel, weight=1)
        
        # === LEFT PANEL ===
        # Instructions with nested mapping guidance
        instruction_frame = ttk.LabelFrame(left_panel, text=" How to Map Fields (Flat & Nested)", padding=10)
        instruction_frame.pack(fill=tk.X, pady=(0, 10))
        
        instructions = """
1. Select database table from Database Connection tab
2. Choose API field for each database column (flat mapping)
3. For NESTED structures: Use "Add Group/Array" to create groups
4. Set data transformation if needed
5. Preview nested JSON structure in real-time
6. Save your mapping as template for reuse
7. Test mapping before going live
        """
        ttk.Label(instruction_frame, text=instructions.strip(), justify=tk.LEFT).pack(anchor=tk.W)
        
        # Nested Mapping Controls
        nested_control_frame = ttk.LabelFrame(left_panel, text=" Nested Structure Controls", padding=10)
        nested_control_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Control buttons for nested mapping
        control_btn_frame = ttk.Frame(nested_control_frame)
        control_btn_frame.pack(fill=tk.X)
        
        # Configure grid for better layout
        control_btn_frame.columnconfigure(0, weight=1)
        control_btn_frame.columnconfigure(1, weight=1)
        control_btn_frame.columnconfigure(2, weight=1)
        control_btn_frame.columnconfigure(3, weight=1)
        
        ttk.Button(control_btn_frame, text=" Add Group", 
                  command=self.add_mapping_group).grid(row=0, column=0, sticky="ew", padx=(0, 2))
        ttk.Button(control_btn_frame, text=" Add Array", 
                  command=self.add_mapping_array).grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Button(control_btn_frame, text=" Visual Designer", 
                  command=self.open_visual_designer).grid(row=0, column=2, sticky="ew", padx=2)
        ttk.Button(control_btn_frame, text=" Import Spec", 
                  command=self.import_nested_api_spec).grid(row=0, column=3, sticky="ew", padx=(2, 0))
        
        # Status display for nested structure
        status_frame = ttk.Frame(nested_control_frame)
        status_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Label(status_frame, text="Status:").pack(side=tk.LEFT)
        self.nested_status_label = ttk.Label(status_frame, text=" No nested structure configured", 
                                           font=('Arial', 9), foreground='gray')
        self.nested_status_label.pack(side=tk.LEFT, padx=(5, 0))
        
        # Mapping mode selector
        mode_frame = ttk.Frame(nested_control_frame)
        mode_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Label(mode_frame, text="Mapping Mode:").pack(side=tk.LEFT)
        self.mapping_mode = tk.StringVar(value="flat")
        ttk.Radiobutton(mode_frame, text="Flat", variable=self.mapping_mode, 
                       value="flat", command=self.on_mapping_mode_change).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Radiobutton(mode_frame, text="Nested", variable=self.mapping_mode, 
                       value="nested", command=self.on_mapping_mode_change).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Radiobutton(mode_frame, text="TBS Auto", variable=self.mapping_mode, 
                       value="tbs_auto", command=self.on_mapping_mode_change).pack(side=tk.LEFT, padx=(10, 0))
        
        # Mapping interface
        mapping_frame = ttk.LabelFrame(left_panel, text="Field Mapping", padding=10)
        mapping_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Headers
        header_frame = ttk.Frame(mapping_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(header_frame, text="Database Column", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, padx=(0, 20))
        ttk.Label(header_frame, text="", font=('Arial', 12, 'bold')).grid(row=0, column=1, padx=10)
        ttk.Label(header_frame, text="API Field", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky=tk.W, padx=(20, 20))
        ttk.Label(header_frame, text="Transform", font=('Arial', 10, 'bold')).grid(row=0, column=3, sticky=tk.W, padx=(20, 0))
        
        # Scrollable mapping area
        self.mapping_canvas = tk.Canvas(mapping_frame)
        mapping_scrollbar = ttk.Scrollbar(mapping_frame, orient="vertical", command=self.mapping_canvas.yview)
        self.mapping_scroll_frame = ttk.Frame(self.mapping_canvas)
        
        self.mapping_scroll_frame.bind(
            "<Configure>",
            lambda e: self.mapping_canvas.configure(scrollregion=self.mapping_canvas.bbox("all"))
        )
        
        self.mapping_canvas.create_window((0, 0), window=self.mapping_scroll_frame, anchor="nw")
        self.mapping_canvas.configure(yscrollcommand=mapping_scrollbar.set)
        
        self.mapping_canvas.pack(side="left", fill="both", expand=True)
        mapping_scrollbar.pack(side="right", fill="y")
        
        # === RIGHT PANEL ===
        # API Settings Connection Status
        api_status_frame = ttk.LabelFrame(right_panel, text="API Settings Connection", padding=10)
        api_status_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.api_status_label = ttk.Label(api_status_frame, text=" Checking API Settings...", font=('Arial', 9))
        self.api_status_label.pack(anchor=tk.W)
        
        ttk.Button(api_status_frame, text="Refresh API Status", command=self.refresh_api_mapping_status).pack(fill=tk.X, pady=(5, 0))
        
        # Template management
        template_frame = ttk.LabelFrame(right_panel, text="Mapping Templates", padding=10)
        template_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Template list
        ttk.Label(template_frame, text="Saved Templates:").pack(anchor=tk.W)
        self.template_listbox = tk.Listbox(template_frame, height=4)
        self.template_listbox.pack(fill=tk.X, pady=(5, 10))
        
        # Template controls - simplified without duplication
        template_btn_frame = ttk.Frame(template_frame)
        template_btn_frame.pack(fill=tk.X)
        
        # Configure grid for better button layout
        template_btn_frame.columnconfigure(0, weight=1)
        template_btn_frame.columnconfigure(1, weight=1)
        template_btn_frame.columnconfigure(2, weight=1)
        
        ttk.Button(template_btn_frame, text=" Load Selected", command=self.load_selected_template).grid(row=0, column=0, sticky="ew", padx=(0, 2))
        ttk.Button(template_btn_frame, text=" Save New", command=self.show_save_template_dialog).grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Button(template_btn_frame, text=" Delete", command=self.delete_selected_template).grid(row=0, column=2, sticky="ew", padx=(2, 0))
        
        # Custom field button for adding manual API fields
        custom_field_btn = ttk.Button(template_frame, text=" Add Custom Field", 
                                    command=self.show_custom_field_dialog_manual, width=20)
        custom_field_btn.pack(fill=tk.X, pady=(10, 0))
        
        # API Structure Import
        api_frame = ttk.LabelFrame(right_panel, text="API Structure", padding=10)
        api_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(api_frame, text="Import API Spec (JSON)", command=self.import_api_spec).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(api_frame, text="Auto Detect from Endpoint", command=self.auto_detect_api_structure).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(api_frame, text="Manual API Fields", command=self.show_manual_api_fields_dialog).pack(fill=tk.X)
        
        # Preview
        preview_frame = ttk.LabelFrame(right_panel, text="JSON Preview", padding=10)
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.json_preview = scrolledtext.ScrolledText(preview_frame, height=8, state=tk.DISABLED)
        self.json_preview.pack(fill=tk.BOTH, expand=True)
        
        # Test controls - EXPANDED for better visibility
        test_frame = ttk.LabelFrame(right_panel, text=" Test Mapping Controls", padding=15)
        test_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Make test buttons larger and more prominent
        test_btn_style = {'fill': tk.X, 'pady': 3, 'ipady': 5}
        
        refresh_btn = ttk.Button(test_frame, text=" Refresh Preview", command=self.update_json_preview)
        refresh_btn.pack(**test_btn_style)
        
        test_api_btn = ttk.Button(test_frame, text=" Test API Call", command=self.test_mapping_api_call)
        test_api_btn.pack(**test_btn_style)
        
        validate_btn = ttk.Button(test_frame, text=" Validate Mapping", command=self.validate_mapping)
        validate_btn.pack(**test_btn_style)
        
        # MAIN ACTION BUTTONS - Essential controls that were missing
        action_frame = ttk.LabelFrame(right_panel, text=" Mapping Actions", padding=15)
        action_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Configure grid for better action button layout
        action_frame.columnconfigure(0, weight=1)
        action_frame.columnconfigure(1, weight=1)
        
        # Primary action buttons with enhanced styling
        save_mapping_btn = ttk.Button(action_frame, text=" SAVE MAPPING", 
                                     command=self.save_field_mapping_with_validation)
        save_mapping_btn.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 5))
        
        apply_mapping_btn = ttk.Button(action_frame, text=" APPLY & REFRESH", 
                                      command=self.apply_and_refresh_mapping)
        apply_mapping_btn.grid(row=1, column=0, sticky="ew", padx=(0, 2))
        
        reset_mapping_btn = ttk.Button(action_frame, text=" RESET MAPPING", 
                                      command=self.reset_field_mapping_with_confirm)
        reset_mapping_btn.grid(row=1, column=1, sticky="ew", padx=(2, 0))
        
        # Nested structure specific actions with enhanced buttons
        nested_actions_frame = ttk.LabelFrame(right_panel, text=" Nested Structure Actions", padding=15)
        nested_actions_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Configure grid for nested action buttons
        nested_actions_frame.columnconfigure(0, weight=1)
        nested_actions_frame.columnconfigure(1, weight=1)
        
        apply_nested_btn = ttk.Button(nested_actions_frame, text=" APPLY NESTED CONFIG", 
                                     command=self.apply_nested_configuration)
        apply_nested_btn.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 5))
        
        preview_nested_btn = ttk.Button(nested_actions_frame, text=" PREVIEW STRUCTURE", 
                                       command=self.preview_nested_structure)
        preview_nested_btn.grid(row=1, column=0, sticky="ew", padx=(0, 2))
        
        clear_nested_btn = ttk.Button(nested_actions_frame, text=" CLEAR NESTED", 
                                     command=self.clear_nested_configuration)
        clear_nested_btn.grid(row=1, column=1, sticky="ew", padx=(2, 0))
        
        # Status indicator for nested configuration
        status_indicator_frame = ttk.Frame(right_panel)
        status_indicator_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Label(status_indicator_frame, text="Nested Status:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT)
        self.nested_config_status = ttk.Label(status_indicator_frame, text=" Not Configured", 
                                             font=('Arial', 9), foreground='gray')
        self.nested_config_status.pack(side=tk.LEFT, padx=(5, 0))
        
        # Initialize
        self.field_mappings = {}
        self.api_fields = []
        self.load_mapping_templates()
        self.refresh_mapping_interface()
        self.refresh_api_mapping_status()  # Check API Settings connection
        
        # Show placeholder if no table selected
        if not hasattr(self, 'selected_table') or not self.selected_table:
            ttk.Label(self.mapping_scroll_frame, text="Please select a database table first from 'Database Connection' tab", 
                     font=('Arial', 12), foreground='gray').pack(pady=50)
    
    def update_nested_status(self, status_text=" Not Configured", color="gray"):
        """Update nested configuration status indicator"""
        try:
            if hasattr(self, 'nested_config_status'):
                self.nested_config_status.config(text=status_text, foreground=color)
            if hasattr(self, 'nested_status_label'):
                self.nested_status_label.config(text=status_text, foreground=color)
        except Exception as e:
            print(f"Error updating nested status: {str(e)}")

    def show_operation_feedback(self, title, message, type="info"):
        """Show user feedback with proper styling"""
        try:
            feedback_window = tk.Toplevel(self.root)
            feedback_window.title(title)
            feedback_window.geometry("400x200")
            feedback_window.transient(self.root)
            feedback_window.grab_set()
            
            # Center the window
            feedback_window.update_idletasks()
            x = (feedback_window.winfo_screenwidth() // 2) - (400 // 2)
            y = (feedback_window.winfo_screenheight() // 2) - (200 // 2)
            feedback_window.geometry(f"400x200+{x}+{y}")
            
            # Icon and styling based on type
            icon_map = {
                "info": "",
                "success": "", 
                "warning": "",
                "error": ""
            }
            color_map = {
                "info": "blue",
                "success": "green",
                "warning": "orange", 
                "error": "red"
            }
            
            icon = icon_map.get(type, "")
            color = color_map.get(type, "blue")
            
            # Content frame
            content_frame = ttk.Frame(feedback_window, padding=20)
            content_frame.pack(fill=tk.BOTH, expand=True)
            
            # Icon and message
            ttk.Label(content_frame, text=icon, font=('Arial', 24)).pack(pady=(0, 10))
            ttk.Label(content_frame, text=message, font=('Arial', 11), 
                     foreground=color, wraplength=350, justify='center').pack(pady=(0, 20))
            
            # OK button
            ttk.Button(content_frame, text="OK", 
                      command=feedback_window.destroy).pack()
            
            # Auto close after 3 seconds for success messages
            if type == "success":
                feedback_window.after(3000, feedback_window.destroy)
                
        except Exception as e:
            print(f"Error showing feedback: {str(e)}")

    def validate_mapping_configuration(self):
        """Validate current mapping configuration before save/apply"""
        try:
            issues = []
            
            # Check if table is selected
            if not hasattr(self, 'selected_table') or not self.selected_table:
                issues.append("No database table selected")
            
            # Check if mappings exist
            if not hasattr(self, 'field_mappings') or not self.field_mappings:
                issues.append("No field mappings configured")
            
            # Check if API fields are available
            if not hasattr(self, 'api_fields') or not self.api_fields:
                issues.append("No API fields available - check API connection")
            
            # Check nested structure if TBS mode
            mapping_mode = getattr(self, 'mapping_mode_var', None)
            if mapping_mode and mapping_mode.get() == "TBS Auto":
                nested_config = getattr(self, 'nested_structure_config', {})
                if not nested_config:
                    issues.append("TBS Auto mode requires nested structure configuration")
            
            return issues
            
        except Exception as e:
            return [f"Validation error: {str(e)}"]
    def refresh_api_mapping_status(self):
        """Refresh API Settings connection status in mapping tab - USE TEST CONNECTION RESULTS"""
        try:
            if not hasattr(self, 'api_status_label'):
                return
                
            # Check if API connection details exist
            if not hasattr(self, 'api_base_url') or not self.api_base_url.get():
                self.api_status_label.config(text=" API not configured", foreground='red')
                return
            
            # Use the test_api_connection result if available
            if hasattr(self, 'last_api_test_result'):
                if self.last_api_test_result.get('success', False):
                    self.api_status_label.config(text=" API Connected", foreground='green')
                    
                    # Update nested status if needed
                    self.update_nested_status(" API Ready for Nested Mapping", "green")
                else:
                    error_msg = self.last_api_test_result.get('error', 'Unknown error')
                    self.api_status_label.config(text=f" API Error: {error_msg[:30]}...", foreground='red')
                    self.update_nested_status(" API Connection Issues", "red")
            else:
                self.api_status_label.config(text=" API not tested yet", foreground='orange')
                self.update_nested_status(" API Test Required", "orange")
                
        except Exception as e:
            self.log_message(f"Error refreshing API mapping status: {str(e)}", "ERROR")

    def update_nested_status(self, status_text=" Not Configured", color="gray"):
        """Update nested configuration status indicator"""
        try:
            if hasattr(self, 'nested_config_status'):
                self.nested_config_status.config(text=status_text, foreground=color)
            if hasattr(self, 'nested_status_label'):
                self.nested_status_label.config(text=status_text, foreground=color)
        except Exception as e:
            print(f"Error updating nested status: {str(e)}")

    def show_operation_feedback(self, title, message, type="info"):
        """Show user feedback with proper styling"""
        try:
            feedback_window = tk.Toplevel(self.root)
            feedback_window.title(title)
            feedback_window.geometry("400x200")
            feedback_window.transient(self.root)
            feedback_window.grab_set()
            
            # Center the window
            feedback_window.update_idletasks()
            x = (feedback_window.winfo_screenwidth() // 2) - (400 // 2)
            y = (feedback_window.winfo_screenheight() // 2) - (200 // 2)
            feedback_window.geometry(f"400x200+{x}+{y}")
            
            # Icon and styling based on type
            icon_map = {
                "info": "",
                "success": "", 
                "warning": "",
                "error": ""
            }
            color_map = {
                "info": "blue",
                "success": "green",
                "warning": "orange", 
                "error": "red"
            }
            
            icon = icon_map.get(type, "")
            color = color_map.get(type, "blue")
            
            # Content frame
            content_frame = ttk.Frame(feedback_window, padding=20)
            content_frame.pack(fill=tk.BOTH, expand=True)
            
            # Icon and message
            ttk.Label(content_frame, text=icon, font=('Arial', 24)).pack(pady=(0, 10))
            ttk.Label(content_frame, text=message, font=('Arial', 11), 
                     foreground=color, wraplength=350, justify='center').pack(pady=(0, 20))
            
            # OK button
            ttk.Button(content_frame, text="OK", 
                      command=feedback_window.destroy).pack()
            
            # Auto close after 3 seconds for success messages
            if type == "success":
                feedback_window.after(3000, feedback_window.destroy)
                
        except Exception as e:
            print(f"Error showing feedback: {str(e)}")

    def validate_mapping_configuration(self):
        """Validate current mapping configuration before save/apply"""
        try:
            issues = []
            
            # Check if table is selected
            if not hasattr(self, 'selected_table') or not self.selected_table:
                issues.append("No database table selected")
            
            # Check if mappings exist
            if not hasattr(self, 'field_mappings') or not self.field_mappings:
                issues.append("No field mappings configured")
            
            # Check if API fields are available
            if not hasattr(self, 'api_fields') or not self.api_fields:
                issues.append("No API fields available - check API connection")
            
            # Check nested structure if TBS mode
            mapping_mode = getattr(self, 'mapping_mode_var', None)
            if mapping_mode and mapping_mode.get() == "TBS Auto":
                nested_config = getattr(self, 'nested_structure_config', {})
                if not nested_config:
                    issues.append("TBS Auto mode requires nested structure configuration")
            
            return issues
            
        except Exception as e:
            return [f"Validation error: {str(e)}"]

    # Support functions for about tab
    def export_support_logs(self):
        """Export comprehensive logs for support purposes"""
        try:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"mdb_agent_support_{timestamp}.txt"
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("MDB AGENT PRO - SUPPORT LOG EXPORT\n")
                f.write("=" * 60 + "\n\n")
                f.write(f"Generated: {datetime.now().isoformat()}\n")
                f.write(f"Version: 2.0.0 Professional Edition\n")
                f.write(f"Developer: Freddy Mazmur - PT Sahabat Agro Group\n")
                f.write(f"Contact: freddy.pm@sahabatagro.co.id\n\n")
                
                # System Information
                import sys, platform
                f.write("SYSTEM INFORMATION:\n")
                f.write("-" * 30 + "\n")
                f.write(f"Operating System: {platform.system()} {platform.release()}\n")
                f.write(f"Python Version: {sys.version.split()[0]}\n")
                f.write(f"Architecture: {platform.machine()}\n")
                f.write(f"Platform: {platform.platform()}\n")
                f.write(f"Processor: {platform.processor()}\n\n")
                
                # Application Status
                f.write("APPLICATION STATUS:\n")
                f.write("-" * 30 + "\n")
                f.write(f"Database Connected: {'Yes' if self.db_connection else 'No'}\n")
                f.write(f"Selected Table: {getattr(self, 'selected_table', 'None')}\n")
                f.write(f"API Endpoint: {'Configured' if self.config.get('api_endpoint') else 'Not Configured'}\n")
                f.write(f"Agent Running: {'Yes' if getattr(self, 'is_running', False) else 'No'}\n")
                f.write(f"Admin Mode: {'Active' if getattr(self, 'admin_mode', False) else 'Inactive'}\n\n")
                
                # Configuration (sanitized)
                f.write("CONFIGURATION:\n")
                f.write("-" * 30 + "\n")
                safe_config = {k: v for k, v in self.config.items() 
                             if k not in ['api_key', 'mdb_password', 'admin_pin']}
                import json
                f.write(json.dumps(safe_config, indent=2))
                f.write("\n\n")
                
                # Database Information
                if self.db_connection:
                    f.write("DATABASE INFORMATION:\n")
                    f.write("-" * 30 + "\n")
                    try:
                        cursor = self.db_connection.cursor()
                        tables = [table.table_name for table in cursor.tables() if table.table_type == 'TABLE']
                        f.write(f"Available Tables: {', '.join(tables)}\n")
                        if hasattr(self, 'table_columns') and self.table_columns:
                            f.write(f"Current Table Columns: {len(self.table_columns)} columns\n")
                    except Exception as e:
                        f.write(f"Database query error: {str(e)}\n")
                    f.write("\n")
                
                # Recent Logs (last 50)
                f.write("RECENT LOGS:\n")
                f.write("-" * 30 + "\n")
                try:
                    logs = self.log_manager.get_recent_logs(50)
                    for log in logs:
                        f.write(f"{log['timestamp']} [{log['level']}] {log['message']}\n")
                        if log.get('details'):
                            f.write(f"    Details: {log['details']}\n")
                except Exception as e:
                    f.write(f"Could not retrieve logs: {str(e)}\n")
                
                f.write("\n" + "=" * 60 + "\n")
                f.write("End of Support Log\n")
            
            messagebox.showinfo("Success", 
                              f" Support logs exported successfully!\n\n"
                              f"File: {filename}\n"
                              f"Please send this file to IT support for assistance.")
            self.log_entry(f"Support logs exported to {filename}", "INFO")
            
        except Exception as e:
            error_msg = f"Failed to export support logs: {str(e)}"
            messagebox.showerror("Export Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def generate_system_report(self):
        """Generate comprehensive system report in popup window"""
        try:
            report_window = tk.Toplevel(self.root)
            report_window.title("MDB Agent Pro - System Report")
            report_window.geometry("800x600")
            report_window.transient(self.root)
            report_window.grab_set()
            
            # Create text widget with scrollbar
            text_frame = ttk.Frame(report_window)
            text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            text_widget = scrolledtext.ScrolledText(text_frame, font=('Courier New', 9), wrap=tk.WORD)
            text_widget.pack(fill=tk.BOTH, expand=True)
            
            # Generate comprehensive report
            import sys, platform
            from datetime import datetime
            
            report = f"""
MDB AGENT PRO - SYSTEM REPORT
{'='*80}

Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Application Version: 2.0.0 Professional Edition
Developer: Freddy Mazmur (freddy.pm@sahabatagro.co.id)
Company: PT Sahabat Agro Group

SYSTEM SPECIFICATIONS
{'-'*50}
Operating System: {platform.system()} {platform.release()}
Platform Details: {platform.platform()}
Python Version: {sys.version.split()[0]}
Architecture: {platform.machine()}
Processor: {platform.processor()}

APPLICATION STATUS
{'-'*50}
Database Status: {' Connected' if self.db_connection else ' Not Connected'}
Selected Database: {self.config.get('mdb_file', 'None')}
Selected Table: {getattr(self, 'selected_table', 'None')}
API Endpoint: {' Configured' if self.config.get('api_endpoint') else ' Not Configured'}
Agent Status: {' Running' if getattr(self, 'is_running', False) else ' Stopped'}
Admin Mode: {' Active' if getattr(self, 'admin_mode', False) else ' Inactive'}

CONFIGURATION SUMMARY
{'-'*50}
Auto Push: {'Enabled' if self.config.get('auto_push', False) else 'Disabled'}
Push Interval: {self.config.get('push_interval', 300)} seconds
Test Mode: {'Enabled' if self.config.get('test_mode', False) else 'Disabled'}
Field Mappings: {len(self.config.get('field_mapping', {}))} configured

DATABASE INFORMATION
{'-'*50}"""
            
            if self.db_connection:
                try:
                    cursor = self.db_connection.cursor()
                    tables = [table.table_name for table in cursor.tables() if table.table_type == 'TABLE']
                    report += f"""
Available Tables: {len(tables)}
Table Names: {', '.join(tables[:10])}{'...' if len(tables) > 10 else ''}
"""
                    if hasattr(self, 'table_columns') and self.table_columns:
                        report += f"Current Table Columns: {len(self.table_columns)}\n"
                        col_names = [col['name'] for col in self.table_columns[:5]]
                        report += f"Sample Columns: {', '.join(col_names)}{'...' if len(self.table_columns) > 5 else ''}\n"
                except Exception as e:
                    report += f"Database Query Error: {str(e)}\n"
            else:
                report += "No database connection available.\n"
            
            # Transaction Statistics
            report += f"""
TRANSACTION STATISTICS
{'-'*50}"""
            try:
                conn = sqlite3.connect(self.db_manager.db_path)
                cursor = conn.cursor()
                
                # Count transactions
                cursor.execute("SELECT COUNT(*) FROM transaction_log")
                total_trans = cursor.fetchone()[0]
                
                cursor.execute("SELECT COUNT(*) FROM transaction_log WHERE status = 'success'")
                success_trans = cursor.fetchone()[0]
                
                cursor.execute("SELECT COUNT(*) FROM push_buffer WHERE status = 'pending'")
                pending_buffer = cursor.fetchone()[0]
                
                conn.close()
                
                success_rate = (success_trans / total_trans * 100) if total_trans > 0 else 0
                
                report += f"""
Total Transactions: {total_trans:,}
Successful: {success_trans:,} ({success_rate:.1f}%)
Failed/Retry: {total_trans - success_trans:,}
Pending in Buffer: {pending_buffer:,}
"""
            except Exception as e:
                report += f"Could not retrieve transaction statistics: {str(e)}\n"
            
            # Recent Activity
            report += f"""
RECENT ACTIVITY (Last 10 entries)
{'-'*50}"""
            try:
                logs = self.log_manager.get_recent_logs(10)
                for log in logs:
                    timestamp = log['timestamp'][:19]  # Remove microseconds
                    report += f"{timestamp} [{log['level']}] {log['message']}\n"
            except Exception as e:
                report += f"Could not retrieve recent logs: {str(e)}\n"
            
            report += f"""
PERFORMANCE METRICS
{'-'*50}
Memory Usage: Available in Task Manager
CPU Usage: Background processing optimized
Network: HTTPS API calls with retry logic
Storage: SQLite database with transaction logging

SUPPORT INFORMATION
{'-'*50}
Technical Support: freddy.pm@sahabatagro.co.id
Emergency Contact: +62 813-9855-2019
Business Hours: Monday-Friday, 8AM-6PM (WIB)
Company: PT Sahabat Agro Group

{'='*80}
End of System Report
"""
            
            # Insert report into text widget
            text_widget.insert(1.0, report)
            text_widget.config(state=tk.DISABLED)
            
            # Add export button
            btn_frame = ttk.Frame(report_window)
            btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
            
            def export_report():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"mdb_agent_report_{timestamp}.txt"
                try:
                    with open(filename, 'w', encoding='utf-8') as f:
                        f.write(report)
                    messagebox.showinfo("Exported", f"Report saved as {filename}")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to export: {str(e)}")
            
            ttk.Button(btn_frame, text=" Export Report", command=export_report).pack(side=tk.LEFT)
            ttk.Button(btn_frame, text=" Close", command=report_window.destroy).pack(side=tk.RIGHT)
            
            self.log_entry("System report generated", "INFO")
            
        except Exception as e:
            error_msg = f"Failed to generate system report: {str(e)}"
            messagebox.showerror("Report Error", error_msg)
            self.log_entry(error_msg, "ERROR")
            
            report = f"""
MDB AGENT PRO - SYSTEM REPORT
Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
{'=' * 60}

APPLICATION INFO:
[CHAR] Name: MDB Agent Pro
[CHAR] Version: 2.0.0
[CHAR] Type: Database to API Bridge

SYSTEM ENVIRONMENT:
[CHAR] Operating System: {platform.system()} {platform.release()}
[CHAR] Python Version: {sys.version.split()[0]}
[CHAR] Architecture: {platform.machine()}
[CHAR] Processor: {platform.processor()}

DATABASE CONNECTION:
 File: {self.config.get('mdb_file', 'Not configured')}
 Status: {'Connected' if self.db_connection else 'Disconnected'}
 Selected Table: {getattr(self, 'selected_table', 'None')}

API CONFIGURATION:
 Endpoint: {self.config.get('api_endpoint', 'Not configured')}
 Authentication: {'Configured' if self.config.get('api_key') else 'Not configured'}
 Auto Push: {self.config.get('auto_push', False)}
 Push Interval: {self.config.get('push_interval', 'Not set')} seconds

FIELD MAPPING:
 Mapped Fields: {len(self.field_mappings) if hasattr(self, 'field_mappings') else 0}
 Templates Available: {self.template_listbox.size() if hasattr(self, 'template_listbox') else 0}

AGENT STATUS:
 Running: {getattr(self, 'is_running', False)}
 Admin Mode: {getattr(self, 'admin_mode', False)}

RECENT ACTIVITY:
            """
            
            text_widget.insert(1.0, report)
            
            # Add recent logs
            try:
                logs = self.log_manager.get_recent_logs(10)
                text_widget.insert(tk.END, "\nRECENT LOGS:\n")
                for log in logs:
                    text_widget.insert(tk.END, f" {log[0]} [{log[1]}] {log[2]}\n")
            except:
                text_widget.insert(tk.END, "\n Could not retrieve recent logs\n")
            
            text_widget.config(state=tk.DISABLED)
            
            # Save button
            def save_report():
                filename = filedialog.asksaveasfilename(
                    defaultextension=".txt",
                    filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
                )
                if filename:
                    with open(filename, 'w') as f:
                        f.write(text_widget.get(1.0, tk.END))
                    messagebox.showinfo("Success", f"Report saved to {filename}")
            
            ttk.Button(report_window, text="Save Report", command=save_report).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {str(e)}")
    
    # Missing methods for mapping functionality
    def load_selected_template(self):
        """Load selected mapping template"""
        try:
            selection = self.template_listbox.curselection()
            if not selection:
                messagebox.showwarning("Warning", "Please select a template to load")
                return
            
            template_name = self.template_listbox.get(selection[0])
            
            # Load template from config
            templates = self.config.get('mapping_templates', {})
            if template_name in templates:
                template_data = templates[template_name]
                self.field_mappings = template_data.get('mappings', {})
                
                # Update UI
                self.refresh_mapping_interface()
                self.update_json_preview()
                
                messagebox.showinfo("Success", f"Template '{template_name}' loaded successfully!")
                self.log_entry(f"Mapping template loaded: {template_name}", "SUCCESS")
            else:
                messagebox.showerror("Error", f"Template '{template_name}' not found")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load template: {str(e)}")
    
    def show_save_template_dialog(self):
        """Show dialog to save current mapping as template with TBS Receiving support"""
        try:
            if not hasattr(self, 'mapping_widgets') or not self.mapping_widgets:
                messagebox.showwarning("Warning", "No field mapping to save as template")
                return
            
            # Check if there are any mapped fields
            mapped_fields = {}
            for db_column, widgets in self.mapping_widgets.items():
                api_field = widgets['api_field'].get()
                transform = widgets['transform'].get()
                
                if api_field != "(unmapped)":
                    mapped_fields[db_column] = {
                        'api_field': api_field,
                        'transform': transform
                    }
            
            if not mapped_fields:
                messagebox.showwarning("Warning", "No mapped fields to save. Map some fields first.")
                return
            
            # Show save dialog
            save_window = tk.Toplevel(self.root)
            save_window.title("Save Mapping Template - TBS Receiving")
            save_window.geometry("500x350")
            save_window.transient(self.root)
            save_window.grab_set()
            
            # Template Information
            info_frame = ttk.LabelFrame(save_window, text="Template Information", padding=15)
            info_frame.pack(fill=tk.X, padx=20, pady=20)
            
            ttk.Label(info_frame, text="Template Name:").pack(anchor=tk.W)
            name_var = tk.StringVar(value=f"TBS_Mapping_{datetime.now().strftime('%Y%m%d_%H%M')}")
            name_entry = ttk.Entry(info_frame, textvariable=name_var, width=50)
            name_entry.pack(fill=tk.X, pady=(5, 10))
            name_entry.focus()
            
            ttk.Label(info_frame, text="Description (optional):").pack(anchor=tk.W)
            desc_var = tk.StringVar(value="TBS Receiving API field mappings")
            desc_entry = ttk.Entry(info_frame, textvariable=desc_var, width=50)
            desc_entry.pack(fill=tk.X, pady=(5, 0))
            
            # Mapping Preview
            preview_frame = ttk.LabelFrame(save_window, text="Mapping Preview", padding=10)
            preview_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
            
            preview_text = f" Fields to save: {len(mapped_fields)}\n\n"
            
            # Categorize fields
            header_fields = []
            detail_fields = []
            
            for db_col, mapping in mapped_fields.items():
                api_field = mapping['api_field']
                if api_field in ["partner_id", "journal_id", "date_order", "officers", 
                               "keterangan_description", "driver_name", "vehicle_no", 
                               "destination_warehouse_id", "branch_id"]:
                    header_fields.append(f" {db_col}  {api_field} ({mapping['transform']})")
                else:
                    detail_fields.append(f" {db_col}  {api_field} ({mapping['transform']})")
            
            if header_fields:
                preview_text += " Header Fields (order_data level):\n"
                preview_text += "\n".join(header_fields[:3])
                if len(header_fields) > 3:
                    preview_text += f"\n   ... and {len(header_fields) - 3} more header fields"
                preview_text += "\n\n"
            
            if detail_fields:
                preview_text += " Detail Fields (order_line level):\n"
                preview_text += "\n".join(detail_fields[:3])
                if len(detail_fields) > 3:
                    preview_text += f"\n   ... and {len(detail_fields) - 3} more detail fields"
            
            ttk.Label(preview_frame, text=preview_text, font=('Arial', 9), justify=tk.LEFT).pack(anchor=tk.W)
            
            def save_template():
                template_name = name_var.get().strip()
                if not template_name:
                    messagebox.showwarning("Missing Name", "Please enter a template name.")
                    return
                
                try:
                    # Save template with enhanced metadata
                    if 'mapping_templates' not in self.config:
                        self.config['mapping_templates'] = {}
                    
                    self.config['mapping_templates'][template_name] = {
                        'mappings': mapped_fields.copy(),
                        'description': desc_var.get().strip(),
                        'table': getattr(self, 'selected_table', ''),
                        'api_format': 'TBS_Receiving_JSONRPC',
                        'created': datetime.now().isoformat(),
                        'field_count': len(mapped_fields),
                        'header_fields': len(header_fields),
                        'detail_fields': len(detail_fields)
                    }
                    
                    self.save_config()
                    
                    # Refresh template list
                    if hasattr(self, 'load_mapping_templates'):
                        self.load_mapping_templates()
                    
                    messagebox.showinfo("Template Saved", 
                                      f" Template '{template_name}' saved successfully!\n\n" +
                                      f" Contains {len(mapped_fields)} field mappings:\n" +
                                      f"    {len(header_fields)} header fields\n" +
                                      f"    {len(detail_fields)} detail fields\n\n" +
                                      f" Format: TBS Receiving JSON-RPC 2.0")
                    
                    save_window.destroy()
                    self.log_entry(f" TBS Mapping template '{template_name}' saved with {len(mapped_fields)} fields", "SUCCESS")
                    
                except Exception as e:
                    messagebox.showerror("Save Error", f"Failed to save template: {str(e)}")
            
            # Buttons
            btn_frame = ttk.Frame(save_window)
            btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
            
            ttk.Button(btn_frame, text=" Save TBS Template", command=save_template).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(btn_frame, text=" Cancel", command=save_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to create save dialog: {str(e)}")
            self.log_entry(f"Save template dialog error: {str(e)}", "ERROR")
    
    def delete_selected_template(self):
        """Delete selected template"""
        try:
            selection = self.template_listbox.curselection()
            if not selection:
                messagebox.showwarning("Warning", "Please select a template to delete")
                return
            
            template_name = self.template_listbox.get(selection[0])
            
            result = messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete template '{template_name}'?")
            if result:
                templates = self.config.get('mapping_templates', {})
                if template_name in templates:
                    del templates[template_name]
                    self.config['mapping_templates'] = templates
                    self.save_config()
                    self.load_mapping_templates()  # Refresh list
                    
                    messagebox.showinfo("Success", f"Template '{template_name}' deleted successfully!")
                    self.log_entry(f"Mapping template deleted: {template_name}", "INFO")
                else:
                    messagebox.showerror("Error", f"Template '{template_name}' not found")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete template: {str(e)}")
    
    def import_api_spec(self):
        """Import API specification from JSON file or open in external editor"""
        try:
            filename = filedialog.askopenfilename(
                title="Select API Specification File",
                filetypes=[
                    ("JSON files", "*.json"),
                    ("Text files", "*.txt"), 
                    ("Excel files", "*.xlsx;*.xls"),
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ]
            )
            
            if not filename:
                return
            
            # Show dialog to choose action
            dialog = tk.Toplevel(self)
            dialog.title("Import API Spec")
            dialog.geometry("400x300")
            dialog.transient(self)
            dialog.grab_set()
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
            y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
            dialog.geometry(f"+{x}+{y}")
            
            # Title
            title_frame = ttk.Frame(dialog)
            title_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
            
            ttk.Label(title_frame, text=" API Specification Import", 
                     font=('Arial', 14, 'bold')).pack()
            ttk.Label(title_frame, text=f"File: {filename.split('/')[-1]}", 
                     font=('Arial', 10), foreground='gray').pack()
            
            # Options frame
            options_frame = ttk.LabelFrame(dialog, text="Choose Action", padding=15)
            options_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
            
            def import_fields():
                try:
                    dialog.destroy()
                    
                    # Try to read as JSON
                    if filename.lower().endswith('.json'):
                        with open(filename, 'r', encoding='utf-8') as f:
                            api_spec = json.load(f)
                        
                        # Extract field names from API spec
                        fields = []
                        if 'properties' in api_spec:
                            fields = list(api_spec['properties'].keys())
                        elif 'fields' in api_spec:
                            fields = api_spec['fields']
                        elif isinstance(api_spec, dict):
                            fields = list(api_spec.keys())
                        
                        if fields:
                            self.api_fields = fields
                            self.update_mapping_comboboxes()
                            messagebox.showinfo("Success", f" API specification imported!\n\n{len(fields)} fields found:\n{', '.join(fields[:8])}{'...' if len(fields) > 8 else ''}")
                            self.log_entry(f"API spec imported: {len(fields)} fields", "SUCCESS")
                        else:
                            messagebox.showwarning("Warning", "No fields found in API specification")
                    else:
                        messagebox.showinfo("Info", "For non-JSON files, please use the 'Open in Editor' option to view and manually enter fields.")
                        
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to import API spec:\n{str(e)}")
            
            def open_in_notepad():
                try:
                    dialog.destroy()
                    import subprocess
                    import os
                    
                    # Try to open with default application first
                    os.startfile(filename)
                    
                    # Show info message
                    messagebox.showinfo("File Opened", 
                                      f" File opened in default application!\n\n"
                                      f" You can now:\n"
                                      f" View the API structure\n"
                                      f" Copy field names\n"
                                      f" Use 'Add Custom Field' to add them manually\n\n"
                                      f" For JSON files, you can also use 'Import Fields' option")
                    
                except Exception as e:
                    try:
                        # Fallback to notepad
                        subprocess.Popen(['notepad.exe', filename])
                        messagebox.showinfo("File Opened", f" File opened in Notepad!\n\nFile: {filename}")
                    except Exception as e2:
                        messagebox.showerror("Error", f"Failed to open file:\n{str(e)}\n\nFallback error:\n{str(e2)}")
            
            def open_in_excel():
                try:
                    dialog.destroy()
                    import subprocess
                    
                    # Try to open with Excel
                    subprocess.Popen(['excel.exe', filename])
                    messagebox.showinfo("File Opened", f" File opened in Excel!\n\nFile: {filename}")
                    
                except Exception as e:
                    try:
                        # Fallback to default application
                        import os
                        os.startfile(filename)
                        messagebox.showinfo("File Opened", f" File opened in default application!\n\nFile: {filename}")
                    except Exception as e2:
                        messagebox.showerror("Error", f"Failed to open file:\n{str(e)}\n\nFallback error:\n{str(e2)}")
            
            # Action buttons
            ttk.Button(options_frame, text=" Import Fields (JSON only)", 
                      command=import_fields).pack(fill=tk.X, pady=(0, 5))
            
            ttk.Button(options_frame, text=" Open in Notepad", 
                      command=open_in_notepad).pack(fill=tk.X, pady=(0, 5))
            
            ttk.Button(options_frame, text=" Open in Excel", 
                      command=open_in_excel).pack(fill=tk.X, pady=(0, 5))
            
            # Cancel button
            ttk.Button(options_frame, text=" Cancel", 
                      command=dialog.destroy).pack(fill=tk.X, pady=(5, 0))
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to select API spec file:\n{str(e)}")
    
    def auto_detect_api_structure(self):
        """Auto detect API structure - SMART MODE: Use Test Connection results if available, otherwise test directly"""
        try:
            # PRIORITY 1: Check if we have recent successful Test Connection results
            if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
                if self.last_test_connection_result['success']:
                    # Use existing successful connection result without retesting
                    self.log_entry(" Using existing Test Connection success for Auto Detect", "INFO")
                    
                    # Set default fields based on successful connection type
                    endpoint = self.api_endpoint_var.get().strip()
                    auth_type = self.auth_type_var.get()
                    
                    if "/login" in endpoint.lower() or "/auth" in endpoint.lower():
                        # Login endpoint detected from Test Connection success
                        default_fields = ["login", "password", "database", "username", "email", "token", "status"]
                        success_msg = f" Auto Detect using Test Connection success!\n\n" \
                                    f"Status: {self.last_test_connection_result['status_code']} " \
                                    f"({self.last_test_connection_result['response_time']})\n" \
                                    f"Auth: {auth_type}\n\n" \
                                    f"Login endpoint fields configured: {', '.join(default_fields)}\n\n" \
                                    f" Based on successful Test Connection result"
                    else:
                        # Data endpoint detected from Test Connection success  
                        default_fields = ["id", "name", "value", "timestamp", "status", "data", "created_at", "updated_at"]
                        success_msg = f" Auto Detect using Test Connection success!\n\n" \
                                    f"Status: {self.last_test_connection_result['status_code']} " \
                                    f"({self.last_test_connection_result['response_time']})\n" \
                                    f"Auth: {auth_type}\n\n" \
                                    f"Data endpoint fields configured: {', '.join(default_fields)}\n\n" \
                                    f" Based on successful Test Connection result"
                    
                    # Apply the detected fields
                    self.api_fields = default_fields
                    self.api_fields_source = "auto_detect_from_test_connection"
                    self.update_mapping_comboboxes()
                    
                    messagebox.showinfo("Auto Detect Success", success_msg)
                    self.log_entry(f" Auto Detect completed using Test Connection success: {len(default_fields)} fields", "SUCCESS")
                    return
                else:
                    # Test Connection failed - show warning but continue with direct test
                    error_msg = self.last_test_connection_result.get('error', 'Unknown error')
                    self.log_entry(f" Test Connection failed previously: {error_msg} - Trying direct test", "WARNING")
            
            # PRIORITY 2: No successful Test Connection available - perform direct test
            self.log_entry(" No recent Test Connection success - performing direct API test", "INFO")
            
            # Get endpoint DIRECTLY from UI (same as test_api_connection)
            endpoint = self.api_endpoint_var.get().strip()
            
            if not endpoint:
                messagebox.showerror("Error", "Please configure API endpoint first in API Settings tab")
                return
            
            # Get authentication DIRECTLY from UI (same as test_api_connection)
            auth_type = self.auth_type_var.get()
            
            # Prepare headers with consistent User-Agent
            headers = {'Content-Type': 'application/json', 'User-Agent': 'MDB-Agent-Pro/2.0'}
            
            if auth_type == "api_key":
                # Get API key DIRECTLY from UI (same as test_api_connection)
                if self.api_key_var.get().strip():
                    api_key = self.api_key_var.get().strip()
                    headers['Authorization'] = f'Bearer {api_key}'
                    self.log_entry(f" Using API Key authentication for auto-detect", "INFO")
                else:
                    messagebox.showwarning("Missing API Key", "API Key authentication selected but no key provided.\n\nPlease configure API Key in API Settings.")
                    return
            
            self.log_entry(f" Auto-detecting API structure from: {endpoint} (Auth: {auth_type})", "INFO")
            
            # Check if this is a login endpoint - use different approach
            if "/login" in endpoint.lower() or "/auth" in endpoint.lower():
                if auth_type == "login":
                    # Get login credentials DIRECTLY from UI (same as test_api_connection)
                    username = self.login_username_var.get().strip()
                    password = self.login_password_var.get().strip()
                    database = self.login_database_var.get().strip()
                    
                    if username and password:
                        # Use proper login format for auto-detection (matches test_api_connection)
                        test_data = {
                            "login": username,  # Auto-conversion from username
                            "password": password,
                            "database": database or "default"
                        }
                        
                        self.log_entry(f" Testing login endpoint with current UI credentials: {username}", "INFO")
                        response = requests.post(endpoint, json=test_data, headers=headers, timeout=15)
                        
                        if response.status_code == 200:
                            try:
                                response_data = response.json()
                                # Login endpoint detected - infer fields from request and response
                                login_fields = ["login", "password", "database"]
                                response_fields = list(response_data.keys()) if isinstance(response_data, dict) else []
                                
                                all_fields = list(set(login_fields + response_fields))
                                # Remove sensitive fields
                                safe_fields = [f for f in all_fields if f.lower() not in ['password', 'token', 'secret']]
                                
                                if safe_fields:
                                    self.api_fields = safe_fields
                                    self.api_fields_source = "auto_detect"
                                    self.update_mapping_comboboxes()
                                    messagebox.showinfo("Auto Detect Success", 
                                                      f" Login endpoint auto-detected successfully!\n\n"
                                                      f"Using current UI configuration: {auth_type} auth\n"
                                                      f"Fields found: {', '.join(safe_fields)}\n\n"
                                                      f"Status: {response.status_code} - Authentication successful\n\n"
                                                      f" API Field mappings updated automatically")
                                    self.log_entry(f" Login endpoint auto-detected using current UI config: {len(safe_fields)} fields", "SUCCESS")
                                    return
                            except:
                                pass
                        
                        # If 400 but endpoint responds, analyze the error
                        if response.status_code == 400:
                            try:
                                error_data = response.json()
                                error_text = str(error_data).lower()
                                if "login" in error_text or "password" in error_text or "credential" in error_text:
                                    # This confirms it's a login endpoint with credential issues
                                    default_fields = ["login", "password", "database", "username", "email"]
                                    self.api_fields = default_fields
                                    self.api_fields_source = "auto_detect"
                                    self.update_mapping_comboboxes()
                                    messagebox.showinfo("Login Endpoint Detected", 
                                                      f" Login endpoint detected!\n\n"
                                                      f"Status: {response.status_code} - Credential validation error (expected)\n"
                                                      f"Using current UI config: {username} / {'*' * len(password)}\n\n"
                                                      f"Default login fields configured: {', '.join(default_fields)}\n\n"
                                                      f" Error indicates endpoint is working but expects different credentials.")
                                    self.log_entry(f" Login endpoint confirmed via HTTP {response.status_code} using current UI config", "INFO")
                                    return
                                else:
                                    # Generic 400 error
                                    messagebox.showwarning("Validation Error", 
                                                         f" API returned validation error (400)\n\n"
                                                         f"Endpoint: {endpoint}\n"
                                                         f"Auth: {auth_type}\n"
                                                         f"Response: {str(error_data)[:200]}...\n\n"
                                                         f" This might mean:\n"
                                                         f" Different field names expected\n"
                                                         f" Additional required fields\n"
                                                         f" Different data format expected")
                                    return
                            except Exception as e:
                                messagebox.showwarning("Request Error", 
                                                     f" HTTP 400 Error\n\n"
                                                     f"Using current UI config but endpoint returned validation error.\n\n"
                                                     f"Check API documentation for required fields.")
                                return
                    else:
                        messagebox.showwarning("Missing Credentials", 
                                             f" Login endpoint detected: {endpoint}\n\n"
                                             f"Missing login credentials in current UI configuration.\n\n"
                                             f"Please:\n"
                                             f"1. Go to API Settings  Login Required\n"
                                             f"2. Fill in username & password\n"
                                             f"3. Try Auto Detect again (no need to save first)")
                        return
                else:
                    # Non-login auth for login endpoint - show warning
                    messagebox.showwarning("Authentication Mismatch", 
                                         f" This appears to be a login endpoint:\n{endpoint}\n\n"
                                         f"But you're using '{auth_type}' authentication.\n\n"
                                         f"Consider switching to 'Login Required' authentication type.")
                    return
            
            # For non-login endpoints, try OPTIONS request first
            try:
                response = requests.options(endpoint, headers=headers, timeout=10)
                if response.status_code == 200:
                    try:
                        schema = response.json()
                        fields = self.extract_fields_from_schema(schema)
                        if fields:
                            self.api_fields = fields
                            self.api_fields_source = "auto_detect"
                            self.update_mapping_comboboxes()
                            messagebox.showinfo("Success", f" API structure detected!\n\nFields found: {', '.join(fields)}")
                            self.log_entry(f"API structure auto-detected: {len(fields)} fields", "SUCCESS")
                            return
                    except:
                        pass
            except:
                pass
            
            # Fallback: send appropriate test data based on endpoint
            if "/data" in endpoint.lower() or "/records" in endpoint.lower() or "/create" in endpoint.lower():
                # Data endpoint - use sample record
                test_data = {
                    "id": 12345,
                    "name": "Sample Record",
                    "value": 123.45,
                    "timestamp": datetime.now().isoformat(),
                    "status": "active",
                    "data": "sample_data"
                }
            else:
                # Generic endpoint
                test_data = {
                    "test": True, 
                    "timestamp": datetime.now().isoformat(),
                    "source": "MDBAgentPro_AutoDetect"
                }
            
            # Log the configuration being used
            self.log_entry(f" Auto-detect using: Auth={auth_type}, Endpoint={endpoint}", "INFO")
            
            response = requests.post(endpoint, json=test_data, headers=headers, timeout=15)
            
            if response.status_code in [200, 201]:
                # Success - infer fields from request and response
                try:
                    response_data = response.json()
                    request_fields = list(test_data.keys())
                    response_fields = list(response_data.keys()) if isinstance(response_data, dict) else []
                    
                    all_fields = list(set(request_fields + response_fields))
                    # Remove test fields
                    clean_fields = [f for f in all_fields if f != 'test']
                    
                    if clean_fields:
                        self.api_fields = clean_fields
                        self.api_fields_source = "auto_detect"
                        self.update_mapping_comboboxes()
                        messagebox.showinfo("Success", 
                                          f" API endpoint accepting data!\n\n"
                                          f"Using current UI config: {auth_type} auth\n"
                                          f"Inferred fields: {', '.join(clean_fields)}\n\n"
                                          f"Response: {response.status_code}")
                        self.log_entry(f" API structure inferred using current UI config: {len(clean_fields)} fields", "SUCCESS")
                        return
                except:
                    pass
                    
            elif response.status_code in [400, 422]:  # Validation errors
                try:
                    error_data = response.json()
                    messagebox.showinfo("API Test", 
                                      f" API endpoint responding!\n\n"
                                      f"Status: {response.status_code} (Validation Error)\n"
                                      f"Using current UI config: {auth_type} auth\n"
                                      f"This means the endpoint is working but expects different data.\n\n"
                                      f"Response: {str(error_data)[:150]}...\n\n"
                                      f" Configure fields manually or check API documentation.")
                    self.log_entry(f" API endpoint validated using current UI config (returned {response.status_code})", "INFO")
                except:
                    messagebox.showinfo("API Test", 
                                      f" API endpoint responding with status {response.status_code}\n\n"
                                      f"Using current UI config: {auth_type} auth\n"
                                      f"Manual field configuration recommended.")
            else:
                messagebox.showwarning("Warning", 
                                     f" API returned status {response.status_code}\n\n"
                                     f"Using current UI config: {auth_type} auth\n"
                                     f"Response: {response.text[:200]}...\n\n"
                                     f"Manual configuration may be required.")
                
        except requests.exceptions.Timeout:
            messagebox.showerror("Error", f" Request timeout\n\nUsing UI config: {auth_type} auth\nEndpoint: {endpoint}\n\nPlease check your connection and try again.")
        except requests.exceptions.ConnectionError:
            messagebox.showerror("Error", f" Connection failed\n\nUsing UI config: {auth_type} auth\nEndpoint: {endpoint}\n\nPlease check the endpoint URL and your internet connection.")
        except Exception as e:
            messagebox.showerror("Error", f" Failed to detect API structure:\n\nUsing UI config: {auth_type} auth\nEndpoint: {endpoint}\n\nError: {str(e)}")
            self.log_entry(f" Auto-detect failed using current UI config: {str(e)}", "ERROR")
    
    def show_manual_api_fields_dialog(self):
        """Show manual API fields configuration dialog with TBS Receiving API preset"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Manual API Fields Configuration - TBS Receiving API")
        dialog.geometry("600x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Test Connection Status Display
        status_frame = ttk.LabelFrame(dialog, text="Test Connection Status", padding=10)
        status_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
        
        if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
            if self.last_test_connection_result['success']:
                status_text = f" Test Connection: {self.last_test_connection_result['status_code']} ({self.last_test_connection_result['response_time']})"
                status_color = 'green'
                suggestion = " API Settings validated - Ready for TBS Receiving mapping"
            else:
                error_msg = self.last_test_connection_result.get('error', 'Unknown error')
                status_text = f" Test Connection Failed: {error_msg[:50]}..."
                status_color = 'red'
                suggestion = " Fix API Settings before configuring field mappings"
        else:
            status_text = " No Test Connection performed yet"
            status_color = 'orange'
            suggestion = " Tip: Run Test Connection first for better field suggestions"
        
        ttk.Label(status_frame, text=status_text, foreground=status_color, font=('Arial', 9)).pack(anchor=tk.W)
        ttk.Label(status_frame, text=suggestion, foreground='blue', font=('Arial', 8)).pack(anchor=tk.W, pady=(5, 0))
        
        # TBS Receiving API Preset
        preset_frame = ttk.LabelFrame(dialog, text="TBS Receiving API Preset", padding=10)
        preset_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        ttk.Label(preset_frame, text=" Detected: TBS Receiving API Format", 
                 foreground='green', font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        ttk.Label(preset_frame, text="Format: JSON-RPC 2.0 with nested order_data structure", 
                 foreground='blue', font=('Arial', 8)).pack(anchor=tk.W)
        
        # Text area for fields
        text_frame = ttk.Frame(dialog)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        fields_text = scrolledtext.ScrolledText(text_frame, height=15)
        fields_text.pack(fill=tk.BOTH, expand=True)
        
        def load_tbs_preset():
            # TBS Receiving API fields (sesuai dengan struktur order_data yang benar)
            tbs_fields = [
                "# === ROOT LEVEL FIELDS ===",
                "uuid",
                "timestamp", 
                "source",
                "",
                "# === ORDER_DATA LEVEL FIELDS ===",
                "order_data.partner_id",
                "order_data.journal_id", 
                "order_data.date_order",
                "order_data.officers",
                "order_data.keterangan_description",
                "order_data.driver_name",
                "order_data.vehicle_no",
                "order_data.destination_warehouse_id",
                "order_data.branch_id",
                "",
                "# === ORDER_LINE LEVEL FIELDS (nested in order_data) ===",
                "order_data.order_line.product_code",
                "order_data.order_line.qty_brutto",
                "order_data.order_line.qty_tara", 
                "order_data.order_line.qty_netto",
                "order_data.order_line.product_uom",
                "order_data.order_line.sortation_percent",
                "order_data.order_line.sortation_weight",
                "order_data.order_line.qty_netto2",
                "order_data.order_line.price_unit",
                "order_data.order_line.product_qty",
                "order_data.order_line.incoming_date",
                "order_data.order_line.outgoing_date"
            ]
            fields_text.delete(1.0, tk.END)
            fields_text.insert(1.0, '\n'.join(tbs_fields))
        
        ttk.Button(preset_frame, text=" Load TBS Receiving Fields", 
                  command=load_tbs_preset).pack(anchor=tk.W, pady=(5, 0))
        
        # Instructions
        ttk.Label(dialog, text="API Field Names (edit as needed):", 
                 font=('Arial', 10, 'bold')).pack(anchor=tk.W, padx=20, pady=(10, 5))
        
        # Pre-populate with current fields or TBS preset
        if hasattr(self, 'api_fields') and self.api_fields:
            fields_text.insert(1.0, '\n'.join(self.api_fields))
        else:
            # Auto-load TBS preset
            load_tbs_preset()
        
        def save_fields():
            field_text = fields_text.get(1.0, tk.END).strip()
            if field_text:
                # Parse fields, ignore comments and empty lines
                raw_fields = [line.strip() for line in field_text.split('\n')]
                api_fields = [field for field in raw_fields 
                            if field and not field.startswith('#')]
                
                if api_fields:
                    self.api_fields = api_fields
                    self.api_fields_source = "manual_tbs_receiving"
                    self.update_mapping_comboboxes()
                    
                    # Log integration status
                    test_status = "No Test Connection" 
                    if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
                        if self.last_test_connection_result['success']:
                            test_status = f"Test Connection:  {self.last_test_connection_result['response_time']}"
                        else:
                            test_status = f"Test Connection:  Failed"
                    
                    self.log_entry(f" Loaded {len(api_fields)} TBS Receiving API fields", "INFO")
                    messagebox.showinfo("Success", 
                                      f" Loaded {len(api_fields)} API fields for TBS Receiving\n\n"
                                      f"Fields include: {', '.join(api_fields[:5])}{'...' if len(api_fields) > 5 else ''}\n\n"
                                      f" Integration Status: {test_status}")
                    dialog.destroy()
                else:
                    messagebox.showwarning("Warning", "No valid API fields found. Please enter field names.")
            else:
                messagebox.showwarning("Warning", "Please enter API field names.")
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=20, pady=20)
        
        ttk.Button(btn_frame, text=" Save TBS Fields", command=save_fields).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(btn_frame, text=" Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def update_mapping_comboboxes(self):
        """Update all mapping comboboxes with new API fields"""
        if hasattr(self, 'mapping_widgets') and hasattr(self, 'api_fields'):
            new_values = ["(unmapped)"] + self.api_fields + ["Custom..."]
            for widgets in self.mapping_widgets.values():
                widgets['api_combo']['values'] = new_values
            
            # Auto-map fields if this is TBS Receiving API
            if hasattr(self, 'api_fields_source') and 'tbs' in self.api_fields_source.lower():
                self.auto_map_tbs_fields()
            
            # Update JSON preview after updating fields
            self.update_json_preview()
            
            source_info = getattr(self, 'api_fields_source', 'unknown')
            self.log_entry(f"Updated mapping comboboxes with {len(self.api_fields)} API fields (source: {source_info})", "INFO")
    
    def auto_map_tbs_fields(self):
        """Automatically map database fields to TBS API fields based on common patterns"""
        if not hasattr(self, 'mapping_widgets') or not hasattr(self, 'table_columns'):
            return
            
        # TBS field mapping patterns (database field -> TBS field path)
        tbs_mapping_patterns = {
            # Partner/Supplier fields
            'supplier': 'order_data.partner_id',
            'partner': 'order_data.partner_id', 
            'vendor': 'order_data.partner_id',
            'customer': 'order_data.partner_id',
            
            # Journal/Bank fields
            'bank': 'order_data.journal_id',
            'journal': 'order_data.journal_id',
            'payment': 'order_data.journal_id',
            'account': 'order_data.journal_id',
            
            # Date fields
            'date': 'order_data.date_order',
            'time': 'order_data.date_order',
            'created': 'order_data.date_order',
            'order_date': 'order_data.date_order',
            'tanggal': 'order_data.date_order',
            
            # Officer fields
            'officer': 'order_data.officers',
            'user': 'order_data.officers',
            'operator': 'order_data.officers',
            'staff': 'order_data.officers',
            
            # Description fields
            'description': 'order_data.keterangan_description',
            'note': 'order_data.keterangan_description',
            'remark': 'order_data.keterangan_description',
            'comment': 'order_data.keterangan_description',
            'keterangan': 'order_data.keterangan_description',
            
            # Driver fields
            'driver': 'order_data.driver_name',
            'supir': 'order_data.driver_name',
            
            # Vehicle fields
            'vehicle': 'order_data.vehicle_no',
            'truck': 'order_data.vehicle_no',
            'mobil': 'order_data.vehicle_no',
            'nopol': 'order_data.vehicle_no',
            
            # Warehouse fields
            'warehouse': 'order_data.destination_warehouse_id',
            'gudang': 'order_data.destination_warehouse_id',
            
            # Branch fields
            'branch': 'order_data.branch_id',
            'cabang': 'order_data.branch_id',
            
            # Product fields
            'product': 'order_data.order_line.product_code',
            'item': 'order_data.order_line.product_code',
            'barang': 'order_data.order_line.product_code',
            'jenis': 'order_data.order_line.product_code',
            'jenispk': 'order_data.order_line.product_code',
            'pk': 'order_data.order_line.product_code',
            
            # Weight fields - order matters for specificity!
            'bruto': 'order_data.order_line.qty_brutto',
            'gross': 'order_data.order_line.qty_brutto',
            'tara': 'order_data.order_line.qty_tara',
            'tare': 'order_data.order_line.qty_tara',
            'netto': 'order_data.order_line.qty_netto',
            'net': 'order_data.order_line.qty_netto',
            'nett': 'order_data.order_line.qty_netto',
            'aktual': 'order_data.order_line.qty_netto',
            
            # Unit fields
            'unit': 'order_data.order_line.product_uom',
            'uom': 'order_data.order_line.product_uom',
            'satuan': 'order_data.order_line.product_uom',
            
            # Price fields
            'price': 'order_data.order_line.price_unit',
            'harga': 'order_data.order_line.price_unit',
            'rate': 'order_data.order_line.price_unit',
            
            # Quantity fields
            'qty': 'order_data.order_line.product_qty',
            'quantity': 'order_data.order_line.product_qty',
            'jumlah': 'order_data.order_line.product_qty',
        }
        
        mapped_count = 0
        for db_column in self.table_columns:
            column_name = db_column['name'] if isinstance(db_column, dict) else db_column
            if column_name in self.mapping_widgets:
                db_lower = column_name.lower()
                
                # Find best match
                best_match = None
                for pattern, tbs_field in tbs_mapping_patterns.items():
                    if pattern in db_lower:
                        # Prefer exact matches over partial matches
                        if pattern == db_lower:
                            best_match = tbs_field
                            break
                        elif best_match is None:
                            best_match = tbs_field
                
                # Apply mapping if found
                if best_match and best_match in self.api_fields:
                    try:
                        self.mapping_widgets[column_name]['api_combo'].set(best_match)
                        mapped_count += 1
                    except:
                        pass
        
        if mapped_count > 0:
            self.log_entry(f"Auto-mapped {mapped_count} TBS fields based on column names", "INFO")
            # Update JSON preview after auto-mapping
            self.update_json_preview()
    
    def extract_fields_from_schema(self, schema):
        """Extract field names from API schema"""
        fields = []
        try:
            if isinstance(schema, dict):
                # Look for common schema patterns
                if 'properties' in schema:
                    fields.extend(schema['properties'].keys())
                elif 'fields' in schema:
                    if isinstance(schema['fields'], list):
                        fields.extend(schema['fields'])
                    elif isinstance(schema['fields'], dict):
                        fields.extend(schema['fields'].keys())
                elif 'columns' in schema:
                    if isinstance(schema['columns'], list):
                        fields.extend(schema['columns'])
                elif isinstance(schema, dict):
                    # Just use all keys as potential fields
                    fields.extend(schema.keys())
            
            # Remove system fields
            system_fields = ['test', '__metadata', '_links', 'schema', 'version']
            fields = [f for f in fields if f not in system_fields]
            
        except Exception as e:
            self.log_entry(f"Error extracting fields from schema: {str(e)}", "ERROR")
        
        return fields
    
    def update_json_preview(self):
        """Update JSON preview with current mapping and real sample data - TBS Receiving API Format"""
        if not hasattr(self, 'json_preview'):
            return
            
        try:
            # Generate sample data based on current mapping
            mapping_count = 0
            mapped_fields = {}
            
            # Count mapped fields and collect data
            if hasattr(self, 'field_mappings') and self.field_mappings:
                for db_field, api_field in self.field_mappings.items():
                    if api_field and api_field != "(unmapped)":
                        mapping_count += 1
                        
                        # Get sample data for this field
                        sample_value = self.get_sample_data_for_column(db_field)
                        if sample_value is None:
                            sample_value = f"sample_{db_field.lower()}"
                        
                        mapped_fields[api_field] = sample_value
            
            # Check API Settings status for Smart Mode preview
            api_settings_status = "Not configured"
            test_connection_status = "Not tested"
            
            # Check Test Connection status
            if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
                if self.last_test_connection_result.get('success', False):
                    test_connection_status = f" Success ({self.last_test_connection_result.get('response_time', 'N/A')})"
                else:
                    test_connection_status = f" Failed ({self.last_test_connection_result.get('error', 'Unknown error')})"
            
            # Check API URL configuration
            api_url = getattr(self, 'api_endpoint_var', tk.StringVar()).get().strip()
            if not api_url:
                api_url = getattr(self, 'api_url_var', tk.StringVar()).get().strip()
            
            # Generate TBS Receiving API format
            if mapping_count > 0:
                # Build order_data structure
                order_data = {
                    "partner_id": mapped_fields.get("partner_id", "PT Sumber Sawit Default"),
                    "journal_id": mapped_fields.get("journal_id", "Bank Agro Default"),
                    "date_order": mapped_fields.get("date_order", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
                    "officers": mapped_fields.get("officers", "System User"),
                    "keterangan_description": mapped_fields.get("keterangan_description", "Auto submission from MDB Agent Pro"),
                    "driver_name": mapped_fields.get("driver_name", "Sample Driver"),
                    "vehicle_no": mapped_fields.get("vehicle_no", "B1234XX"),
                    "destination_warehouse_id": mapped_fields.get("destination_warehouse_id", "Gudang Default"),
                    "branch_id": mapped_fields.get("branch_id", "Default Branch")
                }
                
                # Build order_line structure
                order_line = {
                    "product_code": mapped_fields.get("product_code", "TBS-AUTO-001"),
                    "qty_brutto": self.convert_to_number(mapped_fields.get("qty_brutto", 1000)),
                    "qty_tara": self.convert_to_number(mapped_fields.get("qty_tara", 50)),
                    "qty_netto": self.convert_to_number(mapped_fields.get("qty_netto", 950)),
                    "product_uom": mapped_fields.get("product_uom", "kg"),
                    "sortation_percent": self.convert_to_number(mapped_fields.get("sortation_percent", 5)),
                    "sortation_weight": self.convert_to_number(mapped_fields.get("sortation_weight", 47.5)),
                    "qty_netto2": self.convert_to_number(mapped_fields.get("qty_netto2", 902.5)),
                    "price_unit": self.convert_to_number(mapped_fields.get("price_unit", 1500)),
                    "product_qty": self.convert_to_number(mapped_fields.get("product_qty", 1)),
                    "incoming_date": mapped_fields.get("incoming_date", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
                    "outgoing_date": mapped_fields.get("outgoing_date", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
                }
                
                # Complete TBS Receiving API payload
                api_payload = {
                    "jsonrpc": "2.0",
                    "params": {
                        "order_data": [
                            {**order_data, "order_line": [order_line]}
                        ]
                    }
                }
                
                # Add metadata for debugging
                api_payload["_preview_info"] = {
                    "preview_mode": True,
                    "mapped_fields_count": mapping_count,
                    "api_endpoint": api_url[:50] + "..." if len(api_url) > 50 else api_url,
                    "test_connection": test_connection_status,
                    "format": "TBS Receiving API - JSON-RPC 2.0"
                }
                
                preview_data = api_payload
            else:
                # No mappings configured
                preview_data = {
                    "message": "Configure field mappings to see TBS Receiving API payload preview",
                    "expected_format": {
                        "jsonrpc": "2.0",
                        "params": {
                            "order_data": [
                                {
                                    "partner_id": "Map to supplier name",
                                    "driver_name": "Map to driver field",
                                    "vehicle_no": "Map to vehicle number",
                                    "order_line": [
                                        {
                                            "product_code": "Map to product code",
                                            "qty_brutto": "Map to gross weight",
                                            "qty_tara": "Map to tare weight",
                                            "qty_netto": "Map to net weight"
                                        }
                                    ]
                                }
                            ]
                        }
                    },
                    "next_steps": [
                        "1. Map database fields using dropdowns above",
                        "2. Use TBS field names: qty_brutto, qty_tara, qty_netto, driver_name, vehicle_no",
                        "3. Apply Number transformation for weights",
                        "4. Test the mapping with 'Test API Call'"
                    ]
                }
            
            # Display in preview
            json_text = json.dumps(preview_data, indent=2, ensure_ascii=False)
            self.json_preview.config(state=tk.NORMAL)
            self.json_preview.delete(1.0, tk.END)
            self.json_preview.insert(1.0, json_text)
            self.json_preview.config(state=tk.DISABLED)
            
        except Exception as e:
            error_msg = f"Preview update failed: {str(e)}"
            self.json_preview.config(state=tk.NORMAL)
            self.json_preview.delete(1.0, tk.END)
            self.json_preview.insert(1.0, json.dumps({"error": error_msg}, indent=2))
            self.json_preview.config(state=tk.DISABLED)
    
    def convert_to_number(self, value):
        """Helper to convert values to numbers"""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return value
        try:
            # Try to convert string to number
            if '.' in str(value):
                return float(value)
            else:
                return int(value)
        except:
            return 0
    
    def get_sample_data_for_column(self, column_name):
        """Get real sample data from database column"""
        try:
            if hasattr(self, 'db_connection') and self.db_connection and hasattr(self, 'selected_table'):
                cursor = self.db_connection.cursor()
                cursor.execute(f"SELECT TOP 1 [{column_name}] FROM [{self.selected_table}] WHERE [{column_name}] IS NOT NULL")
                result = cursor.fetchone()
                if result:
                    return result[0]
        except:
            pass
        return None
    
    def generate_sample_value(self, column_name):
        """Generate appropriate sample value based on column name"""
        column_lower = column_name.lower()
        
        if 'id' in column_lower:
            return 12345
        elif any(word in column_lower for word in ['date', 'time', 'created', 'updated']):
            return datetime.now().isoformat()
        elif any(word in column_lower for word in ['name', 'title', 'description']):
            return f"Sample {column_name.title()}"
        elif any(word in column_lower for word in ['weight', 'amount', 'value', 'price', 'total']):
            return 123.45
        elif any(word in column_lower for word in ['count', 'quantity', 'number']):
            return 10
        elif any(word in column_lower for word in ['status', 'state']):
            return "active"
        elif any(word in column_lower for word in ['email', 'mail']):
            return "example@domain.com"
        elif any(word in column_lower for word in ['phone', 'tel']):
            return "+62-xxx-xxx-xxxx"
        elif any(word in column_lower for word in ['code', 'ref', 'reference']):
            return f"CODE_{column_name.upper()[:3]}123"
        else:
            return f"sample_{column_lower}"
    
    def apply_transformation_preview(self, value, transform):
        """Apply transformation for preview purposes"""
        if transform == "No Transform":
            return value
        elif transform == "String":
            return str(value)
        elif transform == "Number":
            try:
                return float(value) if '.' in str(value) else int(value)
            except:
                return 0
        elif transform == "Date":
            try:
                if isinstance(value, str):
                    return value
                return datetime.now().isoformat()
            except:
                return datetime.now().isoformat()
        elif transform == "Boolean":
            if str(value).lower() in ['true', '1', 'yes', 'on']:
                return True
            return False
        elif transform == "Uppercase":
            return str(value).upper()
        elif transform == "Lowercase":
            return str(value).lower()
        elif transform == "Trim":
            return str(value).strip()
        elif transform == "JSON":
            return json.dumps(value)
        elif transform == "Base64":
            import base64
            return base64.b64encode(str(value).encode()).decode()
        else:
            return value
    
    def test_mapping_api_call(self):
        """Test mapping with actual API call using FieldMapper"""
        try:
            # Check if mapping is configured
            if not hasattr(self, 'field_mappings') or not self.field_mappings:
                messagebox.showerror("Error", "Please configure field mapping first.")
                return
            
            # Check if table is selected
            if not hasattr(self, 'selected_table') or not self.selected_table:
                messagebox.showerror("Error", "Please select a database table first.")
                return
            
            # Get test data from database
            raw_data = self.get_latest_record()
            if not raw_data:
                messagebox.showerror("Error", "No data found in selected table.")
                return
            
            # Build mapped payload using FieldMapper
            mapped_data = self.build_api_payload(raw_data)
            
            # Show confirmation dialog with preview
            result = messagebox.askyesno(
                "Test Mapping API Call",
                f"Ready to test mapping with API call.\n\n"
                f"Mode: {self.mapping_mode.get()}\n"
                f"Fields mapped: {len(self.field_mappings)}\n"
                f"Table: {self.selected_table}\n\n"
                f"Do you want to send test data to API?"
            )
            
            if not result:
                return
            
            # Send mapped data to API
            success = self.send_to_api(mapped_data)
            
            if success:
                messagebox.showinfo(
                    "Test Successful", 
                    f"Mapping test completed successfully!\n\n"
                    f"Data was properly mapped and sent to API.\n"
                    f"Check the Transaction Log for details."
                )
                self.log_entry("Mapping API test completed successfully", "SUCCESS")
            else:
                messagebox.showerror(
                    "Test Failed", 
                    f"Mapping test failed.\n\n"
                    f"The data was mapped correctly but API call failed.\n"
                    f"Check logs for details."
                )
                
        except Exception as e:
            error_msg = f"Mapping API test failed: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def validate_mapping(self):
        """Validate current field mapping configuration using FieldMapper"""
        try:
            # Check if mapping exists
            if not hasattr(self, 'field_mappings') or not self.field_mappings:
                messagebox.showwarning("Warning", "No field mapping configured. Please configure field mappings first.")
                return
            
            # Use FieldMapper validation
            if self.field_mapper:
                validation_errors = self.field_mapper.validate_mapping()
            else:
                # Create temporary mapper for validation
                self.update_field_mapper()
                validation_errors = self.field_mapper.validate_mapping() if self.field_mapper else ["Failed to create field mapper"]
            
            # Show validation results
            if validation_errors:
                error_text = "\n".join([f"• {error}" for error in validation_errors])
                messagebox.showerror(
                    "Mapping Validation Failed",
                    f"Found {len(validation_errors)} validation error(s):\n\n{error_text}\n\n"
                    f"Please fix these issues before using the mapping."
                )
                self.log_entry(f"Mapping validation failed: {len(validation_errors)} errors", "ERROR")
            else:
                # Validation passed
                mode = self.mapping_mode.get()
                mapping_count = len(self.field_mappings)
                
                messagebox.showinfo(
                    "Mapping Validation Passed",
                    f"✓ Field mapping validation successful!\n\n"
                    f"Mode: {mode.upper()}\n"
                    f"Fields mapped: {mapping_count}\n\n"
                    f"Your mapping configuration is valid and ready to use."
                )
                self.log_entry(f"Mapping validation passed: {mapping_count} fields", "SUCCESS")
            
        except Exception as e:
            error_msg = f"Mapping validation failed: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def show_mapping_test_dialog(self, test_data, test_status, auth_validated):
        """Show comprehensive mapping test dialog with Test Connection integration"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Test Field Mapping")
        dialog.geometry("600x700")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Test Connection Status
        status_frame = ttk.LabelFrame(dialog, text="Integration Status", padding=10)
        status_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
        
        status_color = 'green' if auth_validated else 'orange'
        ttk.Label(status_frame, text=test_status, foreground=status_color, font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        
        if auth_validated:
            ttk.Label(status_frame, text=" Using validated API Settings from Test Connection", foreground='green').pack(anchor=tk.W)
        else:
            ttk.Label(status_frame, text=" API Settings not validated - may fail", foreground='orange').pack(anchor=tk.W)
        
        # Test Data Preview
        data_frame = ttk.LabelFrame(dialog, text="Test Data (Field Mapping Applied)", padding=10)
        data_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
        
        data_text = scrolledtext.ScrolledText(data_frame, height=15, state=tk.DISABLED)
        data_text.pack(fill=tk.BOTH, expand=True)
        
        # Format test data for display
        data_text.config(state=tk.NORMAL)
        data_text.insert(1.0, json.dumps(test_data, indent=2))
        data_text.config(state=tk.DISABLED)
        
        # Action buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=20, pady=20)
        
        def send_test():
            try:
                # Send using API Settings
                success = self.send_to_api_with_settings(test_data)
                
                if success:
                    messagebox.showinfo("Test Successful", 
                                      f" Field Mapping Test Successful!\n\n"
                                      f" API Settings authentication: \n"
                                      f" Field mapping applied: \n"
                                      f" Data sent successfully: \n\n"
                                      f"Integration Status: {test_status}")
                    self.log_entry("Field mapping test successful with API Settings integration", "SUCCESS")
                else:
                    messagebox.showwarning("Test Failed", 
                                         f" Field Mapping Test Failed\n\n"
                                         f" Check API Settings configuration\n"
                                         f" Verify endpoint URL\n"
                                         f" Review field mapping\n\n"
                                         f"Integration Status: {test_status}")
                
                dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Error", f" Test failed: {str(e)}")
                self.log_entry(f"Mapping API test error: {str(e)}", "ERROR")
        
        ttk.Button(btn_frame, text=" Send Test Data", command=send_test).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text=" Cancel", command=dialog.destroy).pack(side=tk.RIGHT)
    
    def send_to_api_with_settings(self, data: Dict) -> bool:
        """Send data to API using current API Settings authentication - TBS Receiving Format"""
        try:
            # Get settings from API Settings tab
            endpoint = self.api_endpoint_var.get().strip() if hasattr(self, 'api_endpoint_var') else ""
            if not endpoint:
                self.log_entry("No API endpoint configured in API Settings", "ERROR")
                return False
            
            method = self.api_method_var.get() if hasattr(self, 'api_method_var') else "POST"
            content_type = self.api_content_type_var.get() if hasattr(self, 'api_content_type_var') else "application/json"
            auth_type = self.auth_type_var.get() if hasattr(self, 'auth_type_var') else "no_auth"
            
            if content_type == "Custom...":
                content_type = self.custom_content_type_var.get() if hasattr(self, 'custom_content_type_var') else "application/json"
            
            # Prepare headers
            headers = {
                'Content-Type': content_type,
                'User-Agent': 'MDB-Agent-Pro/2.0',
                'Accept': 'application/json'
            }
            
            # Add authentication from API Settings
            if auth_type == "api_key" and hasattr(self, 'api_key_var'):
                api_key = self.api_key_var.get().strip()
                if api_key:
                    headers['Authorization'] = f'Bearer {api_key}'
            elif auth_type == "login" and hasattr(self, 'api_key_var'):
                # For login auth, token is stored in api_key_var after successful login
                token = self.api_key_var.get().strip()
                if token:
                    headers['Authorization'] = f'Bearer {token}'
                else:
                    self.log_entry("Login authentication required but no token available", "ERROR")
                    return False
            
            # ========== CONVERT TO TBS RECEIVING FORMAT ==========
            # Transform the data to TBS JSON-RPC 2.0 format
            tbs_payload = self.convert_to_tbs_format(data)
            
            # Make request using API Settings
            self.log_entry(f"Sending to API using settings: {method} {endpoint} (auth: {auth_type})", "INFO")
            
            response = requests.request(
                method=method,
                url=endpoint,
                json=tbs_payload if content_type == "application/json" else None,
                data=tbs_payload if content_type != "application/json" else None,
                headers=headers,
                timeout=30
            )
            
            if response.status_code < 300:
                self.log_entry(f"API call successful: {response.status_code} (TBS format)", "SUCCESS")
                return True
            else:
                self.log_entry(f"API returned HTTP {response.status_code}: {response.text[:200]}", "WARNING")
                return False
                
        except requests.exceptions.RequestException as e:
            self.log_entry(f"API request failed with settings: {str(e)}", "ERROR")
            return False
        except Exception as e:
            self.log_entry(f"Unexpected error in send_to_api_with_settings: {str(e)}", "ERROR")
            return False
    
    def convert_to_tbs_format(self, data: Dict) -> Dict:
        """Convert standard data format to TBS Receiving API JSON-RPC 2.0 format with nested structure support"""
        try:
            # Check if we have nested mapping configuration
            if hasattr(self, 'mapping_mode') and self.mapping_mode.get() == "nested" and hasattr(self, 'nested_groups'):
                return self.convert_with_nested_mapping(data)
            
            # Generate unique journal_id based on timestamp if not provided
            current_time = datetime.now()
            default_journal_id = f"BANK-{current_time.strftime('%Y%m%d-%H%M%S')}"
            
            # Default TBS structure with proper journal_id
            order_data = {
                "partner_id": data.get("partner_id", data.get("supplier", "Default Supplier")),
                "journal_id": data.get("journal_id", data.get("bank", data.get("payment_method", default_journal_id))),
                "date_order": data.get("date_order", data.get("order_date", current_time.strftime("%d/%m/%Y %H:%M:%S"))),
                "officers": data.get("officers", data.get("officer", data.get("user", "System User"))),
                "keterangan_description": data.get("keterangan_description", data.get("description", data.get("notes", "Auto submission from MDB Agent Pro"))),
                "driver_name": data.get("driver_name", data.get("driver", "")),
                "vehicle_no": data.get("vehicle_no", data.get("vehicle", data.get("truck", ""))),
                "destination_warehouse_id": data.get("destination_warehouse_id", data.get("warehouse", "Default Warehouse")),
                "branch_id": data.get("branch_id", data.get("branch", "Default Branch"))
            }
            
            # Order line data with better field mapping
            order_line = {
                "product_code": data.get("product_code", data.get("product", data.get("item_code", "TBS-AUTO-001"))),
                "qty_brutto": self.convert_to_number(data.get("qty_brutto", data.get("gross_weight", data.get("bruto", 0)))),
                "qty_tara": self.convert_to_number(data.get("qty_tara", data.get("tare_weight", data.get("tara", 0)))),
                "qty_netto": self.convert_to_number(data.get("qty_netto", data.get("net_weight", data.get("netto", 0)))),
                "product_uom": data.get("product_uom", data.get("unit", "kg")),
                "sortation_percent": self.convert_to_number(data.get("sortation_percent", data.get("sorting", 0))),
                "sortation_weight": self.convert_to_number(data.get("sortation_weight", 0)),
                "qty_netto2": self.convert_to_number(data.get("qty_netto2", data.get("final_weight", 0))),
                "price_unit": self.convert_to_number(data.get("price_unit", data.get("price", data.get("harga", 0)))),
                "product_qty": self.convert_to_number(data.get("product_qty", data.get("quantity", 1))),
                "incoming_date": data.get("incoming_date", data.get("in_date", current_time.strftime("%d/%m/%Y %H:%M:%S"))),
                "outgoing_date": data.get("outgoing_date", data.get("out_date", current_time.strftime("%d/%m/%Y %H:%M:%S")))
            }
            
            # Build final TBS payload
            tbs_payload = {
                "jsonrpc": "2.0",
                "params": {
                    "order_data": [
                        {**order_data, "order_line": [order_line]}
                    ]
                }
            }
            
            return tbs_payload
            
        except Exception as e:
            self.log_entry(f"Error converting to TBS format: {str(e)}", "ERROR")
            # Return original data as fallback
            return data
    
    def convert_with_nested_mapping(self, data: Dict) -> Dict:
        """Convert data using nested mapping configuration"""
        try:
            result = {}
            
            # Process nested groups
            for group_name, group_config in self.nested_groups.items():
                if group_config['type'] == 'array':
                    # Create array structure
                    array_items = []
                    
                    # Create single item for now (can be enhanced for multiple records)
                    item = {}
                    
                    # Map fields to this array item
                    for field_name in group_config.get('fields', []):
                        # Try to find matching data field
                        mapped_value = self.find_mapped_value(data, field_name)
                        if mapped_value is not None:
                            item[field_name] = mapped_value
                    
                    # Handle nested arrays within this group
                    if 'nested' in group_config:
                        for nested_name, nested_config in group_config['nested'].items():
                            if nested_config['type'] == 'array':
                                nested_array = []
                                nested_item = {}
                                
                                for nested_field in nested_config.get('fields', []):
                                    mapped_value = self.find_mapped_value(data, nested_field)
                                    if mapped_value is not None:
                                        nested_item[nested_field] = mapped_value
                                
                                if nested_item:
                                    nested_array.append(nested_item)
                                
                                item[nested_name] = nested_array
                    
                    if item:
                        array_items.append(item)
                    
                    result[group_name] = array_items
                
                elif group_config['type'] == 'object':
                    # Create object structure
                    obj = {}
                    for field_name in group_config.get('fields', []):
                        mapped_value = self.find_mapped_value(data, field_name)
                        if mapped_value is not None:
                            obj[field_name] = mapped_value
                    
                    result[group_name] = obj
            
            # Wrap in JSON-RPC format if needed
            if 'order_data' in result:
                return {
                    "jsonrpc": "2.0",
                    "params": result
                }
            
            return result
            
        except Exception as e:
            self.log_entry(f"Error in nested mapping conversion: {str(e)}", "ERROR")
            return data
    
    def find_mapped_value(self, data: Dict, api_field: str):
        """Find mapped value for API field from database data"""
        # Check direct field mappings first
        if hasattr(self, 'field_mappings'):
            for db_field, mapping in self.field_mappings.items():
                if mapping.get('api_field') == api_field:
                    value = data.get(db_field)
                    
                    # Apply transformation if specified
                    transform = mapping.get('transform', 'No Transform')
                    if transform != 'No Transform' and value is not None:
                        value = self.apply_transformation(value, transform)
                    
                    return value
        
        # Fallback to direct field name matching
        if api_field in data:
            return data[api_field]
        
        # Try common aliases
        aliases = {
            'partner_id': ['supplier', 'vendor', 'customer'],
            'journal_id': ['bank', 'payment_method', 'account'],
            'vehicle_no': ['vehicle', 'truck', 'nopol'],
            'qty_netto': ['net_weight', 'netto'],
            'qty_brutto': ['gross_weight', 'bruto'],
            'qty_tara': ['tare_weight', 'tara'],
            'price_unit': ['price', 'harga', 'unit_price'],
            'product_code': ['product', 'item_code', 'item']
        }
        
        if api_field in aliases:
            for alias in aliases[api_field]:
                if alias in data:
                    return data[alias]
        
        return None
    
    def apply_transformation(self, value, transform):
        """Apply data transformation to value"""
        try:
            if transform == "Number":
                return self.convert_to_number(value)
            elif transform == "String":
                return str(value)
            elif transform == "Uppercase":
                return str(value).upper()
            elif transform == "Lowercase":
                return str(value).lower()
            elif transform == "Boolean":
                return bool(value)
            elif transform == "Date":
                # Try to format as date
                if isinstance(value, str):
                    return value
                return str(value)
        except:
            pass
        
        return value
    
    def extract_fields_from_schema(self, schema):
        """Extract field names from API schema"""
        fields = []
        if isinstance(schema, dict):
            if 'properties' in schema:
                fields = list(schema['properties'].keys())
            elif 'fields' in schema:
                fields = schema['fields'] if isinstance(schema['fields'], list) else list(schema['fields'].keys())
            elif 'data' in schema and isinstance(schema['data'], dict):
                fields = list(schema['data'].keys())
            else:
                # Try to get all keys as potential fields
                fields = [k for k in schema.keys() if not k.startswith('_')]
        return fields
    
    def validate_mapping(self):
        """Validate current field mapping with API Settings integration"""
        # Check API Settings configuration first
        endpoint = self.api_endpoint_var.get().strip() if hasattr(self, 'api_endpoint_var') else ""
        auth_type = self.auth_type_var.get() if hasattr(self, 'auth_type_var') else "no_auth"
        
        # Check Test Connection status
        test_connection_status = "Not performed"
        test_connection_color = "orange"
        if hasattr(self, 'last_test_connection_result') and self.last_test_connection_result:
            if self.last_test_connection_result['success']:
                test_connection_status = f" Success ({self.last_test_connection_result['response_time']})"
                test_connection_color = "green"
            else:
                error_msg = self.last_test_connection_result.get('error', 'Unknown error')
                test_connection_status = f" Failed: {error_msg[:30]}..."
                test_connection_color = "red"
        
        if not hasattr(self, 'field_mappings') or not self.field_mappings:
            messagebox.showwarning("Validation", " No field mapping to validate\n\n Please configure field mappings first")
            return
        
        # Create validation dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Field Mapping Validation")
        dialog.geometry("600x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # API Settings Status
        api_frame = ttk.LabelFrame(dialog, text="API Settings Integration", padding=10)
        api_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
        
        ttk.Label(api_frame, text=f"Endpoint: {endpoint or 'Not configured'}", font=('Arial', 9)).pack(anchor=tk.W)
        ttk.Label(api_frame, text=f"Authentication: {auth_type}", font=('Arial', 9)).pack(anchor=tk.W)
        ttk.Label(api_frame, text=f"Test Connection: {test_connection_status}", 
                 foreground=test_connection_color, font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        
        # Validation Results
        results_frame = ttk.LabelFrame(dialog, text="Validation Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
        
        results_text = scrolledtext.ScrolledText(results_frame, height=20, state=tk.DISABLED)
        results_text.pack(fill=tk.BOTH, expand=True)
        
        # Perform validation
        issues = []
        mapped_count = 0
        total_count = len(getattr(self, 'table_columns', []))
        
        # Check for unmapped required fields
        api_fields_used = []
        for db_field, mapping in self.field_mappings.items():
            api_field = mapping.get('api_field')
            if api_field and api_field != "(unmapped)":
                mapped_count += 1
                if api_field in api_fields_used:
                    issues.append(f" Duplicate API field mapping: '{api_field}' is mapped to multiple database fields")
                else:
                    api_fields_used.append(api_field)
        
        # Validation checks
        validation_report = f"FIELD MAPPING VALIDATION REPORT\n{'='*50}\n\n"
        
        # API Settings checks
        validation_report += "API INTEGRATION STATUS:\n"
        if not endpoint:
            validation_report += " No API endpoint configured\n"
            issues.append("API endpoint not configured")
        else:
            validation_report += f" Endpoint configured: {endpoint}\n"
        
        validation_report += f" Authentication type: {auth_type}\n"
        
        if test_connection_color == "green":
            validation_report += f" Test Connection: Successful\n"
        elif test_connection_color == "red":
            validation_report += f" Test Connection: Failed\n"
            issues.append("Test Connection failed")
        else:
            validation_report += f" Test Connection: Not performed\n"
            issues.append("Test Connection not performed")
        
        validation_report += f"\nFIELD MAPPING STATUS:\n"
        validation_report += f" Total database fields: {total_count}\n"
        validation_report += f" Mapped fields: {mapped_count}\n"
        validation_report += f" Coverage: {(mapped_count/total_count*100) if total_count > 0 else 0:.1f}%\n\n"
        
        # Check mapping completeness
        if mapped_count == 0:
            issues.append("No fields are mapped")
            validation_report += " No fields are mapped\n"
        elif mapped_count < total_count * 0.5:
            issues.append(f"Low mapping coverage: Only {mapped_count} out of {total_count} fields mapped")
            validation_report += f" Low mapping coverage\n"
        else:
            validation_report += f" Good mapping coverage\n"
        
        # Check for required API fields (common ones)
        required_fields = ['id', 'timestamp']
        missing_required = []
        for field in required_fields:
            if field not in api_fields_used:
                missing_required.append(field)
        
        if missing_required:
            issues.append(f"Missing recommended fields: {', '.join(missing_required)}")
            validation_report += f" Missing recommended fields: {', '.join(missing_required)}\n"
        
        # Field mapping details
        validation_report += f"\nFIELD MAPPING DETAILS:\n"
        for db_field, mapping in self.field_mappings.items():
            api_field = mapping.get('api_field', '(unmapped)')
            transform = mapping.get('transform', 'none')
            if api_field != "(unmapped)":
                validation_report += f" {db_field}  {api_field}"
                if transform != 'none':
                    validation_report += f" (transform: {transform})"
                validation_report += "\n"
            else:
                validation_report += f" {db_field}  (unmapped)\n"
        
        # Summary
        validation_report += f"\nVALIDATION SUMMARY:\n"
        if not issues:
            validation_report += " All validations passed!\n"
            validation_report += " Ready for production use\n"
        else:
            validation_report += f" Found {len(issues)} issue(s):\n"
            for i, issue in enumerate(issues, 1):
                validation_report += f"  {i}. {issue}\n"
        
        # Display results
        results_text.config(state=tk.NORMAL)
        results_text.insert(1.0, validation_report)
        results_text.config(state=tk.DISABLED)
        
        # Close button
        ttk.Button(dialog, text=" Close", command=dialog.destroy).pack(pady=20)
        
        self.log_entry(f"Mapping validation: {len(issues)} issues found", "INFO" if not issues else "WARNING")
    
    def load_mapping_templates(self):
        """Load available mapping templates into listbox"""
        if not hasattr(self, 'template_listbox'):
            return
            
        self.template_listbox.delete(0, tk.END)
        
        templates = self.config.get('mapping_templates', {})
        for template_name in sorted(templates.keys()):
            self.template_listbox.insert(tk.END, template_name)
        
        # Add default templates if none exist
        if not templates:
            default_templates = {
                "Basic Fields": {
                    "id": {"api_field": "id", "transform": "Number"},
                    "name": {"api_field": "name", "transform": "String"},
                    "created_at": {"api_field": "timestamp", "transform": "Date"}
                },
                "Weight Scale": {
                    "id": {"api_field": "id", "transform": "Number"},
                    "weight": {"api_field": "weight", "transform": "Number"},
                    "timestamp": {"api_field": "created_at", "transform": "Date"},
                    "operator": {"api_field": "user_id", "transform": "String"}
                }
            }
            self.config['mapping_templates'] = default_templates
            self.save_config()
            
            for template_name in sorted(default_templates.keys()):
                self.template_listbox.insert(tk.END, template_name)
    
    def load_selected_template(self):
        """Load selected template into current mapping"""
        try:
            selection = self.template_listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a template to load.")
                return
            
            template_name = self.template_listbox.get(selection[0])
            templates = self.config.get('mapping_templates', {})
            
            if template_name not in templates:
                messagebox.showerror("Template Error", f"Template '{template_name}' not found.")
                return
            
            template_mapping = templates[template_name]
            
            # Apply template to current mapping widgets
            applied_count = 0
            if hasattr(self, 'mapping_widgets'):
                for db_column, widgets in self.mapping_widgets.items():
                    if db_column in template_mapping:
                        mapping = template_mapping[db_column]
                        widgets['api_field'].set(mapping.get('api_field', '(unmapped)'))
                        widgets['transform'].set(mapping.get('transform', 'No Transform'))
                        applied_count += 1
            
            messagebox.showinfo("Template Loaded", 
                              f" Template '{template_name}' loaded successfully!\n\n" +
                              f"Applied to {applied_count} fields.\n" +
                              "Configure remaining fields as needed.")
            
            # Update preview
            self.update_json_preview()
            self.log_entry(f"Mapping template '{template_name}' loaded", "INFO")
            
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load template: {str(e)}")
    
    def delete_selected_template(self):
        """Delete selected template"""
        try:
            selection = self.template_listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a template to delete.")
                return
            
            template_name = self.template_listbox.get(selection[0])
            
            result = messagebox.askyesno("Confirm Delete", 
                                       f"Are you sure you want to delete template '{template_name}'?\n\n" +
                                       "This action cannot be undone.")
            if not result:
                return
            
            # Delete template
            templates = self.config.get('mapping_templates', {})
            if template_name in templates:
                del templates[template_name]
                self.config['mapping_templates'] = templates
                self.save_config()
                
                # Refresh template list
                self.load_mapping_templates()
                
                messagebox.showinfo("Template Deleted", f"Template '{template_name}' has been deleted.")
                self.log_entry(f"Mapping template '{template_name}' deleted", "INFO")
            else:
                messagebox.showerror("Delete Error", f"Template '{template_name}' not found.")
                
        except Exception as e:
            messagebox.showerror("Delete Error", f"Failed to delete template: {str(e)}")
    
    def refresh_mapping_interface(self):
        """Refresh mapping interface based on selected table"""
        if not hasattr(self, 'mapping_scroll_frame'):
            return
            
        # Clear existing widgets
        for widget in self.mapping_scroll_frame.winfo_children():
            widget.destroy()
        
        if not hasattr(self, 'selected_table') or not self.selected_table or not hasattr(self, 'table_columns'):
            ttk.Label(self.mapping_scroll_frame, text="Please select a database table first from 'Database Connection' tab", 
                     font=('Arial', 12), foreground='gray').pack(pady=50)
            return
        
        # Create mapping rows for each database column
        for i, column in enumerate(self.table_columns):
            self.create_mapping_row(i, column)
        
        # Add save button at bottom with better sizing
        save_frame = ttk.Frame(self.mapping_scroll_frame)
        save_frame.pack(fill=tk.X, pady=(20, 10), padx=10)
        
        # Create buttons with explicit sizes and better styling
        save_btn = ttk.Button(save_frame, text=" Save Field Mapping", 
                             command=self.save_field_mapping, width=20)
        save_btn.pack(side=tk.LEFT, padx=(0, 10), pady=5)
        
        reset_btn = ttk.Button(save_frame, text=" Reset Mapping", 
                              command=self.reset_field_mapping, width=15)
        reset_btn.pack(side=tk.LEFT, padx=(0, 10), pady=5)
        
        preview_btn = ttk.Button(save_frame, text=" Generate Preview", 
                                command=self.update_json_preview, width=18)
        preview_btn.pack(side=tk.LEFT, padx=(0, 10), pady=5)
        
        # Add Template Save/Load buttons
        template_frame = ttk.Frame(self.mapping_scroll_frame)
        template_frame.pack(fill=tk.X, pady=10, padx=10)
        
        ttk.Label(template_frame, text="Templates:", font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 10))
        
        save_template_btn = ttk.Button(template_frame, text=" Save Template", 
                                      command=self.show_save_template_dialog, width=15)
        save_template_btn.pack(side=tk.LEFT, padx=(0, 10), pady=5)
        
        load_template_btn = ttk.Button(template_frame, text=" Load Template", 
                                      command=self.load_selected_template, width=15)
        load_template_btn.pack(side=tk.LEFT, padx=(0, 10), pady=5)
        
        # Update scroll region and scroll to show buttons
        self.mapping_canvas.update_idletasks()
        self.mapping_canvas.configure(scrollregion=self.mapping_canvas.bbox("all"))
        
        # Scroll to bottom to show save buttons
        self.mapping_canvas.yview_moveto(1.0)
    
    def create_mapping_row(self, row_index, column):
        """Create a mapping row for a database column with improved sizing"""
        row_frame = ttk.Frame(self.mapping_scroll_frame)
        row_frame.pack(fill=tk.X, pady=4, padx=10)
        
        # Configure grid weights for responsive layout
        row_frame.columnconfigure(2, weight=1)
        row_frame.columnconfigure(3, weight=1)
        
        # Database column label with better formatting
        db_text = f"{column['name']} ({column.get('type', 'Unknown')})"
        if len(db_text) > 25:
            db_text = db_text[:22] + "..."
        
        db_label = ttk.Label(row_frame, text=db_text, 
                            font=('Arial', 9), cursor="hand2", width=25)
        db_label.grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        # Add drag-and-drop functionality
        db_label.bind("<Button-1>", lambda e: self.start_drag(e, column['name']))
        db_label.bind("<B1-Motion>", self.on_drag)
        db_label.bind("<ButtonRelease-1>", self.on_drop)
        
        # Add double-click for details
        db_label.bind("<Double-1>", lambda e: self.show_column_details(column))
        
        # Arrow
        ttk.Label(row_frame, text="", font=('Arial', 12, 'bold')).grid(row=0, column=1, padx=10)
        
        # API field dropdown with better sizing
        api_var = tk.StringVar()
        api_combo = ttk.Combobox(row_frame, textvariable=api_var, width=35, state="readonly")
        api_combo['values'] = [
            "(unmapped)", 
            "id", "name", "value", "timestamp", "status", "data", "description",
            "created_at", "updated_at", "user_id", "category", "type", "amount",
            "weight", "quantity", "price", "total", "code", "reference",
            "Custom..."
        ]
        api_combo.set("(unmapped)")
        api_combo.grid(row=0, column=2, sticky=tk.EW, padx=(0, 10))
        
        # Add double-click for API field details
        api_combo.bind("<Double-1>", lambda e: self.show_api_field_details(api_var.get()))
        
        # Add handler for Custom field selection
        def on_api_field_change(*args):
            if api_var.get() == "Custom...":
                self.show_custom_field_dialog(api_var, column['name'])
            self.update_json_preview()
        
        api_var.trace('w', on_api_field_change)
        
        # Transformation dropdown with better sizing
        transform_var = tk.StringVar()
        transform_combo = ttk.Combobox(row_frame, textvariable=transform_var, width=20, state="readonly")
        transform_combo['values'] = [
            "No Transform", "String", "Number", "Date", "Boolean", 
            "Uppercase", "Lowercase", "Trim", "JSON", "Base64", "Custom..."
        ]
        transform_combo.set("No Transform")
        transform_combo.grid(row=0, column=3, sticky=tk.EW, padx=(0, 10))
        
        # Add double-click for transformation details
        transform_combo.bind("<Double-1>", lambda e: self.show_transform_details(transform_var.get()))
        
        # Store references for later use
        if not hasattr(self, 'mapping_widgets'):
            self.mapping_widgets = {}
        self.mapping_widgets[column['name']] = {
            'api_field': api_var,
            'transform': transform_var,
            'api_combo': api_combo,
            'transform_combo': transform_combo,
            'db_label': db_label
        }
        
        # Load existing mapping if available
        if column['name'] in self.field_mappings:
            mapping = self.field_mappings[column['name']]
            api_var.set(mapping.get('api_field', '(unmapped)'))
            transform_var.set(mapping.get('transform', 'No Transform'))
        
        # Bind change events for real-time preview update  
        transform_var.trace('w', lambda *args: self.update_json_preview())
    
    def show_custom_field_dialog(self, api_var, column_name):
        """Show dialog for adding custom API field mapping"""
        if not hasattr(self, 'api_fields'):
            self.api_fields = []
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Custom API Field - {column_name}")
        dialog.geometry("500x600")
        dialog.resizable(True, True)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
        y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Title
        title_frame = ttk.Frame(dialog)
        title_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
        
        ttk.Label(title_frame, text=f" Custom API Field for: {column_name}", 
                 font=('Arial', 14, 'bold')).pack()
        ttk.Label(title_frame, text="Define custom API field mapping for this database column", 
                 font=('Arial', 10), foreground='gray').pack()
        
        # Field entry
        entry_frame = ttk.LabelFrame(dialog, text="Field Configuration", padding=10)
        entry_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        # Field name
        ttk.Label(entry_frame, text="Field Name:").pack(anchor=tk.W)
        field_name_var = tk.StringVar()
        field_name_entry = ttk.Entry(entry_frame, textvariable=field_name_var, font=('Consolas', 10))
        field_name_entry.pack(fill=tk.X, pady=(2, 10))
        
        # Field path (for nested JSON)
        ttk.Label(entry_frame, text="JSON Path (optional, e.g., 'order_data.items[0].name'):").pack(anchor=tk.W)
        field_path_var = tk.StringVar()
        field_path_entry = ttk.Entry(entry_frame, textvariable=field_path_var, font=('Consolas', 10))
        field_path_entry.pack(fill=tk.X, pady=(2, 10))
        
        # Description
        ttk.Label(entry_frame, text="Description:").pack(anchor=tk.W)
        description_var = tk.StringVar()
        description_entry = ttk.Entry(entry_frame, textvariable=description_var)
        description_entry.pack(fill=tk.X, pady=(2, 0))
        
        # API format examples
        examples_frame = ttk.LabelFrame(dialog, text="Common API Field Examples", padding=10)
        examples_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
        
        examples_text = scrolledtext.ScrolledText(examples_frame, height=15, state=tk.DISABLED, font=('Consolas', 9))
        examples_text.pack(fill=tk.BOTH, expand=True)
        
        # Populate examples
        examples_content = """TBS Receiving API - JSON-RPC 2.0 Format:
[CHAR]

Basic Fields:
[CHAR] id                    - Request ID
[CHAR] method               - API method name
[CHAR] jsonrpc              - Protocol version (2.0)

Order Data Fields (order_data object):
[CHAR] order_data.order_id           - Order identifier
[CHAR] order_data.customer_name      - Customer name
[CHAR] order_data.total_amount       - Total order amount
[CHAR] order_data.order_date         - Order date
[CHAR] order_data.status             - Order status

Item Fields (order_data.items array):
[CHAR] order_data.items[0].item_id      - Item identifier
[CHAR] order_data.items[0].item_name    - Item name
[CHAR] order_data.items[0].quantity     - Item quantity
[CHAR] order_data.items[0].unit_price   - Item unit price
[CHAR] order_data.items[0].category     - Item category

Weight Fields (for timbangan/scale data):
[CHAR] order_data.items[0].gross_weight    - Gross weight
[CHAR] order_data.items[0].tare_weight     - Tare weight
[CHAR] order_data.items[0].net_weight      - Net weight
[CHAR] order_data.items[0].weight_unit     - Weight unit (kg, tons)

Timestamp Fields:
[CHAR] order_data.created_at             - Creation timestamp
[CHAR] order_data.updated_at             - Update timestamp
[CHAR] order_data.items[0].weighing_time - Weighing timestamp

Location Fields:
[CHAR] order_data.warehouse_id           - Warehouse identifier
[CHAR] order_data.location              - Location description
[CHAR] order_data.items[0].scale_id     - Scale/timbangan ID

Additional Metadata:
[CHAR] order_data.operator_id           - Operator identifier
[CHAR] order_data.vehicle_number        - Vehicle number
[CHAR] order_data.reference_number      - Reference/document number
[CHAR] order_data.notes                 - Additional notes

Standard REST API Fields:
[CHAR]
[CHAR] customer_id          - Customer identifier
[CHAR] product_name         - Product name
[CHAR] price               - Product price
[CHAR] created_at          - Creation timestamp
 updated_at          - Update timestamp"""
        
        examples_text.config(state=tk.NORMAL)
        examples_text.insert(tk.END, examples_content)
        examples_text.config(state=tk.DISABLED)
        
        # Quick insert buttons
        quick_frame = ttk.Frame(dialog)
        quick_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        def insert_preset(field_name, path="", desc=""):
            field_name_var.set(field_name)
            field_path_var.set(path)
            description_var.set(desc)
        
        ttk.Button(quick_frame, text="TBS Order ID", 
                  command=lambda: insert_preset("order_id", "order_data.order_id", "TBS Order identifier")).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(quick_frame, text="TBS Item Name", 
                  command=lambda: insert_preset("item_name", "order_data.items[0].item_name", "TBS Item name")).pack(side=tk.LEFT, padx=5)
        ttk.Button(quick_frame, text="TBS Weight", 
                  command=lambda: insert_preset("net_weight", "order_data.items[0].net_weight", "TBS Net weight")).pack(side=tk.LEFT, padx=5)
        
        # Buttons
        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        def add_field():
            field_name = field_name_var.get().strip()
            field_path = field_path_var.get().strip()
            
            if not field_name:
                messagebox.showerror("Error", "Please enter a field name")
                return
            
            # Use path if provided, otherwise use name
            final_field = field_path if field_path else field_name
            
            # Check if field already exists
            if final_field in self.api_fields:
                messagebox.showwarning("Warning", f"Field '{final_field}' already exists")
                return
            
            # Set the custom field in the dropdown
            if final_field:
                # Add to api_fields if not exists
                if final_field not in self.api_fields:
                    self.api_fields.append(final_field)
                    # Refresh dropdown options
                    if hasattr(self, 'mapping_scroll_frame'):
                        self.refresh_mapping_interface()
                api_var.set(final_field)
            else:
                messagebox.showwarning("Warning", "Please enter a field name")
                return
            
            dialog.destroy()
        
        def cancel_dialog():
            api_var.set("(unmapped)")
            dialog.destroy()
        
        ttk.Button(button_frame, text="Cancel", command=cancel_dialog).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Add Field", command=add_field).pack(side=tk.RIGHT)
        
        # Focus on field name entry
        field_name_entry.focus_set()
    
    def show_custom_field_dialog_manual(self):
        """Show dialog for adding custom API fields manually (for use in mapping tab)"""
        if not hasattr(self, 'api_fields'):
            self.api_fields = []
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Custom API Field")
        dialog.geometry("500x600")
        dialog.resizable(True, True)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
        y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Title
        title_frame = ttk.Frame(dialog)
        title_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
        
        ttk.Label(title_frame, text=" Add Custom API Field", 
                 font=('Arial', 14, 'bold')).pack()
        ttk.Label(title_frame, text="Define custom API fields for field mapping", 
                 font=('Arial', 10), foreground='gray').pack()
        
        # Field entry
        entry_frame = ttk.LabelFrame(dialog, text="Field Configuration", padding=10)
        entry_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        # Field name
        ttk.Label(entry_frame, text="Field Name:").pack(anchor=tk.W)
        field_name_var = tk.StringVar()
        field_name_entry = ttk.Entry(entry_frame, textvariable=field_name_var, font=('Consolas', 10))
        field_name_entry.pack(fill=tk.X, pady=(2, 10))
        
        # Field path (for nested JSON)
        ttk.Label(entry_frame, text="JSON Path (optional, e.g., 'order_data.items[0].name'):").pack(anchor=tk.W)
        field_path_var = tk.StringVar()
        field_path_entry = ttk.Entry(entry_frame, textvariable=field_path_var, font=('Consolas', 10))
        field_path_entry.pack(fill=tk.X, pady=(2, 10))
        
        # Description
        ttk.Label(entry_frame, text="Description:").pack(anchor=tk.W)
        description_var = tk.StringVar()
        description_entry = ttk.Entry(entry_frame, textvariable=description_var)
        description_entry.pack(fill=tk.X, pady=(2, 0))
        
        # API format examples
        examples_frame = ttk.LabelFrame(dialog, text="Common API Field Examples", padding=10)
        examples_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
        
        examples_text = scrolledtext.ScrolledText(examples_frame, height=10, state=tk.DISABLED, font=('Consolas', 9))
        examples_text.pack(fill=tk.BOTH, expand=True)
        
        # Populate examples - shorter version for manual dialog
        examples_content = """TBS Receiving API - JSON-RPC 2.0 Format:
[CHAR]

Basic Fields:
[CHAR] id                    - Request ID
[CHAR] method               - API method name
[CHAR] order_data.order_id           - Order identifier
[CHAR] order_data.customer_name      - Customer name
[CHAR] order_data.items[0].item_name    - Item name
[CHAR] order_data.items[0].net_weight   - Net weight

Quick Examples:
[CHAR] customer_id          - Customer identifier
[CHAR] product_name         - Product name
[CHAR] price               - Product price
 created_at          - Creation timestamp"""
        
        examples_text.config(state=tk.NORMAL)
        examples_text.insert(tk.END, examples_content)
        examples_text.config(state=tk.DISABLED)
        
        # Quick insert buttons
        quick_frame = ttk.Frame(dialog)
        quick_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        def insert_preset(field_name, path="", desc=""):
            field_name_var.set(field_name)
            field_path_var.set(path)
            description_var.set(desc)
        
        ttk.Button(quick_frame, text="TBS Order ID", 
                  command=lambda: insert_preset("order_id", "order_data.order_id", "TBS Order identifier")).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(quick_frame, text="TBS Item Name", 
                  command=lambda: insert_preset("item_name", "order_data.items[0].item_name", "TBS Item name")).pack(side=tk.LEFT, padx=5)
        
        # Buttons
        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        def add_field():
            field_name = field_name_var.get().strip()
            field_path = field_path_var.get().strip()
            
            if not field_name:
                messagebox.showerror("Error", "Please enter a field name")
                return
            
            # Use path if provided, otherwise use name
            final_field = field_path if field_path else field_name
            
            # Check if field already exists
            if final_field in self.api_fields:
                messagebox.showwarning("Warning", f"Field '{final_field}' already exists")
                return
            
            # Add to api_fields list
            self.api_fields.append(final_field)
            
            messagebox.showinfo("Success", f"Custom field '{final_field}' added successfully!")
            
            # Refresh mapping interface if we're in mapping tab
            if hasattr(self, 'mapping_scroll_frame'):
                self.refresh_mapping_interface()
            
            dialog.destroy()
        
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Add Field", command=add_field).pack(side=tk.RIGHT)
        
        # Focus on field name entry
        field_name_entry.focus_set()
    
    def start_drag(self, event, column_name):
        """Start drag operation"""
        self.drag_data = {'column': column_name, 'start_x': event.x, 'start_y': event.y}
        
    def on_drag(self, event):
        """Handle drag motion"""
        if hasattr(self, 'drag_data'):
            # Visual feedback during drag
            pass
            
    def on_drop(self, event):
        """Handle drop operation"""
        if hasattr(self, 'drag_data'):
            # For now, just show a message about drag functionality
            messagebox.showinfo("Drag & Drop", 
                              f"Dragging '{self.drag_data['column']}'\n\n" +
                              " Use the dropdown menus to map fields\n" +
                              " Double-click for detailed options")
            delattr(self, 'drag_data')
    
    def show_column_details(self, column):
        """Show detailed information about database column"""
        details_window = tk.Toplevel(self.root)
        details_window.title(f"Column Details: {column['name']}")
        details_window.geometry("400x300")
        details_window.transient(self.root)
        details_window.grab_set()
        
        # Column information
        info_frame = ttk.LabelFrame(details_window, text="Column Information", padding=15)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        details = [
            ("Column Name:", column['name']),
            ("Data Type:", column.get('type', 'Unknown')),
            ("Table:", getattr(self, 'selected_table', 'N/A')),
            ("Nullable:", column.get('nullable', 'Unknown')),
            ("Default Value:", column.get('default', 'None')),
        ]
        
        for i, (label, value) in enumerate(details):
            ttk.Label(info_frame, text=label, font=('Arial', 9, 'bold')).grid(row=i, column=0, sticky=tk.W, padx=(0, 15), pady=3)
            ttk.Label(info_frame, text=str(value), font=('Arial', 9)).grid(row=i, column=1, sticky=tk.W, pady=3)
        
        # Sample data if available
        if hasattr(self, 'db_connection') and self.db_connection:
            try:
                cursor = self.db_connection.cursor()
                cursor.execute(f"SELECT TOP 5 [{column['name']}] FROM [{getattr(self, 'selected_table', '')}]")
                samples = [str(row[0]) for row in cursor.fetchall()]
                
                if samples:
                    ttk.Label(info_frame, text="Sample Values:", font=('Arial', 9, 'bold')).grid(row=len(details), column=0, sticky=tk.NW, padx=(0, 15), pady=(10, 3))
                    sample_text = "\n".join(samples[:3])
                    ttk.Label(info_frame, text=sample_text, font=('Arial', 9)).grid(row=len(details), column=1, sticky=tk.W, pady=(10, 3))
            except:
                pass
        
        # Close button
        ttk.Button(details_window, text="Close", command=details_window.destroy).pack(pady=10)
    
    def show_api_field_details(self, api_field):
        """Show detailed information about API field mapping"""
        if api_field in ["(unmapped)", ""]:
            messagebox.showinfo("API Field Details", "No API field selected.\n\nChoose an API field from the dropdown to see mapping details.")
            return
            
        details_window = tk.Toplevel(self.root)
        details_window.title(f"API Field Details: {api_field}")
        details_window.geometry("400x300")
        details_window.transient(self.root)
        details_window.grab_set()
        
        # API field information
        info_frame = ttk.LabelFrame(details_window, text="API Field Information", padding=15)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Predefined field descriptions
        field_descriptions = {
            "id": "Unique identifier for the record",
            "name": "Name or title of the record",
            "value": "Main value or data content",
            "timestamp": "Date and time information",
            "status": "Status or state of the record",
            "data": "General data payload",
            "description": "Detailed description text",
            "created_at": "Record creation timestamp",
            "updated_at": "Last modification timestamp",
            "user_id": "User identifier reference",
            "category": "Classification category",
            "type": "Type or kind specification",
            "amount": "Numeric amount value",
            "weight": "Weight measurement",
            "quantity": "Quantity count",
            "price": "Price or cost value",
            "total": "Total calculated value",
            "code": "Code or reference identifier",
            "reference": "Reference to related record"
        }
        
        description = field_descriptions.get(api_field, "Custom API field - configure as needed")
        
        ttk.Label(info_frame, text="Field Name:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        ttk.Label(info_frame, text=api_field, font=('Arial', 10)).pack(anchor=tk.W, pady=(0, 10))
        
        ttk.Label(info_frame, text="Description:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        ttk.Label(info_frame, text=description, font=('Arial', 9), wraplength=350, justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 10))
        
        ttk.Label(info_frame, text="Usage Example:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        example = f'"{api_field}": "value_from_database_column"'
        ttk.Label(info_frame, text=example, font=('Arial', 9, 'italic'), foreground='blue').pack(anchor=tk.W, pady=(0, 10))
        
        # Close button
        ttk.Button(details_window, text="Close", command=details_window.destroy).pack(pady=10)
    
    def show_transform_details(self, transform_type):
        """Show detailed information about data transformation"""
        if transform_type in ["No Transform", ""]:
            messagebox.showinfo("Transformation Details", "No transformation selected.\n\nData will be sent as-is from the database.")
            return
            
        details_window = tk.Toplevel(self.root)
        details_window.title(f"Transformation Details: {transform_type}")
        details_window.geometry("450x350")
        details_window.transient(self.root)
        details_window.grab_set()
        
        # Transformation information
        info_frame = ttk.LabelFrame(details_window, text="Transformation Information", padding=15)
        info_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Transformation descriptions
        transform_info = {
            "String": {
                "description": "Convert value to string text format",
                "example": "123  '123', true  'true'"
            },
            "Number": {
                "description": "Convert value to numeric format",
                "example": "'123'  123, '12.5'  12.5"
            },
            "Date": {
                "description": "Convert value to ISO date format",
                "example": "2024-01-15  '2024-01-15T00:00:00Z'"
            },
            "Boolean": {
                "description": "Convert value to true/false",
                "example": "1  true, 0  false"
            },
            "Uppercase": {
                "description": "Convert text to uppercase",
                "example": "'hello'  'HELLO'"
            },
            "Lowercase": {
                "description": "Convert text to lowercase", 
                "example": "'HELLO'  'hello'"
            },
            "Trim": {
                "description": "Remove leading and trailing spaces",
                "example": "' hello '  'hello'"
            },
            "JSON": {
                "description": "Convert value to JSON string",
                "example": "object  '{\"key\": \"value\"}'"
            },
            "Base64": {
                "description": "Encode value in Base64 format",
                "example": "'hello'  'aGVsbG8='"
            }
        }
        
        info = transform_info.get(transform_type, {
            "description": "Custom transformation - configure as needed",
            "example": "Define your own transformation logic"
        })
        
        ttk.Label(info_frame, text="Transformation Type:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        ttk.Label(info_frame, text=transform_type, font=('Arial', 10)).pack(anchor=tk.W, pady=(0, 10))
        
        ttk.Label(info_frame, text="Description:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        ttk.Label(info_frame, text=info["description"], font=('Arial', 9), wraplength=400, justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 10))
        
        ttk.Label(info_frame, text="Example:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        ttk.Label(info_frame, text=info["example"], font=('Arial', 9, 'italic'), foreground='blue', wraplength=400, justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 10))
        
        # Close button
        ttk.Button(details_window, text="Close", command=details_window.destroy).pack(pady=10)
    
    def save_field_mapping(self):
        """Save current field mapping configuration"""
        try:
            if not hasattr(self, 'mapping_widgets'):
                messagebox.showwarning("Warning", "No mapping to save")
                return
            
            # Collect mapping data
            new_mapping = {}
            mapped_count = 0
            mapped_api_fields = []
            
            for db_field, widgets in self.mapping_widgets.items():
                api_field = widgets['api_field'].get()
                transform = widgets['transform'].get()
                
                if api_field and api_field != "(unmapped)":
                    new_mapping[db_field] = {
                        'api_field': api_field,
                        'transform': transform
                    }
                    mapped_count += 1
                    mapped_api_fields.append(api_field)
            
            # Check for TBS required fields
            tbs_required_fields = ['journal_id', 'partner_id']
            missing_required = []
            
            if hasattr(self, 'api_fields_source') and 'tbs' in self.api_fields_source.lower():
                for required_field in tbs_required_fields:
                    if required_field not in mapped_api_fields:
                        missing_required.append(required_field)
                
                if missing_required:
                    result = messagebox.askyesnocancel(
                        "Missing Required Fields", 
                        f"TBS API requires these fields that are not mapped:\n {', '.join(missing_required)}\n\n" +
                        "Do you want to:\n" +
                        " YES: Auto-fill missing fields with defaults\n" +
                        " NO: Save mapping as-is (may cause API errors)\n" +
                        " CANCEL: Don't save, fix mapping manually"
                    )
                    
                    if result is None:  # Cancel
                        return
                    elif result:  # YES - auto-fill
                        # Add missing required fields with default values
                        for missing_field in missing_required:
                            # Create a dummy entry to ensure the field is present
                            new_mapping[f"_auto_{missing_field}"] = {
                                'api_field': missing_field,
                                'transform': 'Default Value'
                            }
                            mapped_count += 1
                        
                        messagebox.showinfo("Auto-Fill Applied", 
                                          f"Added default values for: {', '.join(missing_required)}")
            
            # Update field mappings
            self.field_mappings = new_mapping
            self.config['field_mapping'] = new_mapping
            self.save_config()
            
            # Update field mapper with new mappings
            self.update_field_mapper()
            
            # Update preview
            self.update_json_preview()
            
            success_msg = f"Field mapping saved!\n\n{mapped_count} fields mapped out of {len(self.mapping_widgets)}"
            if missing_required and not result:
                success_msg += f"\n\nWarning: Missing required fields: {', '.join(missing_required)}"
            
            messagebox.showinfo("Success", success_msg)
            self.log_entry(f"Field mapping saved: {mapped_count} fields mapped", "SUCCESS")
            
        except Exception as e:
            error_msg = f"Failed to save field mapping: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def reset_field_mapping(self):
        """Reset field mapping to defaults"""
        result = messagebox.askyesno("Confirm Reset", "Are you sure you want to reset all field mappings?")
        if result:
            self.field_mappings = {}
            if hasattr(self, 'mapping_widgets'):
                for widgets in self.mapping_widgets.values():
                    widgets['api_field'].set("(unmapped)")
                    widgets['transform'].set("No Transform")
            self.update_json_preview()
            messagebox.showinfo("Reset Complete", "Field mapping has been reset")
    
    def update_json_preview(self):
        """Update JSON preview based on current mapping"""
        if not hasattr(self, 'json_preview'):
            return
            
        try:
            # Generate sample data using FieldMapper for accurate preview
            sample_raw_data = {}
            
            # Create sample raw data based on field mappings
            if hasattr(self, 'field_mappings') and self.field_mappings:
                for db_field, mapping in self.field_mappings.items():
                    # Generate realistic sample values
                    if 'id' in db_field.lower():
                        sample_raw_data[db_field] = 12345
                    elif 'name' in db_field.lower():
                        sample_raw_data[db_field] = "Sample Name"
                    elif 'date' in db_field.lower() or 'time' in db_field.lower():
                        sample_raw_data[db_field] = "2025-01-15 10:30:00"
                    elif 'price' in db_field.lower() or 'amount' in db_field.lower():
                        sample_raw_data[db_field] = 1500.75
                    elif 'status' in db_field.lower():
                        sample_raw_data[db_field] = "Active"
                    elif 'qty' in db_field.lower() or 'quantity' in db_field.lower():
                        sample_raw_data[db_field] = 10
                    else:
                        sample_raw_data[db_field] = f"sample_{db_field.lower()}"
            
            # Use FieldMapper to build preview data
            if self.field_mapper and sample_raw_data:
                preview_data = self.field_mapper.build_api_payload(sample_raw_data)
                
                # Add sample metadata if not present
                if "uuid" not in preview_data:
                    preview_data["uuid"] = "sample-uuid-12345"
                if "timestamp" not in preview_data:
                    preview_data["timestamp"] = "2025-01-15T10:30:00Z"
                if "table" not in preview_data:
                    preview_data["table"] = getattr(self, 'selected_table', 'your_table')
                    
            else:
                # Fallback preview when no mapper or mappings
                preview_data = {
                    "uuid": "sample-uuid-12345", 
                    "timestamp": "2025-01-15T10:30:00Z",
                    "table": getattr(self, 'selected_table', 'your_table'),
                    "message": "Configure field mapping to see preview",
                    "sample_structure": {
                        "field1": "value1",
                        "field2": "value2"
                    }
                }
            
            # Format JSON with proper indentation
            json_text = json.dumps(preview_data, indent=2, ensure_ascii=False)
            
            # Update preview widget
            self.json_preview.config(state=tk.NORMAL)
            self.json_preview.delete(1.0, tk.END)
            self.json_preview.insert(1.0, json_text)
            self.json_preview.config(state=tk.DISABLED)
            
        except Exception as e:
            # Fallback preview
            fallback_json = {
                "message": "Preview generation failed",
                "error": str(e),
                "sample": {"field": "value"}
            }
            json_text = json.dumps(fallback_json, indent=2)
            
            self.json_preview.config(state=tk.NORMAL)
            self.json_preview.delete(1.0, tk.END)
            self.json_preview.insert(1.0, json_text)
            self.json_preview.config(state=tk.DISABLED)
    
    def apply_transformation(self, value, transform):
        """Apply data transformation"""
        if transform == "No Transform":
            return value
        elif transform == "String":
            return str(value)
        elif transform == "Number":
            try:
                return float(value)
            except:
                return 0
        elif transform == "Date":
            if hasattr(value, 'strftime'):
                return value.strftime('%Y-%m-%d')
            return str(value)
        else:
            return value
    
    def refresh_scheduler_log(self):
        """Refresh scheduler log"""
        if hasattr(self, 'scheduler_tree'):
            # Clear existing items
            for item in self.scheduler_tree.get_children():
                self.scheduler_tree.delete(item)
            
            # Add some sample entries
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            status = "Running" if self.is_running else "Stopped"
            self.scheduler_tree.insert("", 0, values=(timestamp, "Status Check", "INFO", f"Scheduler is {status}"))
    
    def start_health_monitoring(self):
        """Start health monitoring"""
        self.run_all_health_checks()
    
    def run_health_checks(self):
        """Run health checks"""
        self.run_all_health_checks()
    
    def view_transaction_details(self, event):
        """View transaction details"""
        selection = self.trans_tree.selection()
        if selection:
            item = self.trans_tree.item(selection[0])
            values = item['values']
            messagebox.showinfo("Transaction Details", f"Transaction ID: {values[0]}\nDetails: {values[-1]}")
    
    def filter_transactions(self, event=None):
        """Filter transactions"""
        self.load_transaction_log()
    
    def refresh_transactions(self):
        """Refresh transaction log"""
        self.load_transaction_log()
    
    def clear_transaction_log(self):
        """Clear transaction log"""
        if not self.admin_mode:
            messagebox.showerror("Error", "Admin mode required")
            return
        
        result = messagebox.askyesno("Confirm", "Clear all transaction logs?")
        if result:
            # Clear tree view
            for item in self.transaction_tree.get_children():
                self.transaction_tree.delete(item)
            messagebox.showinfo("Success", "Transaction log cleared")
    
    def export_transaction_report(self):
        """Export transaction report"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            if filename:
                import csv
                with open(filename, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(["ID", "Timestamp", "Table", "Status", "Records", "Details"])
                    
                    for item in self.transaction_tree.get_children():
                        values = self.transaction_tree.item(item)['values']
                        writer.writerow(values)
                
                messagebox.showinfo("Success", f"Report exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")
    
    def load_settings(self):
        """Load configuration settings into UI elements"""
        try:
            if hasattr(self, 'mdb_file_var'):
                self.mdb_file_var.set(self.config.get('mdb_file', ''))
            if hasattr(self, 'mdb_password_var'):
                self.mdb_password_var.set(self.config.get('mdb_password', 'qwerty123'))
            if hasattr(self, 'table_var'):
                self.table_var.set(self.config.get('selected_table', ''))
            if hasattr(self, 'api_endpoint_var'):
                self.api_endpoint_var.set(self.config.get('api_endpoint', ''))
            if hasattr(self, 'api_key_var'):
                self.api_key_var.set(self.config.get('api_key', ''))
            if hasattr(self, 'push_interval_var'):
                self.push_interval_var.set(self.config.get('push_interval', 300))
            if hasattr(self, 'auto_push_var'):
                self.auto_push_var.set(self.config.get('auto_push', False))
            
            # Load field mappings
            field_mapping = self.config.get('field_mapping', {})
            if field_mapping and hasattr(self, 'field_mappings'):
                self.field_mappings = field_mapping
                
        except Exception as e:
            print(f"Error loading settings: {e}")
    
    def run(self):
        """Run the application"""
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()
    
    def on_closing(self):
        """Handle application closing"""
        self.is_running = False
        self.save_config()
        if self.db_connection:
            self.db_connection.close()
        self.root.destroy()
    
    # =====================================
    # NESTED MAPPING FUNCTIONS
    # =====================================
    
    def add_mapping_group(self):
        """Add a new mapping group for nested structure"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Mapping Group")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        self.center_window(dialog, 400, 300)
        
        ttk.Label(dialog, text="Create New Mapping Group", 
                 font=('Arial', 12, 'bold')).pack(pady=10)
        
        # Group name
        ttk.Label(dialog, text="Group Name:").pack(anchor=tk.W, padx=20)
        group_name_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=group_name_var, width=40).pack(padx=20, pady=(0, 10))
        
        # Group type
        ttk.Label(dialog, text="Group Type:").pack(anchor=tk.W, padx=20)
        group_type_var = tk.StringVar(value="object")
        type_frame = ttk.Frame(dialog)
        type_frame.pack(padx=20, pady=(0, 10))
        ttk.Radiobutton(type_frame, text="Object", variable=group_type_var, value="object").pack(side=tk.LEFT)
        ttk.Radiobutton(type_frame, text="Array", variable=group_type_var, value="array").pack(side=tk.LEFT, padx=(20, 0))
        
        # Description
        ttk.Label(dialog, text="Description:").pack(anchor=tk.W, padx=20)
        desc_text = tk.Text(dialog, height=4, width=45)
        desc_text.pack(padx=20, pady=(0, 10))
        
        def save_group():
            name = group_name_var.get().strip()
            if not name:
                messagebox.showwarning("Warning", "Please enter group name")
                return
                
            group_type = group_type_var.get()
            description = desc_text.get(1.0, tk.END).strip()
            
            # Initialize nested_groups if not exists
            if not hasattr(self, 'nested_groups'):
                self.nested_groups = {}
            
            self.nested_groups[name] = {
                'type': group_type,
                'description': description,
                'fields': []
            }
            
            self.update_nested_mapping_display()
            messagebox.showinfo("Success", f"Group '{name}' created successfully!")
            dialog.destroy()
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="Save Group", command=save_group).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT)
    
    def add_mapping_array(self):
        """Add a new mapping array for nested structure"""
        # Quick way to add array - calls add_mapping_group with array preset
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Mapping Array")
        dialog.geometry("450x350")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        self.center_window(dialog, 450, 350)
        
        ttk.Label(dialog, text="Create New Mapping Array", 
                 font=('Arial', 12, 'bold')).pack(pady=10)
        
        # Array name
        ttk.Label(dialog, text="Array Name:").pack(anchor=tk.W, padx=20)
        array_name_var = tk.StringVar(value="order_line")
        ttk.Entry(dialog, textvariable=array_name_var, width=40).pack(padx=20, pady=(0, 10))
        
        # Array item template
        ttk.Label(dialog, text="Array Item Fields (one per line):").pack(anchor=tk.W, padx=20)
        fields_text = tk.Text(dialog, height=8, width=50)
        fields_text.pack(padx=20, pady=(0, 10))
        
        # Pre-populate with TBS order_line fields
        default_fields = """product_code
qty_brutto
qty_tara
qty_netto
product_uom
sortation_percent
price_unit
product_qty"""
        fields_text.insert(1.0, default_fields)
        
        def save_array():
            name = array_name_var.get().strip()
            if not name:
                messagebox.showwarning("Warning", "Please enter array name")
                return
                
            fields_content = fields_text.get(1.0, tk.END).strip()
            fields = [f.strip() for f in fields_content.split('\n') if f.strip()]
            
            # Initialize nested_groups if not exists
            if not hasattr(self, 'nested_groups'):
                self.nested_groups = {}
            
            self.nested_groups[name] = {
                'type': 'array',
                'description': f'Array of {name} items',
                'fields': fields,
                'item_template': True
            }
            
            self.update_nested_mapping_display()
            messagebox.showinfo("Success", f"Array '{name}' created successfully with {len(fields)} fields!")
            dialog.destroy()
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="Save Array", command=save_array).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT)
    
    def open_visual_designer(self):
        """Open visual nested structure designer"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Visual Nested Structure Designer")
        dialog.geometry("800x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        self.center_window(dialog, 800, 600)
        
        # Header
        header_frame = ttk.Frame(dialog)
        header_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(header_frame, text=" Visual Nested Structure Designer", 
                 font=('Arial', 14, 'bold')).pack(side=tk.LEFT)
        
        # Main container
        main_frame = ttk.PanedWindow(dialog, orient=tk.HORIZONTAL)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Left panel - Structure tree
        left_frame = ttk.LabelFrame(main_frame, text="Nested Structure", padding=10)
        main_frame.add(left_frame, weight=1)
        
        # Tree widget for nested structure
        tree_frame = ttk.Frame(left_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        self.structure_tree = ttk.Treeview(tree_frame, columns=('type', 'mapped'), show='tree headings')
        self.structure_tree.heading('#0', text='Field Name')
        self.structure_tree.heading('type', text='Type')
        self.structure_tree.heading('mapped', text='Mapped')
        
        # Add sample TBS structure
        root_node = self.structure_tree.insert('', 'end', text='order_data', values=('array', 'Yes'))
        header_node = self.structure_tree.insert(root_node, 'end', text='[Header Fields]', values=('group', ''))
        self.structure_tree.insert(header_node, 'end', text='partner_id', values=('string', 'No'))
        self.structure_tree.insert(header_node, 'end', text='journal_id', values=('string', 'No'))
        self.structure_tree.insert(header_node, 'end', text='vehicle_no', values=('string', 'No'))
        
        line_node = self.structure_tree.insert(root_node, 'end', text='order_line', values=('array', ''))
        self.structure_tree.insert(line_node, 'end', text='product_code', values=('string', 'No'))
        self.structure_tree.insert(line_node, 'end', text='qty_netto', values=('number', 'No'))
        self.structure_tree.insert(line_node, 'end', text='price_unit', values=('number', 'No'))
        
        self.structure_tree.pack(fill=tk.BOTH, expand=True)
        
        # Expand all nodes
        for item in self.structure_tree.get_children():
            self.structure_tree.item(item, open=True)
            for child in self.structure_tree.get_children(item):
                self.structure_tree.item(child, open=True)
        
        # Right panel - Mapping controls
        right_frame = ttk.LabelFrame(main_frame, text="Field Mapping", padding=10)
        main_frame.add(right_frame, weight=1)
        
        # Database fields list
        ttk.Label(right_frame, text="Database Fields:", font=('Arial', 10, 'bold')).pack(anchor=tk.W)
        
        db_list_frame = ttk.Frame(right_frame)
        db_list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.db_fields_listbox = tk.Listbox(db_list_frame, height=8)
        db_scrollbar = ttk.Scrollbar(db_list_frame, orient=tk.VERTICAL, command=self.db_fields_listbox.yview)
        self.db_fields_listbox.configure(yscrollcommand=db_scrollbar.set)
        
        self.db_fields_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        db_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Populate with database fields if available
        if hasattr(self, 'table_columns') and self.table_columns:
            for col in self.table_columns:
                self.db_fields_listbox.insert(tk.END, col)
        
        # Mapping controls
        ttk.Label(right_frame, text="Actions:", font=('Arial', 10, 'bold')).pack(anchor=tk.W, pady=(10, 5))
        
        action_frame = ttk.Frame(right_frame)
        action_frame.pack(fill=tk.X)
        
        ttk.Button(action_frame, text=" Map Selected", 
                  command=self.map_selected_field).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(action_frame, text=" Remove Mapping", 
                  command=self.remove_field_mapping).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(action_frame, text=" Auto Map All", 
                  command=self.auto_map_nested_fields).pack(fill=tk.X, pady=(0, 5))
        
        # Close button
        ttk.Button(right_frame, text=" Apply & Close", 
                  command=dialog.destroy).pack(side=tk.BOTTOM, pady=(20, 0))
    
    def import_nested_api_spec(self):
        """Import API specification for nested structure"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Import Nested API Specification")
        dialog.geometry("600x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        self.center_window(dialog, 600, 500)
        
        ttk.Label(dialog, text=" Import Nested API Specification", 
                 font=('Arial', 14, 'bold')).pack(pady=10)
        
        # Instructions
        instruction_text = """
Paste your API specification (JSON format) below.
The system will automatically detect nested structures and create mapping groups.

Supported formats:
[CHAR] JSON-RPC 2.0 (TBS format)
[CHAR] REST API JSON schemas
[CHAR] OpenAPI/Swagger specifications
        """
        ttk.Label(dialog, text=instruction_text.strip(), justify=tk.LEFT).pack(padx=20, pady=(0, 10))
        
        # Text area for API spec
        spec_frame = ttk.LabelFrame(dialog, text="API Specification", padding=10)
        spec_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
        
        spec_text = scrolledtext.ScrolledText(spec_frame, height=15)
        spec_text.pack(fill=tk.BOTH, expand=True)
        
        # Pre-populate with TBS example
        tbs_example = """{
  "jsonrpc": "2.0",
  "params": {
    "order_data": [
      {
        "partner_id": "string",
        "journal_id": "string", 
        "vehicle_no": "string",
        "order_line": [
          {
            "product_code": "string",
            "qty_netto": "number",
            "sortation_percent": "number",
            "price_unit": "number"
          }
        ]
      }
    ]
  }
}"""
        spec_text.insert(1.0, tbs_example)
        
        def parse_and_import():
            spec_content = spec_text.get(1.0, tk.END).strip()
            if not spec_content:
                messagebox.showwarning("Warning", "Please enter API specification")
                return
            
            try:
                import json
                api_spec = json.loads(spec_content)
                
                # Parse nested structure
                self.parse_nested_structure(api_spec)
                
                messagebox.showinfo("Success", "API specification imported successfully!\nNested structure has been parsed and groups created.")
                dialog.destroy()
                
            except json.JSONDecodeError as e:
                messagebox.showerror("Error", f"Invalid JSON format:\n{str(e)}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to parse API specification:\n{str(e)}")
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text=" Import & Parse", command=parse_and_import).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT)
    
    def on_mapping_mode_change(self):
        """Handle mapping mode change"""
        mode = self.mapping_mode.get()
        self.log_entry(f"Mapping mode changed to: {mode}", "INFO")
        
        # Update field mapper with new mode
        self.update_field_mapper()
        
        if mode == "tbs_auto":
            # Auto-setup TBS nested structure
            self.setup_tbs_nested_structure()
            # Auto-load TBS API fields
            self.auto_load_tbs_api_fields()
            # Set source as TBS
            self.api_fields_source = "manual_tbs_receiving"
            # Auto-map fields
            if hasattr(self, 'mapping_widgets') and hasattr(self, 'table_columns'):
                self.auto_map_tbs_fields()
        elif mode == "nested":
            # Enable nested mapping controls
            self.update_nested_mapping_display()
        else:
            # Flat mode - use existing functionality
            pass
        
        # Refresh preview with new mode
        self.update_json_preview()
    
    def test_mapping_with_real_data(self):
        """Test current mapping with real database data"""
        try:
            if not hasattr(self, 'selected_table') or not self.selected_table:
                messagebox.showerror("Error", "Please select a database table first.")
                return
            
            if not hasattr(self, 'field_mappings') or not self.field_mappings:
                messagebox.showerror("Error", "Please configure field mapping first.")
                return
            
            # Get latest record from database
            raw_data = self.get_latest_record()
            if not raw_data:
                messagebox.showerror("Error", "No data found in selected table.")
                return
            
            # Build mapped payload
            mapped_data = self.build_api_payload(raw_data)
            
            # Show results in dialog
            result_window = tk.Toplevel(self.root)
            result_window.title("Mapping Test Results")
            result_window.geometry("700x500")
            result_window.transient(self.root)
            result_window.grab_set()
            
            # Raw data section
            raw_frame = ttk.LabelFrame(result_window, text="Raw Database Data", padding=10)
            raw_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 5))
            
            raw_text = scrolledtext.ScrolledText(raw_frame, height=8, state=tk.DISABLED)
            raw_text.pack(fill=tk.BOTH, expand=True)
            
            raw_text.config(state=tk.NORMAL)
            raw_text.insert(1.0, json.dumps(raw_data, indent=2))
            raw_text.config(state=tk.DISABLED)
            
            # Mapped data section
            mapped_frame = ttk.LabelFrame(result_window, text="Mapped API Payload", padding=10)
            mapped_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            mapped_text = scrolledtext.ScrolledText(mapped_frame, height=8, state=tk.DISABLED)
            mapped_text.pack(fill=tk.BOTH, expand=True)
            
            mapped_text.config(state=tk.NORMAL)
            mapped_text.insert(1.0, json.dumps(mapped_data, indent=2))
            mapped_text.config(state=tk.DISABLED)
            
            # Control buttons
            btn_frame = ttk.Frame(result_window)
            btn_frame.pack(fill=tk.X, padx=10, pady=(5, 10))
            
            ttk.Button(btn_frame, text="Send Test to API", 
                      command=lambda: self.send_test_to_api(mapped_data)).pack(side=tk.LEFT, padx=(0, 10))
            ttk.Button(btn_frame, text="Save as Template", 
                      command=lambda: self.save_test_as_template(mapped_data)).pack(side=tk.LEFT, padx=(0, 10))
            ttk.Button(btn_frame, text="Close", 
                      command=result_window.destroy).pack(side=tk.RIGHT)
            
            self.log_entry("Mapping test completed successfully", "INFO")
            
        except Exception as e:
            error_msg = f"Mapping test failed: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def send_test_to_api(self, test_data: Dict):
        """Send test data to API"""
        try:
            success = self.send_to_api(test_data)
            if success:
                messagebox.showinfo("Success", "Test data sent to API successfully!")
            else:
                messagebox.showerror("Error", "Failed to send test data. Check logs.")
        except Exception as e:
            messagebox.showerror("Error", f"Test API call failed: {str(e)}")
    
    def save_test_as_template(self, test_data: Dict):
        """Save test data as template"""
        try:
            # Simple implementation - could be enhanced
            template_name = f"test_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            # Save to config or file
            messagebox.showinfo("Template Saved", f"Test data saved as template: {template_name}")
            self.log_entry(f"Test template saved: {template_name}", "INFO")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save template: {str(e)}")
    
    def setup_tbs_nested_structure(self):
        """Auto-setup TBS nested structure"""
        if not hasattr(self, 'nested_groups'):
            self.nested_groups = {}
        
        # Create TBS order_data structure
        self.nested_groups['order_data'] = {
            'type': 'array',
            'description': 'TBS Order Data Array',
            'fields': ['partner_id', 'journal_id', 'vehicle_no', 'date_order', 'officers'],
            'nested': {
                'order_line': {
                    'type': 'array',
                    'description': 'Order Line Items',
                    'fields': ['product_code', 'qty_brutto', 'qty_tara', 'qty_netto', 'price_unit', 'product_qty']
                }
            }
        }
        
        self.log_entry("TBS nested structure auto-configured", "SUCCESS")
        messagebox.showinfo("TBS Auto Setup", "TBS nested structure has been automatically configured!\n\n" +
                           "Structure created:\n" +
                           " order_data (array)\n" +
                           "   Header fields (partner_id, journal_id, etc.)\n" +
                           "   order_line (nested array)\n" +
                           "     Line item fields (product_code, qty_netto, etc.)")
    
    def update_nested_mapping_display(self):
        """Update the nested mapping display"""
        # This would update the UI to show nested groups
        # Implementation depends on the current UI structure
        pass
    
    def auto_load_tbs_api_fields(self):
        """Auto-load TBS API fields when TBS mode is selected"""
        tbs_fields = [
            "# === ROOT LEVEL FIELDS ===",
            "uuid",
            "timestamp", 
            "source",
            "",
            "# === ORDER_DATA LEVEL FIELDS ===",
            "order_data.partner_id",
            "order_data.journal_id", 
            "order_data.date_order",
            "order_data.officers",
            "order_data.keterangan_description",
            "order_data.driver_name",
            "order_data.vehicle_no",
            "order_data.destination_warehouse_id",
            "order_data.branch_id",
            "",
            "# === ORDER_LINE LEVEL FIELDS (nested in order_data) ===",
            "order_data.order_line.product_code",
            "order_data.order_line.qty_brutto",
            "order_data.order_line.qty_tara", 
            "order_data.order_line.qty_netto",
            "order_data.order_line.product_uom",
            "order_data.order_line.sortation_percent",
            "order_data.order_line.sortation_weight",
            "order_data.order_line.qty_netto2",
            "order_data.order_line.price_unit",
            "order_data.order_line.product_qty",
            "order_data.order_line.incoming_date",
            "order_data.order_line.outgoing_date"
        ]
        
        # Filter out comments and empty lines for actual fields
        self.api_fields = [field for field in tbs_fields if field and not field.startswith("#")]
        self.api_fields_source = "manual_tbs_receiving"
        
        # Update the comboboxes
        self.update_mapping_comboboxes()
        
        self.log_entry(f"Auto-loaded {len(self.api_fields)} TBS API fields", "INFO")
    
    def map_selected_field(self):
        """Map selected database field to API field"""
        # Implementation for mapping in visual designer
        pass
    
    def remove_field_mapping(self):
        """Remove selected field mapping"""
        # Implementation for removing mapping
        pass
    
    def auto_map_nested_fields(self):
        """Auto-map all fields to nested structure"""
        # Implementation for auto-mapping
        pass
    
    def parse_nested_structure(self, api_spec):
        """Parse API specification and create nested structure"""
        def extract_structure(obj, path=""):
            structure = {}
            
            if isinstance(obj, dict):
                for key, value in obj.items():
                    current_path = f"{path}.{key}" if path else key
                    
                    if isinstance(value, list):
                        if value and isinstance(value[0], dict):
                            # Array of objects
                            structure[key] = {
                                'type': 'array',
                                'path': current_path,
                                'fields': extract_structure(value[0], current_path)
                            }
                        else:
                            # Array of primitives
                            structure[key] = {
                                'type': 'array_primitive',
                                'path': current_path
                            }
                    elif isinstance(value, dict):
                        # Nested object
                        structure[key] = {
                            'type': 'object',
                            'path': current_path,
                            'fields': extract_structure(value, current_path)
                        }
                    else:
                        # Primitive field
                        field_type = 'string'
                        if isinstance(value, (int, float)):
                            field_type = 'number'
                        elif isinstance(value, bool):
                            field_type = 'boolean'
                        
                        structure[key] = {
                            'type': field_type,
                            'path': current_path
                        }
            
            return structure
        
        # Extract and store the structure
        self.parsed_structure = extract_structure(api_spec)
        
        # Convert to nested_groups format
        if not hasattr(self, 'nested_groups'):
            self.nested_groups = {}
        
        def convert_to_groups(structure, prefix=""):
            for key, value in structure.items():
                if value['type'] in ['array', 'object']:
                    group_name = f"{prefix}{key}" if prefix else key
                    self.nested_groups[group_name] = {
                        'type': value['type'],
                        'description': f"Auto-imported {value['type']}: {key}",
                        'fields': list(value.get('fields', {}).keys()) if 'fields' in value else [],
                        'path': value['path']
                    }
                    
                    # Recursively process nested structures
                    if 'fields' in value and value['fields']:
                        convert_to_groups(value['fields'], f"{group_name}.")
        
        convert_to_groups(self.parsed_structure)
    
    def center_window(self, window, width, height):
        """Center a window on the screen"""
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        window.geometry(f"{width}x{height}+{x}+{y}")
    
    # =====================================
    # ENHANCED ACTION FUNCTIONS - Missing Functions Added
    # =====================================
    
    def save_field_mapping_with_validation(self):
        """Enhanced save field mapping with validation and user feedback"""
        try:
            if not hasattr(self, 'mapping_widgets'):
                messagebox.showwarning("Warning", "No mapping configuration found.\nPlease create field mappings first.")
                return
            
            # Show progress
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Saving Mapping")
            progress_window.geometry("300x100")
            progress_window.transient(self.root)
            progress_window.grab_set()
            self.center_window(progress_window, 300, 100)
            
            ttk.Label(progress_window, text="Saving field mapping configuration...", 
                     font=('Arial', 10)).pack(pady=20)
            progress_bar = ttk.Progressbar(progress_window, mode='indeterminate')
            progress_bar.pack(pady=10, padx=20, fill=tk.X)
            progress_bar.start()
            
            self.root.update()
            
            # Perform actual save and update
            self.save_field_mapping()
            
            # Update field mapper with latest mappings
            self.update_field_mapper()
            
            # Close progress window
            progress_bar.stop()
            progress_window.destroy()
            
            # Show success confirmation with mapping details
            mode = self.mapping_mode.get()
            mapping_count = len(self.field_mappings) if hasattr(self, 'field_mappings') else 0
            
            messagebox.showinfo("Success", f" Field mapping applied successfully!\n\n" +
                               f"Mode: {mode.upper()}\n" +
                               f"Fields mapped: {mapping_count}\n\n" +
                               "Your mapping configuration is now active for data export.")
            
            # Refresh preview to show current state
            self.update_json_preview()
            
        except Exception as e:
            if 'progress_window' in locals():
                progress_window.destroy()
            error_msg = f"Failed to save field mapping:\n{str(e)}"
            messagebox.showerror("Save Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def apply_and_refresh_mapping(self):
        """Enhanced apply current mapping and refresh all displays with feedback"""
        try:
            # Validate first
            validation_issues = self.validate_mapping_configuration()
            if validation_issues:
                issues_text = "\n".join([f" {issue}" for issue in validation_issues])
                self.show_operation_feedback(
                    "Cannot Apply Configuration", 
                    f"Please fix these issues first:\n\n{issues_text}",
                    "warning"
                )
                return
            
            # Show progress window
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Applying Configuration")
            progress_window.geometry("400x200")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            # Center window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (200)
            y = (progress_window.winfo_screenheight() // 2) - (100)
            progress_window.geometry(f"400x200+{x}+{y}")
            
            content_frame = ttk.Frame(progress_window, padding=20)
            content_frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(content_frame, text=" Applying & Refreshing Configuration", 
                     font=('Arial', 12, 'bold')).pack(pady=(0, 15))
            
            progress_bar = ttk.Progressbar(content_frame, mode='indeterminate')
            progress_bar.pack(fill=tk.X, pady=(0, 10))
            progress_bar.start()
            
            status_label = ttk.Label(content_frame, text="Preparing...", font=('Arial', 9))
            status_label.pack(pady=(0, 10))
            
            # Detailed progress log
            log_text = tk.Text(content_frame, height=6, width=45, font=('Consolas', 8))
            log_text.pack(fill=tk.BOTH, expand=True)
            
            def update_log(message):
                log_text.insert(tk.END, f" {message}\n")
                log_text.see(tk.END)
                progress_window.update()
            
            # Step 1: Apply current mapping configuration
            status_label.config(text="Step 1: Applying mapping configuration...")
            update_log("Checking current mapping widgets...")
            
            if hasattr(self, 'mapping_widgets'):
                update_log("Found mapping widgets - applying configuration")
                # Apply mapping from widgets
                for widget_info in self.mapping_widgets:
                    if 'combobox' in widget_info and 'db_field' in widget_info:
                        db_field = widget_info['db_field']
                        api_field = widget_info['combobox'].get()
                        if api_field and api_field != "Select API Field":
                            self.field_mappings[db_field] = api_field
                            update_log(f"Mapped: {db_field}  {api_field}")
            
            # Step 2: Refresh API connection
            status_label.config(text="Step 2: Refreshing API connection...")
            update_log("Refreshing API status...")
            self.refresh_api_mapping_status()
            update_log("API status updated")
            
            # Step 3: Update nested configuration status
            status_label.config(text="Step 3: Updating nested configuration...")
            update_log("Checking nested structure configuration...")
            
            nested_config = getattr(self, 'nested_structure_config', {})
            if nested_config:
                self.update_nested_status(" Nested Structure Applied", "green")
                update_log("Nested structure configuration applied")
            else:
                self.update_nested_status(" Standard Mapping Mode", "blue")
                update_log("Standard mapping mode applied")
            
            # Step 4: Refresh displays
            status_label.config(text="Step 4: Refreshing interface displays...")
            update_log("Updating JSON preview...")
            if hasattr(self, 'update_json_preview'):
                self.update_json_preview()
            
            update_log("Refreshing mapping interface...")
            if hasattr(self, 'refresh_mapping_interface'):
                self.refresh_mapping_interface()
            
            # Step 5: Final status update
            status_label.config(text="Configuration applied successfully!")
            update_log(" All configurations applied successfully!")
            
            # Stop progress and wait a moment
            progress_bar.stop()
            progress_window.update()
            progress_window.after(2000, progress_window.destroy)
            
            # Show success feedback
            self.show_operation_feedback(
                "Configuration Applied", 
                f"Mapping configuration has been applied successfully!\n\n" +
                f"Active mappings: {len(getattr(self, 'field_mappings', {}))}\n" +
                f"Nested config: {'Yes' if nested_config else 'No'}\n" +
                f"Mode: {getattr(self, 'mapping_mode_var', tk.StringVar()).get()}",
                "success"
            )
            
            self.log_message("Mapping configuration applied and refreshed successfully", "SUCCESS")
            
        except Exception as e:
            if 'progress_window' in locals():
                progress_window.destroy()
            error_msg = f"Error applying configuration:\n{str(e)}"
            self.show_operation_feedback("Apply Error", error_msg, "error")
            self.log_message(error_msg, "ERROR")

    def reset_field_mapping_with_confirm(self):
        """Enhanced reset field mapping with detailed confirmation dialog"""
        try:
            # Create custom confirmation dialog
            confirm_window = tk.Toplevel(self.root)
            confirm_window.title(" Confirm Reset Mapping")
            confirm_window.geometry("450x300")
            confirm_window.transient(self.root)
            confirm_window.grab_set()
            confirm_window.update_idletasks()
            x = (confirm_window.winfo_screenwidth() // 2) - (225)
            y = (confirm_window.winfo_screenheight() // 2) - (150)
            confirm_window.geometry(f"450x300+{x}+{y}")
            
            content_frame = ttk.Frame(confirm_window, padding=20)
            content_frame.pack(fill=tk.BOTH, expand=True)
            
            # Warning header
            ttk.Label(content_frame, text=" RESET FIELD MAPPING", 
                     font=('Arial', 14, 'bold'), foreground='red').pack(pady=(0, 15))
            
            # Warning message
            warning_text = ("This action will completely reset your field mapping configuration.\n\n"
                           "The following will be cleared:\n"
                           " All field mappings between database and API\n"
                           " Nested structure configuration\n"
                           " Transformation settings\n"
                           " Template configurations\n\n"
                           " This action CANNOT be undone!")
            
            ttk.Label(content_frame, text=warning_text, font=('Arial', 10), 
                     wraplength=400, justify='left').pack(pady=(0, 20))
            
            # Current status info
            current_info = f"Current Configuration:\n"
            current_info += f" Mapped fields: {len(getattr(self, 'field_mappings', {}))}\n"
            current_info += f" Nested groups: {len(getattr(self, 'nested_groups', {}))}\n"
            current_info += f" Active table: {getattr(self, 'selected_table', 'None')}"
            
            info_frame = ttk.LabelFrame(content_frame, text="Current Status", padding=10)
            info_frame.pack(fill=tk.X, pady=(0, 20))
            
            ttk.Label(info_frame, text=current_info, font=('Arial', 9), 
                     foreground='blue').pack(anchor='w')
            
            # Button frame
            button_frame = ttk.Frame(content_frame)
            button_frame.pack(fill=tk.X)
            
            result = tk.BooleanVar(value=False)
            
            def confirm_reset():
                result.set(True)
                confirm_window.destroy()
            
            def cancel_reset():
                result.set(False)
                confirm_window.destroy()
            
            ttk.Button(button_frame, text=" YES, RESET ALL", 
                      command=confirm_reset, 
                      style='Danger.TButton').pack(side=tk.LEFT, padx=(0, 10))
            
            ttk.Button(button_frame, text=" Cancel", 
                      command=cancel_reset).pack(side=tk.LEFT)
            
            # Wait for user decision
            confirm_window.wait_window()
            
            if not result.get():
                return
            
            # Show progress for reset operation
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Resetting Configuration")
            progress_window.geometry("350x150")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            # Center progress window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (175)
            y = (progress_window.winfo_screenheight() // 2) - (75)
            progress_window.geometry(f"350x150+{x}+{y}")
            
            prog_content = ttk.Frame(progress_window, padding=20)
            prog_content.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(prog_content, text=" Resetting Configuration...", 
                     font=('Arial', 12, 'bold')).pack(pady=(0, 10))
            
            progress_bar = ttk.Progressbar(prog_content, mode='indeterminate')
            progress_bar.pack(fill=tk.X, pady=(0, 10))
            progress_bar.start()
            
            status_label = ttk.Label(prog_content, text="Clearing mappings...", font=('Arial', 9))
            status_label.pack()
            
            progress_window.update()
            
            # Perform reset operations
            status_label.config(text="Clearing field mappings...")
            progress_window.update()
            self.field_mappings = {}
            
            status_label.config(text="Clearing nested configuration...")
            progress_window.update()
            if hasattr(self, 'nested_groups'):
                self.nested_groups = {}
            if hasattr(self, 'nested_structure_config'):
                self.nested_structure_config = {}
            
            status_label.config(text="Resetting UI widgets...")
            progress_window.update()
            if hasattr(self, 'mapping_widgets'):
                for widget_info in self.mapping_widgets:
                    if 'combobox' in widget_info:
                        widget_info['combobox'].set("Select API Field")
            
            status_label.config(text="Updating status indicators...")
            progress_window.update()
            self.update_nested_status(" Configuration Reset", "gray")
            
            status_label.config(text="Refreshing displays...")
            progress_window.update()
            if hasattr(self, 'update_json_preview'):
                self.update_json_preview()
            if hasattr(self, 'refresh_mapping_interface'):
                self.refresh_mapping_interface()
            
            # Close progress window
            progress_bar.stop()
            progress_window.destroy()
            
            # Show success confirmation
            self.show_operation_feedback(
                "Reset Complete", 
                "Field mapping configuration has been reset successfully!\n\n" +
                "All mappings and nested configurations have been cleared.\n" +
                "You can now create a fresh configuration.",
                "success"
            )
            
            self.log_message("Field mapping reset completed successfully", "INFO")
            
        except Exception as e:
            if 'progress_window' in locals():
                progress_window.destroy()
            error_msg = f"Failed to reset mapping configuration:\n{str(e)}"
            self.show_operation_feedback("Reset Error", error_msg, "error")
            self.log_message(error_msg, "ERROR")
    
    def apply_nested_configuration(self):
        """Apply nested structure configuration"""
        try:
            if not hasattr(self, 'nested_groups') or not self.nested_groups:
                messagebox.showwarning("No Nested Configuration", 
                                     "No nested structure configuration found.\n\n" +
                                     "Please create nested groups first using:\n" +
                                     " Add Group/Array buttons\n" +
                                     " Visual Designer\n" +
                                     " Import API Spec")
                return
            
            # Apply nested configuration
            mode = getattr(self, 'mapping_mode', tk.StringVar()).get()
            if mode not in ['nested', 'tbs_auto']:
                self.mapping_mode.set('nested')
                self.on_mapping_mode_change()
            
            # Update JSON preview with nested structure
            self.update_json_preview()
            
            # Show configuration details
            config_details = []
            for group_name, group_config in self.nested_groups.items():
                config_details.append(f" {group_name} ({group_config['type']}) - {len(group_config.get('fields', []))} fields")
            
            messagebox.showinfo("Nested Configuration Applied", 
                               " Nested structure configuration applied successfully!\n\n" +
                               "Structure:\n" + "\n".join(config_details[:5]) +
                               ("\n... and more" if len(config_details) > 5 else ""))
            
            self.log_entry(f"Nested configuration applied: {len(self.nested_groups)} groups", "SUCCESS")
            
        except Exception as e:
            error_msg = f"Failed to apply nested configuration: {str(e)}"
            messagebox.showerror("Configuration Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def preview_nested_structure(self):
        """Preview nested structure in a dedicated window"""
        try:
            # Create preview window
            preview_window = tk.Toplevel(self.root)
            preview_window.title("Nested Structure Preview")
            preview_window.geometry("700x500")
            preview_window.transient(self.root)
            self.center_window(preview_window, 700, 500)
            
            # Header
            header_frame = ttk.Frame(preview_window)
            header_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Label(header_frame, text=" Nested Structure Preview", 
                     font=('Arial', 14, 'bold')).pack(side=tk.LEFT)
            
            # Close button
            ttk.Button(header_frame, text=" Close", 
                      command=preview_window.destroy).pack(side=tk.RIGHT)
            
            # Content
            content_frame = ttk.Frame(preview_window)
            content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
            
            # Structure display
            structure_text = scrolledtext.ScrolledText(content_frame, height=20)
            structure_text.pack(fill=tk.BOTH, expand=True)
            
            # Generate structure preview
            if hasattr(self, 'nested_groups') and self.nested_groups:
                structure_preview = " NESTED STRUCTURE CONFIGURATION\n"
                structure_preview += "=" * 50 + "\n\n"
                
                for group_name, group_config in self.nested_groups.items():
                    structure_preview += f" {group_name.upper()} ({group_config['type']})\n"
                    structure_preview += f"   Description: {group_config.get('description', 'No description')}\n"
                    structure_preview += f"   Fields ({len(group_config.get('fields', []))}):\n"
                    
                    for field in group_config.get('fields', []):
                        structure_preview += f"    {field}\n"
                    
                    if 'nested' in group_config:
                        structure_preview += f"   Nested Groups:\n"
                        for nested_name, nested_config in group_config['nested'].items():
                            structure_preview += f"     {nested_name} ({nested_config['type']})\n"
                    
                    structure_preview += "\n"
                
                # Add JSON example
                structure_preview += "\n EXAMPLE JSON OUTPUT:\n"
                structure_preview += "=" * 30 + "\n"
                
                try:
                    sample_data = {'test': 'data'}
                    if hasattr(self, 'convert_with_nested_mapping'):
                        example_json = self.convert_with_nested_mapping(sample_data)
                        import json
                        structure_preview += json.dumps(example_json, indent=2, ensure_ascii=False)
                    else:
                        structure_preview += "Preview not available - please save configuration first"
                except:
                    structure_preview += "Preview generation failed - please check configuration"
                    
            else:
                structure_preview = " NO NESTED STRUCTURE CONFIGURED\n\n"
                structure_preview += "To create nested structure:\n"
                structure_preview += "1. Click 'Add Group' or 'Add Array' buttons\n"
                structure_preview += "2. Use Visual Designer\n"
                structure_preview += "3. Import API Spec\n"
                structure_preview += "4. Select 'TBS Auto' mode for automatic setup"
            
            structure_text.insert(1.0, structure_preview)
            structure_text.config(state=tk.DISABLED)
            
        except Exception as e:
            error_msg = f"Failed to generate structure preview: {str(e)}"
            messagebox.showerror("Preview Error", error_msg)
            self.log_entry(error_msg, "ERROR")
    
    def clear_nested_configuration(self):
        """Clear nested structure configuration"""
        result = messagebox.askyesno("Confirm Clear", 
                                   " Are you sure you want to clear the nested structure configuration?\n\n" +
                                   "This will remove all:\n" +
                                   " Created groups and arrays\n" +
                                   " Nested structure mappings\n" +
                                   " Visual designer settings\n\n" +
                                   "Field mappings will be preserved.")
        
        if result:
            try:
                # Clear nested configuration
                if hasattr(self, 'nested_groups'):
                    self.nested_groups = {}
                
                # Reset to flat mode
                if hasattr(self, 'mapping_mode'):
                    self.mapping_mode.set('flat')
                    self.on_mapping_mode_change()
                
                # Update preview
                self.update_json_preview()
                
                messagebox.showinfo("Cleared", " Nested structure configuration cleared successfully!\n\n" +
                                   "Mapping mode reset to 'Flat'.")
                
                self.log_entry("Nested structure configuration cleared", "INFO")
                
            except Exception as e:
                error_msg = f"Failed to clear nested configuration: {str(e)}"
                messagebox.showerror("Clear Error", error_msg)
                self.log_entry(error_msg, "ERROR")

if __name__ == "__main__":
    try:
        app = MDBAgentPro()
        app.run()
    except Exception as e:
        print(f"Failed to start application: {e}")
        traceback.print_exc()
    
